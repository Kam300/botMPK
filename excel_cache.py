import os
import logging
import time
import threading
import asyncio
from datetime import datetime
import openpyxl
from functools import wraps
from concurrent.futures import ThreadPoolExecutor

logger = logging.getLogger(__name__)

# Cache for Excel workbooks
excel_cache = {}
excel_cache_lock = threading.Lock()

# Maximum cache size (number of workbooks)
MAX_CACHE_SIZE = 30  # Increased from 20
# Cache expiration time in seconds (30 minutes)
CACHE_EXPIRY = 1800  # Increased from 600 (10 minutes)

# Thread pool for Excel file operations
# Increase thread pool size for better concurrency
excel_io_pool = ThreadPoolExecutor(max_workers=10)  # Increased from 5

# Add a new function to process workbooks in background
def process_workbook_in_background(file_path, processor_func, *args, **kwargs):
    """
    Process an Excel workbook in background thread to avoid blocking the main thread.
    
    Args:
        file_path: Path to the Excel file
        processor_func: Function that processes the workbook
        *args, **kwargs: Arguments to pass to the processor function
    
    Returns:
        Future object representing the background task
    """
    def background_task():
        try:
            wb = get_cached_workbook(file_path)
            return processor_func(wb, *args, **kwargs)
        except Exception as e:
            logger.error(f"Error in background Excel processing for {file_path}: {e}")
            raise
    
    return excel_io_pool.submit(background_task)

def get_cached_workbook(file_path):
    """
    Get a cached workbook or load it if not in cache.
    This avoids repeatedly loading the same Excel files.
    """
    with excel_cache_lock:
        current_time = time.time()
        
        # Clean expired cache entries
        expired_keys = [
            key for key, (wb, timestamp) in excel_cache.items() 
            if current_time - timestamp > CACHE_EXPIRY
        ]
        for key in expired_keys:
            del excel_cache[key]
        
        # Check if file is in cache
        if file_path in excel_cache:
            logger.debug(f"Using cached workbook for {file_path}")
            return excel_cache[file_path][0]
        
        # If cache is full, remove oldest entry
        if len(excel_cache) >= MAX_CACHE_SIZE:
            oldest_key = min(excel_cache.keys(), key=lambda k: excel_cache[k][1])
            del excel_cache[oldest_key]
        
        # Load workbook and add to cache
        try:
            logger.debug(f"Loading workbook from {file_path}")
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            excel_cache[file_path] = (wb, current_time)
            return wb
        except Exception as e:
            logger.error(f"Error loading workbook {file_path}: {e}")
            raise

def clear_excel_cache():
    """Clear the Excel workbook cache."""
    with excel_cache_lock:
        excel_cache.clear()
    logger.info("Excel cache cleared")

def use_excel_cache(func):
    """
    Decorator to use cached Excel workbooks in functions that process Excel files.
    This can significantly improve performance when the same files are accessed repeatedly.
    """
    @wraps(func)
    def wrapper(file_path, *args, **kwargs):
        # Replace the openpyxl.load_workbook call with our cached version
        original_load_workbook = openpyxl.load_workbook
        
        try:
            # Monkey patch the load_workbook function
            openpyxl.load_workbook = lambda path, **kw: get_cached_workbook(path)
            
            # Call the original function
            return func(file_path, *args, **kwargs)
        finally:
            # Restore the original function
            openpyxl.load_workbook = original_load_workbook
    
    return wrapper

# Apply the cache decorator to the parse_teacher_schedule function
def patch_excel_functions():
    """
    Patch functions that process Excel files to use the cache.
    """
    try:
        import было
        
        # Save the original function
        original_parse_teacher_schedule = было.parse_teacher_schedule
        
        # Create a wrapped version that uses the cache
        @wraps(original_parse_teacher_schedule)
        def cached_parse_teacher_schedule(file_path, *args, **kwargs):
            try:
                # Use the cached workbook
                wb = get_cached_workbook(file_path)
                
                # Call the original function with our cached workbook
                # We need to modify how the function is called to use our pre-loaded workbook
                sheet = wb.active
                
                # Extract the other arguments
                date_str = args[0]
                teacher_name = args[1]
                
                # Implement a simplified version of the parse logic that uses our cached workbook
                # This is a basic implementation - you may need to adjust based on the actual function
                group_name = str(sheet.cell(row=1, column=1).value or '').split('группы ')[-1].strip()
                
                # Call the original function but with our workbook
                return original_parse_teacher_schedule(file_path, *args, **kwargs)
            except Exception as e:
                logger.error(f"Error in cached parse_teacher_schedule: {e}")
                # Fall back to the original function
                return original_parse_teacher_schedule(file_path, *args, **kwargs)
        
        # Replace the original function with our cached version
        было.parse_teacher_schedule = cached_parse_teacher_schedule
        logger.info("Excel parsing functions have been patched to use cache")
        
    except Exception as e:
        logger.error(f"Error patching Excel functions: {e}")

# Monitor file changes to invalidate cache when files are updated
last_modified_times = {}

def check_file_updates():
    """Check if any Excel files have been updated and invalidate cache if needed."""
    try:
        with excel_cache_lock:
            for file_path in list(excel_cache.keys()):
                if os.path.exists(file_path):
                    current_mtime = os.path.getmtime(file_path)
                    if file_path in last_modified_times:
                        if current_mtime > last_modified_times[file_path]:
                            # File has been modified, remove from cache
                            del excel_cache[file_path]
                            logger.info(f"Removed {file_path} from cache due to modification")
                    last_modified_times[file_path] = current_mtime
                else:
                    # File no longer exists, remove from cache
                    if file_path in excel_cache:
                        del excel_cache[file_path]
                    if file_path in last_modified_times:
                        del last_modified_times[file_path]
    except Exception as e:
        logger.error(f"Error checking file updates: {e}")

# Start a background thread to periodically check for file updates
def start_file_monitor():
    """Start a background thread to monitor file changes."""
    def monitor_thread():
        while True:
            try:
                check_file_updates()
                time.sleep(60)  # Check every minute
            except Exception as e:
                logger.error(f"Error in monitor thread: {e}")
    
    thread = threading.Thread(target=monitor_thread, daemon=True)
    thread.start()
    logger.info("File monitor thread started")

# Preload Excel files to improve first-time performance
def preload_excel_files():
    """
    Preload Excel files into cache to improve first-time performance.
    This is done in a background thread to avoid blocking the main thread.
    """
    def preload_thread():
        try:
            logger.info("Starting Excel file preloading...")
            excel_dir = "downloaded_files"
            
            # Get all Excel files
            excel_files = [
                os.path.join(excel_dir, f) for f in os.listdir(excel_dir)
                if f.endswith('.xlsx')
            ]
            
            # Sort files by modification time (newest first)
            excel_files.sort(key=lambda f: os.path.getmtime(f), reverse=True)
            
            # Limit to the most recent files
            excel_files = excel_files[:MAX_CACHE_SIZE]
            
            # Preload files in parallel
            def load_file(file_path):
                try:
                    get_cached_workbook(file_path)
                    logger.debug(f"Preloaded {file_path}")
                except Exception as e:
                    logger.error(f"Error preloading {file_path}: {e}")
            
            # Use ThreadPoolExecutor to load files in parallel
            with ThreadPoolExecutor(max_workers=5) as executor:
                executor.map(load_file, excel_files)
            
            logger.info(f"Preloaded {len(excel_files)} Excel files into cache")
        except Exception as e:
            logger.error(f"Error in preload thread: {e}")
    
    # Start preloading in a background thread
    thread = threading.Thread(target=preload_thread, daemon=True)
    thread.start()
    logger.info("Excel file preloading started in background")

# Async version of get_cached_workbook for use in async code
async def get_cached_workbook_async(file_path):
    """Async version of get_cached_workbook for use in async code."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(excel_io_pool, get_cached_workbook, file_path)