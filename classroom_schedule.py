import os
import runpy
import openpyxl
from log import logger

async def process_file(file):
    try:
        group_name = os.path.splitext(os.path.basename(file))[0]
        logger.info(f"Processing file for group: {group_name}")
        
        async with file_access_semaphore:
            wb = await run_blocking(lambda: openpyxl.load_workbook(
                os.path.join("downloaded_files", file), 
                read_only=True, 
                data_only=True
            ))
            sheet = wb.active
        
        # Find the schedule starting point
        start_row = 1
        for row in range(1, 30):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value and "расписание" in str(cell_value).lower():
                start_row = row + 1
                break
        
        # Find the day column
        is_even_week = week_type == "четная"
        start_row, day_col = find_day_column(sheet, weekday_ru, start_row, is_even_week)
        
        if not start_row or not day_col:
            logger.warning(f"Day {weekday_ru} not found in {file}")
            return []
        
        logger.info(f"Found start row: {start_row}, day column: {day_col}")
        
        # Process lessons
        results = []
        current_row = start_row + 1  # Skip header row
        current_start_row = current_row
        
        while current_row < current_start_row + 20:
            try:
                lesson_num = sheet.cell(row=current_row, column=day_col).value
                if not lesson_num:
                    current_row += 1
                    continue
                    
                # Check both subgroups for the classroom
                room_first = str(sheet.cell(row=current_row + 1, column=day_col + 2).value or '')
                room_second = str(sheet.cell(row=current_row + 1, column=day_col + 4).value or '')
                
                # Initialize variables to avoid potential reference errors
                subject_first = None
                subject_second = None
                teacher_first = None
                teacher_second = None
                
                # Get subject and teacher information
                subject_first = sheet.cell(row=current_row, column=day_col + 1).value
                teacher_first = sheet.cell(row=current_row + 1, column=day_col + 1).value
                subject_second = sheet.cell(row=current_row, column=day_col + 3).value
                teacher_second = sheet.cell(row=current_row + 1, column=day_col + 3).value
                
                # Check if this is a common class (like КП)
                is_common_class = False
                if subject_first and '(КП)' in str(subject_first):
                    is_common_class = True
                    # For common КП classes, check both room cells
                    if classroom == room_first.strip() or classroom == room_second.strip():
                        lesson_key = f"{lesson_num}_{group_name}"
                        results.append({
                            'lesson_num': lesson_num,
                            'lesson_key': lesson_key,
                            'subject': subject_first,
                            'teacher': teacher_first,
                            'room': classroom,
                            'group': group_name,
                            'subgroup': None,  # Common class has no subgroup
                            'is_common': True
                        })
                        logger.info(f"Found common КП class in {classroom} for {group_name}, period {lesson_num}")
                
                # If not a common class, check individual subgroups
                if not is_common_class:
                    # Add a lesson to the classroom schedule if the classroom exactly matches
                    if classroom == room_first.strip():
                        if subject_first:  # Only add if we have a subject
                            lesson_key = f"{lesson_num}_1_{group_name}" if not is_theory_lesson(subject_first) else f"{lesson_num}_{group_name}"
                            results.append({
                                'lesson_num': lesson_num,
                                'lesson_key': lesson_key,
                                'subject': subject_first,
                                'teacher': teacher_first,
                                'room': room_first.strip(),
                                'group': group_name,
                                'subgroup': 1 if not is_theory_lesson(subject_first) else None,
                                'is_common': False
                            })
                    
                    if classroom == room_second.strip():
                        # Get the subject from the first subgroup if the second subgroup doesn't have one
                        if not subject_second and subject_first:
                            subject_second = subject_first
                            teacher_second = teacher_second or teacher_first  # Use first teacher if second is not set
                        
                        if subject_second:  # Only add if we have a subject
                            lesson_key = f"{lesson_num}_2_{group_name}" if not is_theory_lesson(subject_second) else f"{lesson_num}_{group_name}"
                            results.append({
                                'lesson_num': lesson_num,
                                'lesson_key': lesson_key,
                                'subject': subject_second,
                                'teacher': teacher_second,
                                'room': room_second.strip(),
                                'group': group_name,
                                'subgroup': 2 if not is_theory_lesson(subject_second) else None,
                                'is_common': False
                            })
                
                current_row += 2
            except Exception as e:
                logger.error(f"Error processing row in file {file}: {e}")
                current_row += 1  # Continue with the next row
        
        return results
    except Exception as e:
        logger.error(f"Error processing file {file}: {e}")
        return [] 