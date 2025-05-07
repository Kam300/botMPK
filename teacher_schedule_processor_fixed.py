import os
import logging
import asyncio
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
from functools import partial
import openpyxl
import threading
import time
import json
import re
from cache_utils import get_cached_teacher_schedule, cache_teacher_schedule
from excel_cache import use_excel_cache, get_cached_workbook_async

# Import necessary functions from было.py without modifying it
# We'll use these imported functions to maintain compatibility
from было import (
    parse_teacher_schedule,
    format_teacher_schedule,
    run_blocking
)

logger = logging.getLogger(__name__)

# Create a dedicated thread pool for Excel file processing
# This is separate from the main thread_pool in было.py
excel_thread_pool = ThreadPoolExecutor(max_workers=25)  # Increased from 20

# Create a semaphore to limit concurrent file operations
# This prevents overwhelming the system with too many file operations
file_semaphore = asyncio.Semaphore(20)  # Increased from 15

# Lock for thread-safe operations
excel_lock = threading.Lock()

# Track files being processed to avoid duplicate processing
files_being_processed = {}
files_lock = threading.Lock()

# Store file applicability results to avoid recalculating
file_applicability_cache = {}
file_applicability_lock = threading.Lock()

# Popular teachers list for prioritized processing
POPULAR_TEACHERS = []
popular_teachers_lock = threading.Lock()

# Dictionary to track teacher access frequency
teacher_access_counts = {}
access_counts_lock = threading.Lock()

# Background task control
background_processor_running = False
background_processor_lock = threading.Lock()

# Global schedule index for optimized file lookup
schedule_index = {
    "teachers": {},  # ФИО преподавателя -> список файлов
    "dates": {},     # Дата -> список файлов
    "replacements": {}  # Дата -> список файлов замен
}
schedule_index_lock = threading.Lock()
schedule_index_initialized = False

def update_teacher_access(teacher_name):
    """Track teacher access to identify popular teachers"""
    with access_counts_lock:
        if teacher_name not in teacher_access_counts:
            teacher_access_counts[teacher_name] = 0
        teacher_access_counts[teacher_name] += 1
        
        # Update popular teachers list if this teacher is frequently accessed
        if teacher_access_counts[teacher_name] >= 3 and teacher_name not in POPULAR_TEACHERS:
            with popular_teachers_lock:
                if teacher_name not in POPULAR_TEACHERS:
                    POPULAR_TEACHERS.append(teacher_name)
                    logger.info(f"Added {teacher_name} to popular teachers list")
                    
                    # Save popular teachers to disk for persistence across restarts
                    try:
                        with open("popular_teachers.json", "w") as f:
                            json.dump(POPULAR_TEACHERS, f)
                    except Exception as e:
                        logger.error(f"Error saving popular teachers: {e}")

def load_popular_teachers():
    """Load popular teachers from disk to persist across restarts"""
    global POPULAR_TEACHERS
    try:
        if os.path.exists("popular_teachers.json"):
            with open("popular_teachers.json", "r") as f:
                teachers = json.load(f)
                with popular_teachers_lock:
                    POPULAR_TEACHERS = teachers
                logger.info(f"Loaded {len(POPULAR_TEACHERS)} popular teachers from disk")
    except Exception as e:
        logger.error(f"Error loading popular teachers: {e}")

async def run_excel_task(func, *args, **kwargs):
    """Run Excel processing tasks in a dedicated thread pool with semaphore control."""
    loop = asyncio.get_running_loop()
    async with file_semaphore:
        return await loop.run_in_executor(excel_thread_pool, partial(func, *args, **kwargs))

# Apply the Excel cache decorator to our processing function
@use_excel_cache
def cached_parse_teacher_schedule(file_path, date_str, teacher_name):
    """Cached version of parse_teacher_schedule that uses the Excel cache."""
    try:
        return parse_teacher_schedule(file_path, date_str, teacher_name)
    except Exception as e:
        logger.error(f"Error in cached_parse_teacher_schedule for {file_path}, {date_str}, {teacher_name}: {e}")
        return {}

async def process_excel_file_for_teacher(file_path, date_str, teacher_name):
    """Process a single Excel file for a teacher on a specific date with error handling."""
    # Create a unique key for this file processing task
    task_key = f"{file_path}_{date_str}_{teacher_name}"
    
    # Check if this file is already being processed
    with files_lock:
        if task_key in files_being_processed:
            # Wait for the existing task to complete
            try:
                logger.debug(f"Waiting for existing processing of {file_path} for {date_str}")
                return await files_being_processed[task_key]
            except Exception:
                # If the existing task failed, we'll try again
                if task_key in files_being_processed:
                    del files_being_processed[task_key]
    
    # Use a simple Future instead of a Task to avoid task creation overhead
    processing_future = asyncio.Future()
    
    # Register the future
    with files_lock:
        files_being_processed[task_key] = processing_future
    
    try:
        # Process the file with a timeout
        try:
            # Process the file directly with a timeout
            result = await asyncio.wait_for(
                _process_excel_file(file_path, date_str, teacher_name),
                timeout=15  # 15 seconds timeout per file
            )
            # Set the result for any waiting tasks
            if not processing_future.done():
                processing_future.set_result(result)
            return result
        except asyncio.TimeoutError:
            logger.warning(f"Timeout processing {file_path} for teacher {teacher_name} on date {date_str}")
            # Provide an empty result on timeout
            empty_result = {}
            if not processing_future.done():
                processing_future.set_result(empty_result)
            return empty_result
    except Exception as e:
        # Handle any error
        logger.error(f"Error in process_excel_file_for_teacher for {file_path}, {date_str}, {teacher_name}: {e}")
        empty_result = {}
        if not processing_future.done():
            processing_future.set_exception(e)
        raise
    finally:
        # Always clean up
        with files_lock:
            if task_key in files_being_processed and files_being_processed[task_key] == processing_future:
                del files_being_processed[task_key]

async def _process_excel_file(file_path, date_str, teacher_name):
    """Internal function to process an Excel file."""
    try:
        # Skip Excel file preloading to avoid potential hangs
        # Use direct parsing with synchronous process
        return await run_blocking(parse_teacher_schedule, file_path, date_str, teacher_name)
    except Exception as e:
        logger.error(f"Error processing {file_path} for {date_str}: {e}")
        return {}

# Function to check if a file is applicable for a specific date with caching
def is_file_applicable_for_date(file_name, date_str):
    """Check if an Excel file is applicable for a specific date with caching."""
    # Create a cache key for this file and date
    cache_key = f"{file_name}_{date_str}"
    
    # Check if we have a cached result
    with file_applicability_lock:
        if cache_key in file_applicability_cache:
            return file_applicability_cache[cache_key]
    
    try:
        result = False
        # Define a regex pattern for date range files
        import re
        date_pattern = re.compile(r'^(\d{2}\.\d{2}\.\d{2,4})-(\d{2}\.\d{2}\.\d{2,4})\.xlsx$')
        match = date_pattern.match(file_name)
        
        if match:
            # This is a replacement file with dates
            dates = extract_dates_from_replacement_file(file_name)
            target_date = datetime.strptime(date_str, '%d.%m.%Y').date()
            
            # Check if the target date is in the extracted dates
            result = date_str in dates
            
            # If extraction failed, try direct comparison
            if not dates:
                start_str, end_str = match.groups()
                
                # Try different date formats
                try:
                    # Try with short year format (DD.MM.YY)
                    start_date = datetime.strptime(start_str, '%d.%m.%y').date()
                    end_date = datetime.strptime(end_str, '%d.%m.%y').date()
                except ValueError:
                    try:
                        # Try with full year format (DD.MM.YYYY)
                        start_date = datetime.strptime(start_str, '%d.%m.%Y').date()
                        end_date = datetime.strptime(end_str, '%d.%m.%Y').date()
                    except ValueError:
                        # If we can't parse the dates, assume not applicable
                        result = False
                
                # Check if target date is in range
                if not result and 'start_date' in locals() and 'end_date' in locals():
                    result = start_date <= target_date <= end_date
        else:
            # For regular schedule files (like ИСпВ-24-1.xlsx)
            # These are always applicable
            result = True
    except Exception as e:
        logger.error(f"Error checking if file {file_name} is applicable for date {date_str}: {e}")
        # In case of any error, assume the file is applicable to be safe
        result = True
    
    # Cache the result
    with file_applicability_lock:
        file_applicability_cache[cache_key] = result
    
    return result

async def get_teacher_schedule_optimized(teacher_name: str, start_date: str, end_date: str) -> str:
    """Optimized version that only processes relevant files for each date."""
    try:
        # Track teacher access to identify popular teachers
        update_teacher_access(teacher_name)
        
        # Check cache first
        cached_schedule = get_cached_teacher_schedule(teacher_name, start_date, end_date)
        if cached_schedule:
            logger.info(f"Using cached schedule for {teacher_name} from {start_date} to {end_date}")
            return cached_schedule

        all_schedules = {}
        start_date_obj = datetime.strptime(start_date, '%d.%m.%Y').date()
        end_date_obj = datetime.strptime(end_date, '%d.%m.%Y').date()

        # Get all files once
        all_files = await run_blocking(lambda: os.listdir("downloaded_files"))
        excel_files = [f for f in all_files if f.endswith('.xlsx')]
        
        # Определяем регулярное выражение для файлов с датами замен
        import re
        date_pattern = re.compile(r'^(\d{2}\.\d{2}\.\d{2,4})-(\d{2}\.\d{2}\.\d{2,4})\.xlsx$')
        
        # Организуем файлы по категориям
        replacement_files = []
        regular_files = []
        
        # Классифицируем файлы правильно
        for f in excel_files:
            if date_pattern.match(f):
                # Это файл замен с датами
                replacement_files.append(f)
            else:
                # Это обычный файл расписания
                regular_files.append(f)
        
        # Sort replacement files by date (newest first) to prioritize recent replacements
        replacement_files.sort(reverse=True)
        
        # Combine files with prioritization
        prioritized_files = regular_files + replacement_files
        
        # Preload Excel files in parallel
        preload_tasks = []
        for file in prioritized_files[:25]:  # Limit to 25 files to avoid overloading
            file_path = os.path.join("downloaded_files", file)
            preload_tasks.append(get_cached_workbook_async(file_path))
        
        # Start preloading but don't wait for completion
        preload_task = asyncio.create_task(asyncio.gather(*preload_tasks))
        
        # Get date ranges from replacement files
        replacement_date_ranges = []
        for file in replacement_files:
            dates = extract_dates_from_replacement_file(file)
            if dates:
                start_date_str = dates[0]
                end_date_str = dates[-1]
                try:
                    start_file_date = datetime.strptime(start_date_str, '%d.%m.%Y').date()
                    end_file_date = datetime.strptime(end_date_str, '%d.%m.%Y').date()
                    replacement_date_ranges.append((start_file_date, end_file_date))
                except Exception as e:
                    logger.error(f"Error parsing date range for file {file}: {e}")
        
        # Process each date
        tasks = []
        file_date_pairs = []
        
        current_date = start_date_obj
        while current_date <= end_date_obj:
            if current_date.weekday() != 6:  # Skip Sundays
                date_str = current_date.strftime('%d.%m.%Y')
                
                # Check if the current date is covered by any replacement file
                is_date_in_replacements = False
                for start_file_date, end_file_date in replacement_date_ranges:
                    if start_file_date <= current_date <= end_file_date:
                        is_date_in_replacements = True
                        break
                
                # Process all dates in the requested range
                if True:
                    # Filter files applicable for this date
                applicable_files = []
                
                # Regular schedule files are always applicable
                for file in regular_files:
                    applicable_files.append(file)
                
                # Process applicable replacement files
                for file in replacement_files:
                    if await run_blocking(is_file_applicable_for_date, file, date_str):
                        applicable_files.append(file)
                
                # Process each applicable file
                for file in applicable_files:
                    file_path = os.path.join("downloaded_files", file)
                    tasks.append(process_excel_file_for_teacher(file_path, date_str, teacher_name))
                    file_date_pairs.append((file, date_str))
            
            current_date += timedelta(days=1)

        # Wait for all tasks to complete
        results = await asyncio.gather(*tasks)

        # Process results
        for i, result in enumerate(results):
            if result:
                file, date_str = file_date_pairs[i]
                if date_str not in all_schedules:
                    all_schedules[date_str] = {}
                all_schedules[date_str].update(result)

        # Format the final schedule
        formatted_schedule = await run_blocking(
            format_teacher_schedule,
            all_schedules,
            teacher_name,
            start_date,
            end_date
        )

        # Cache the result with a longer expiration for popular teachers
        if teacher_name in POPULAR_TEACHERS:
            cache_teacher_schedule(teacher_name, start_date, end_date, formatted_schedule, expiration=3600)  # 1 hour
        else:
            cache_teacher_schedule(teacher_name, start_date, end_date, formatted_schedule)  # Default expiration
        
        # Make sure preloading is complete before returning
        try:
            await asyncio.wait_for(preload_task, timeout=0.1)
        except asyncio.TimeoutError:
            # It's okay if preloading is still ongoing
            pass
        
        return formatted_schedule

    except Exception as e:
        logger.error(f"Error getting optimized teacher schedule: {e}")
        return f"Ошибка при получении расписания преподавателя: {str(e)}"

# Function to preload schedules for popular teachers in the background
async def preload_teacher_schedules():
    """Preload schedules for popular teachers in the background."""
    try:
        # Get today's date and date a week from now
        today = datetime.now().date()
        next_week = today + timedelta(days=7)
        
        start_date = today.strftime('%d.%m.%Y')
        end_date = next_week.strftime('%d.%m.%Y')
        
        # Make a copy of the popular teachers list to avoid lock contention
        with popular_teachers_lock:
            teachers_to_preload = POPULAR_TEACHERS.copy()
        
        if not teachers_to_preload:
            logger.info("No popular teachers to preload schedules for")
            return
        
        logger.info(f"Preloading schedules for {len(teachers_to_preload)} popular teachers")
        
        # Preload schedules for each popular teacher
        for teacher_name in teachers_to_preload:
            try:
                # Check if we already have a cached schedule
                if get_cached_teacher_schedule(teacher_name, start_date, end_date):
                    continue
                
                # Generate and cache the schedule
                logger.info(f"Preloading schedule for popular teacher: {teacher_name}")
                await get_teacher_schedule_optimized(teacher_name, start_date, end_date)
                # Small delay to avoid overwhelming the system
                await asyncio.sleep(0.5)
            except Exception as e:
                logger.error(f"Error preloading schedule for {teacher_name}: {e}")
    except Exception as e:
        logger.error(f"Error in preload_teacher_schedules: {e}")

def start_background_processor():
    """Start a background processor to preload and maintain schedules."""
    global background_processor_running
    
    # Load popular teachers from disk
    load_popular_teachers()
    
    with background_processor_lock:
        if background_processor_running:
            logger.info("Background processor already running")
            return
        background_processor_running = True
    
    # Create a dedicated thread for the background processor
    thread = threading.Thread(target=_run_background_processor, daemon=True)
    thread.start()
    logger.info("Started background teacher schedule processor")

def _run_background_processor():
    """Run the background processor in a dedicated thread."""
    try:
        # Create a new event loop for this thread
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        # Run the background processor loop
        loop.run_until_complete(_background_processor_loop())
    except Exception as e:
        logger.error(f"Error in background processor: {e}")
        with background_processor_lock:
            global background_processor_running
            background_processor_running = False

async def _background_processor_loop():
    """Background processor loop to preload and maintain schedules."""
    try:
        logger.info("Starting background processor loop")
        
        # First, build the schedule index
        await build_schedule_index()
        
        # Second, precache popular teacher schedules
        await precache_popular_teachers()
        
        # Now continue with other background tasks
        while True:
            try:
                # Define the current time period for processing
                now = datetime.now()
                today = now.strftime('%d.%m.%Y')
                
                # Process next week's schedule for popular teachers
                future_date = (now + timedelta(days=7)).strftime('%d.%m.%Y')
                
                with popular_teachers_lock:
                    teachers_to_process = POPULAR_TEACHERS.copy() if POPULAR_TEACHERS else []
                
                if teachers_to_process:
                    logger.info(f"Background processing schedules for {len(teachers_to_process)} popular teachers")
                    for teacher in teachers_to_process[:10]:  # Process top 10 to avoid overload
                        try:
                            # Use indexed version for better performance
                            await get_teacher_schedule_with_index(teacher, today, future_date)
                        except Exception as e:
                            logger.error(f"Error background processing teacher {teacher}: {e}")
                
                # Sleep for a while before next processing cycle
                await asyncio.sleep(3600)  # Sleep for 1 hour
                
            except Exception as e:
                logger.error(f"Error in background processor cycle: {e}")
                await asyncio.sleep(1800)  # Sleep for 30 minutes on error
                
    except Exception as e:
        logger.error(f"Background processor loop terminated: {e}")

def extract_teachers_from_file(file_path):
    """Извлекает всех преподавателей из файла расписания"""
    teachers = set()
    try:
        # Use read_only for better performance
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        
        # Ищем ячейки с именами преподавателей (обычно под предметами)
        # Формат обычно: Фамилия И.О.
        teacher_pattern = re.compile(r'[А-Яа-я]+\s+[А-Я]\.[А-Я]\.')
        
        # Limit search to rows 1-50 and columns 1-12 to improve performance
        max_row = min(sheet.max_row, 50)
        max_col = min(sheet.max_column, 12)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    # Ищем все совпадения с паттерном ФИО
                    matches = teacher_pattern.findall(cell_value)
                    teachers.update(matches)
        
        # Close the workbook to release resources
        wb.close()
        return teachers
    except Exception as e:
        logger.error(f"Ошибка при извлечении преподавателей из {file_path}: {e}")
        return set()

def extract_dates_from_replacement_file(file_name):
    """Извлекает даты из имени файла замен (формат: DD.MM.YY-DD.MM.YY.xlsx)"""
    dates = []
    try:
        # Убираем расширение .xlsx
        base_name = file_name.replace('.xlsx', '')
        
        # Пытаемся извлечь диапазон дат
        if '-' in base_name:
            # Проверяем, соответствует ли имя файла формату дат
            # Шаблон для даты: NN.NN.NN-NN.NN.NN или NN.NN.NNNN-NN.NN.NNNN
            import re
            date_pattern = re.compile(r'^(\d{2}\.\d{2}\.\d{2,4})-(\d{2}\.\d{2}\.\d{2,4})$')
            match = date_pattern.match(base_name)
            
            if match:
                start_str, end_str = match.groups()
                
                # Пробуем разные форматы даты
                try:
                    # Формат DD.MM.YY
                    start_date = datetime.strptime(start_str, '%d.%m.%y')
                    end_date = datetime.strptime(end_str, '%d.%m.%y')
                except ValueError:
                    try:
                        # Формат DD.MM.YYYY
                        start_date = datetime.strptime(start_str, '%d.%m.%Y')
                        end_date = datetime.strptime(end_str, '%d.%m.%Y')
                    except ValueError:
                        return dates
                
                # Генерируем все даты в диапазоне
                current_date = start_date
                while current_date <= end_date:
                    dates.append(current_date.strftime('%d.%m.%Y'))
                    current_date += timedelta(days=1)
            else:
                # Это не файл дат, а файл группы (например, ИСпВ-24-1.xlsx)
                logger.debug(f"Файл {file_name} не соответствует формату замен с датами")
    
    except Exception as e:
        logger.error(f"Ошибка при извлечении дат из файла замен {file_name}: {e}")
    
    return dates

async def build_schedule_index():
    """Строит индекс всех файлов расписания и замен"""
    global schedule_index, schedule_index_initialized
    
    # Проверяем, не выполняется ли уже индексация
    with schedule_index_lock:
        if schedule_index_initialized:
            return
    
    try:
        # Временные структуры для построения индекса
        teachers_index = {}
        dates_index = {}
        replacements_index = {}
        
        files_dir = "downloaded_files"
        if not os.path.exists(files_dir):
            logger.warning(f"Директория {files_dir} не существует")
            return
        
        logger.info("Начинаем индексацию файлов расписания...")
        
        # Получаем список всех файлов
        file_list = os.listdir(files_dir)
        excel_files = [f for f in file_list if f.endswith('.xlsx')]
        
        # Определяем регулярное выражение для файлов с датами замен
        import re
        date_pattern = re.compile(r'^(\d{2}\.\d{2}\.\d{2,4})-(\d{2}\.\d{2}\.\d{2,4})\.xlsx$')
        
        # Организуем файлы по категориям
        replacement_files = []
        regular_files = []
        
        # Классифицируем файлы правильно
        for f in excel_files:
            if date_pattern.match(f):
                # Это файл замен с датами
                replacement_files.append(f)
            else:
                # Это обычный файл расписания
                regular_files.append(f)
        
        logger.info(f"Найдено {len(regular_files)} обычных файлов расписания и {len(replacement_files)} файлов замен")
        
        # Обрабатываем файлы замен
        for file_name in replacement_files:
            file_path = os.path.join(files_dir, file_name)
            
            # Индексируем файл замен по датам
            dates = extract_dates_from_replacement_file(file_name)
            
            for date in dates:
                if date not in replacements_index:
                    replacements_index[date] = []
                replacements_index[date].append(file_path)
                
                # Также добавляем в общий индекс дат
                if date not in dates_index:
                    dates_index[date] = []
                dates_index[date].append(file_path)
            
            # Добавляем преподавателей из файла замен
            teachers = await run_excel_task(extract_teachers_from_file, file_path)
            for teacher in teachers:
                if teacher not in teachers_index:
                    teachers_index[teacher] = []
                teachers_index[teacher].append(file_path)
        
        # Обрабатываем обычные файлы расписания
        for file_name in regular_files:
            file_path = os.path.join(files_dir, file_name)
            
            # Добавляем преподавателей из обычного файла
            teachers = await run_excel_task(extract_teachers_from_file, file_path)
            for teacher in teachers:
                if teacher not in teachers_index:
                    teachers_index[teacher] = []
                teachers_index[teacher].append(file_path)
        
        # Обновляем глобальный индекс
        with schedule_index_lock:
            schedule_index = {
                "teachers": teachers_index,
                "dates": dates_index,
                "replacements": replacements_index
            }
            schedule_index_initialized = True
        
        teacher_count = len(teachers_index)
        date_count = len(dates_index)
        replacement_count = len(replacements_index)
        
        logger.info(f"Индексация завершена: найдено {teacher_count} преподавателей, " + 
                   f"{date_count} дат, {replacement_count} периодов замен")
                   
    except Exception as e:
        logger.error(f"Ошибка при индексации файлов расписания: {e}")

def get_relevant_files_for_teacher_and_date(teacher_name, date_str):
    """Возвращает список релевантных файлов для преподавателя и даты"""
    relevant_files = set()
    
    try:
        with schedule_index_lock:
            # Проверяем, инициализирован ли индекс
            if not schedule_index_initialized:
                return []
            
            # Получаем файлы, в которых есть преподаватель
            if teacher_name in schedule_index["teachers"]:
                relevant_files.update(schedule_index["teachers"][teacher_name])
            
            # Получаем файлы для конкретной даты
            if date_str in schedule_index["dates"]:
                relevant_files.update(schedule_index["dates"][date_str])
            
            # Получаем файлы замен для этой даты
            if date_str in schedule_index["replacements"]:
                relevant_files.update(schedule_index["replacements"][date_str])
            
    except Exception as e:
        logger.error(f"Ошибка при поиске релевантных файлов: {e}")
    
    return list(relevant_files)

async def precache_popular_teachers():
    """Предварительно кэширует расписания популярных преподавателей"""
    try:
        # Получаем список популярных преподавателей
        with popular_teachers_lock:
            teachers_to_cache = POPULAR_TEACHERS.copy() if POPULAR_TEACHERS else []
        
        if not teachers_to_cache:
            logger.info("Нет популярных преподавателей для кэширования")
            return

        # Получаем текущую неделю и следующую
        today = datetime.now()
        start_date = today.strftime('%d.%m.%Y')
        end_date = (today + timedelta(days=14)).strftime('%d.%m.%Y')
        
        logger.info(f"Начинаем предварительное кэширование для {len(teachers_to_cache)} преподавателей")
        
        # Запускаем задачи кэширования для каждого преподавателя
        tasks = []
        for teacher_name in teachers_to_cache:
            task = asyncio.create_task(
                get_teacher_schedule_optimized(teacher_name, start_date, end_date)
            )
            tasks.append(task)
        
        # Ждем завершения всех задач кэширования
        await asyncio.gather(*tasks)
        logger.info("Предварительное кэширование популярных преподавателей завершено")
        
    except Exception as e:
        logger.error(f"Ошибка при предварительном кэшировании: {e}")

# Модифицируем существующую функцию для использования индекса файлов
async def get_teacher_schedule_with_index(teacher_name: str, start_date: str, end_date: str) -> str:
    """Использует индекс для оптимизации поиска расписания преподавателя"""
    try:
        # Track teacher access to identify popular teachers
        update_teacher_access(teacher_name)
        
        # Check cache first
        cached_schedule = get_cached_teacher_schedule(teacher_name, start_date, end_date)
        if cached_schedule:
            logger.info(f"Using cached schedule for {teacher_name} from {start_date} to {end_date}")
            return cached_schedule

        all_schedules = {}
        start_date_obj = datetime.strptime(start_date, '%d.%m.%Y').date()
        end_date_obj = datetime.strptime(end_date, '%d.%m.%Y').date()
        
        # Собираем информацию о файлах замен и их датах
        replacement_files_info = []
        files_dir = "downloaded_files"
        
        # Получаем список всех файлов Excel
        file_list = os.listdir(files_dir)
        excel_files = [f for f in file_list if f.endswith('.xlsx')]
        
        # Находим файлы с заменами (имя файла содержит дату в формате DD.MM.YY-DD.MM.YY.xlsx)
        import re
        date_pattern = re.compile(r'^(\d{2}\.\d{2}\.\d{2,4})-(\d{2}\.\d{2}\.\d{2,4})\.xlsx$')
        
        for file in excel_files:
            match = date_pattern.match(file)
            if match:
                try:
                    dates = file.replace('.xlsx', '').split('-')
                    if len(dates) == 2:
                        # Пробуем разные форматы даты
                        try:
                            # Сначала пробуем формат с двузначным годом
                            start_file_date = datetime.strptime(dates[0], '%d.%m.%y').date()
                            end_file_date = datetime.strptime(dates[1], '%d.%m.%y').date()
                    except ValueError:
                            try:
                                # Затем пробуем формат с четырехзначным годом
                                start_file_date = datetime.strptime(dates[0], '%d.%m.%Y').date()
                                end_file_date = datetime.strptime(dates[1], '%d.%m.%Y').date()
                            except ValueError:
                                logger.warning(f"Не удалось распознать формат даты в файле: {file}")
                                continue
                        
                        file_path = os.path.join(files_dir, file)
                        replacement_files_info.append((start_file_date, end_file_date, file_path))
                except Exception as e:
                    logger.error(f"Ошибка при обработке файла {file}: {str(e)}")
                    continue
        
        # Собираем только даты, которые действительно относятся к файлам с заменами
        dates_to_check = set()
        
        # Находим самую раннюю и самую позднюю даты для определения диапазона
        latest_end_date = None
        earliest_start_date = None
        
        for start_file_date, end_file_date, _ in replacement_files_info:
            if not latest_end_date or end_file_date > latest_end_date:
                latest_end_date = end_file_date
            if not earliest_start_date or start_file_date < earliest_start_date:
                earliest_start_date = start_file_date
        
        # Устанавливаем earliest_start_date как максимум между запрошенной начальной датой и самой ранней датой из файлов
        earliest_start_date = max(start_date_obj, earliest_start_date) if earliest_start_date else start_date_obj
        
        # Ограничиваем latest_end_date запрошенной конечной датой
        latest_end_date = min(end_date_obj, latest_end_date) if latest_end_date else end_date_obj
        
        # Перебираем все возможные даты между earliest_start_date и latest_end_date
        if earliest_start_date and latest_end_date:
            current_date = earliest_start_date
            while current_date <= latest_end_date:
                if current_date.weekday() != 6:  # Пропускаем воскресенье
                    # Проверяем, входит ли текущая дата в диапазон хотя бы одного файла с заменами
                    is_date_in_replacement = False
                    for start_file_date, end_file_date, _ in replacement_files_info:
                        if start_file_date <= current_date <= end_file_date:
                            is_date_in_replacement = True
                            break
                    
                    # Process all dates in the requested range
                    if True:
                        dates_to_check.add(current_date.strftime('%d.%m.%Y'))
                
                current_date += timedelta(days=1)
        
        logger.info(f"Даты для проверки: {sorted(dates_to_check)}")
        
        # Если нет дат для проверки, возвращаем сообщение
        if not dates_to_check:
            return f"Расписание для {teacher_name} на указанный период не найдено (нет актуальных файлов замен)"
        
        # Получаем все файлы, которые содержат упоминание преподавателя
        teacher_files = set()
        with schedule_index_lock:
            if schedule_index_initialized and teacher_name in schedule_index["teachers"]:
                teacher_files = set(schedule_index["teachers"][teacher_name])
        
        # Добавляем файлы замен
        replacement_files = set([file_path for _, _, file_path in replacement_files_info])
        
        # Объединяем все релевантные файлы
        all_relevant_files = teacher_files.union(replacement_files)
        
        if not all_relevant_files:
            logger.warning(f"Не найдено релевантных файлов для {teacher_name} на период {start_date} - {end_date}")
            return f"Расписание для {teacher_name} на указанный период не найдено"
            
        # Логирование найденных файлов
        logger.info(f"Для {teacher_name} найдено {len(teacher_files)} файлов с упоминанием преподавателя")
        logger.info(f"Найдено {len(replacement_files)} файлов замен для дат в диапазоне {start_date} - {end_date}")
        
        # Устанавливаем общий таймаут на всю операцию
        overall_start_time = time.time()
        overall_timeout = 30  # 30 секунд на весь процесс
        
        # Ограничиваем количество одновременно обрабатываемых файлов
        MAX_CONCURRENT_TASKS = 10
        dates_processed = []
        
        # Обрабатываем только отфильтрованные даты
        for date_str in sorted(dates_to_check):
            if time.time() - overall_start_time >= overall_timeout:
                break
                
            dates_processed.append(date_str)
            # Обрабатываем файлы для этой даты
            date_schedule = {}
            
            # Создаем задачи для обработки файлов
            files_to_process = []
            
            for file_path in all_relevant_files:
                if os.path.exists(file_path):
                    # Проверяем, применим ли файл к текущей дате
                    file_name = os.path.basename(file_path)
                    if is_file_applicable_for_date(file_name, date_str) or not '-' in file_name:
                        files_to_process.append(file_path)
                else:
                    logger.warning(f"Файл не найден: {file_path}")
            
            # Обрабатываем файлы партиями, чтобы избежать перегрузки
            for i in range(0, len(files_to_process), MAX_CONCURRENT_TASKS):
                if time.time() - overall_start_time >= overall_timeout:
                    break
                    
                batch = files_to_process[i:i+MAX_CONCURRENT_TASKS]
                batch_tasks = [process_excel_file_for_teacher(file_path, date_str, teacher_name) 
                              for file_path in batch]
                
                if batch_tasks:
                    logger.info(f"Обрабатываем партию из {len(batch_tasks)} файлов для {teacher_name} на дату {date_str}")
                    try:
                        # Добавляем таймаут на уровне партии
                        results = await asyncio.wait_for(
                            asyncio.gather(*batch_tasks, return_exceptions=True),
                            timeout=20  # 20 секунд на партию файлов
                        )
                        
                        # Объединяем результаты, игнорируя ошибки
                        for result in results:
                            if isinstance(result, Exception):
                                logger.error(f"Ошибка при обработке файла: {result}")
                            elif result:
                                date_schedule.update(result)
                    except asyncio.TimeoutError:
                        logger.warning(f"Превышен таймаут обработки партии файлов для {date_str}")
                    except Exception as e:
                        logger.error(f"Ошибка при выполнении задач для даты {date_str}: {e}")
            
            # Сохраняем результаты для этой даты
            if date_schedule:
                all_schedules[date_str] = date_schedule
                logger.info(f"Для даты {date_str} найдено {len(date_schedule)} пар")
            else:
                logger.info(f"Для даты {date_str} не найдено пар")
        
        # Проверяем, не истек ли общий таймаут
        if time.time() - overall_start_time >= overall_timeout:
            logger.warning(f"Превышен общий таймаут обработки расписания для {teacher_name}")
            # Если есть хотя бы какие-то данные, продолжаем обработку
            if not all_schedules:
                return f"Не удалось получить расписание для {teacher_name} в указанный срок. Пожалуйста, попробуйте еще раз."
        
        # Если мы не смогли найти ни одной пары для всех дат
        if not all_schedules:
            logger.warning(f"Расписание не найдено для {teacher_name} на {start_date}-{end_date} (обработано дат: {len(dates_processed)})")
            return f"Расписание для {teacher_name} на указанный период не найдено"
        
        # Форматируем результат
        logger.info(f"Форматирование расписания для {teacher_name} на {start_date}-{end_date}")
        formatted_schedule = format_teacher_schedule(all_schedules, teacher_name, start_date, end_date)
        
        # Кэшируем результат
        cache_teacher_schedule(teacher_name, start_date, end_date, formatted_schedule)
        
        # Логирование результата
        schedule_lines = formatted_schedule.count('\n') + 1
        logger.info(f"Успешно сформировано расписание для {teacher_name} ({schedule_lines} строк)")
        
        return formatted_schedule
        
    except Exception as e:
        logger.error(f"Ошибка при получении расписания для {teacher_name}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return f"Произошла ошибка при обработке запроса: {str(e)[:100]}" 