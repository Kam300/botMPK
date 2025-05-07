import os
import logging
import openpyxl
import asyncio
from datetime import datetime, timedelta
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram import ReplyKeyboardMarkup, KeyboardButton
from telegram import BotCommand
from telegram.ext import PicklePersistence 
from functools import wraps  # Add this import
from telegram.ext import (
    ContextTypes,
    Application,
    CommandHandler,
    ConversationHandler,
    MessageHandler,
    filters,
    CallbackQueryHandler
)
import sys
import signal
from dropbox_sync import sync_files, sync_files_async, get_dropbox_client, schedule_sync, is_update_in_progress, get_update_status_message
import threading
from functools import lru_cache
from concurrent.futures import ThreadPoolExecutor, as_completed
import asyncio
import re
import json  # Added missing json import
from cache_utils import (
    get_cached_student_schedule,
    cache_student_schedule,
    init_cache,
    get_cached_teacher_schedule,
    cache_teacher_schedule,
    cache_classroom_schedule,
    get_cached_classroom_schedule,
    selective_cache_clear
)

from concurrent.futures import ThreadPoolExecutor
from functools import partial
from threading import Thread
import httpx  # Add this import





TELEGRAM_TOKEN="5849256613:AAH34MtjRPyBhrtQouFseQzVw5G9KJsX1WQ"
file_access_semaphore = asyncio.Semaphore(3)
# ID администраторов, которые могут очищать кэш
# Чтобы узнать свой ID, отправьте боту команду /clear_cache и посмотрите ID в ответном сообщении
# или воспользуйтесь ботом @userinfobot в Telegram
ADMIN_IDS = ["1809028797", "1809028797"]  # Замените на реальные ID администраторов

# Настройка логирования
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# Добавляем состояния разговора
ENTER_GROUP, CHOOSE_SUBGROUP, CHOOSE_ACTION = range(3)
ENTER_TEACHER, CHOOSE_DATE_FOR_TEACHER = range(4, 6)
ENTER_CLASSROOM, CHOOSE_DATE_FOR_CLASSROOM = range(6, 8) 
SUBSCRIBERS_FILE = "subscribers.json"

# Добавляем в начало файла после других импортов
schedule_cache = {}
cache_lock = threading.Lock()

# Словарь для перевода дней недели
days_ru = {
    0: 'понедельник',
    1: 'вторник',
    2: 'среда',
    3: 'четверг',
    4: 'пятница',
    5: 'суббота',
    6: 'воскресенье'
}

# Глобальный ThreadPoolExecutor для обработки тяжелых задач
thread_pool = ThreadPoolExecutor(max_workers=20)  # Increased number of workers
# Improve the non_blocking_handler to handle connection errors during shutdown
# Improve the non_blocking_handler to handle connection errors during shutdown
def non_blocking_handler(handler_func):
    """Decorator to make command handlers non-blocking"""
    @wraps(handler_func)
    async def wrapper(update, context):
        # Send immediate response
        if update.message and not update.callback_query:
            try:
                await update.message.reply_text("Обрабатываю запрос, пожалуйста, подождите...")
            except Exception as e:
                logger.error(f"Error sending initial response: {e}")
        
        # Create a task that runs in the background
        asyncio.create_task(_run_handler(update, context, handler_func))
    
    return wrapper


    # First, make sure we have a proper run_blocking function
async def run_blocking(func, *args, **kwargs):
    """Run blocking function in threadpool."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(
        thread_pool, 
        lambda: func(*args, **kwargs)
    )

async def _run_handler(update, context, handler_func):
    """Helper function to run the handler in the current event loop"""
    try:
        await handler_func(update, context)
    except httpx.ReadError:
        logger.warning("Connection closed while sending message - application may be shutting down")
    except RuntimeError as e:
        if "Event loop is closed" in str(e):
            logger.warning("Event loop is closed, cannot complete the request")
        else:
            logger.error(f"Runtime error in handler: {e}")
            try:
                await update.message.reply_text("Произошла ошибка при обработке запроса")
            except Exception:
                pass
    except Exception as e:
        logger.error(f"Error in background handler: {e}")
        try:
            await update.message.reply_text(f"Произошла ошибка при обработке запроса: {str(e)}")
        except Exception:
            logger.error("Could not send error message to user")

# Добавляем функцию для кэширования расписания
def cache_schedule(file_path, date_str, teacher_name, schedule):
    """Кэширует расписание с временем жизни"""
    cache_key = f"{file_path}:{date_str}:{teacher_name}"
    with cache_lock:
        schedule_cache[cache_key] = {
            'schedule': schedule,
            'timestamp': datetime.now()
        }


def get_cached_schedule(file_path, date_str, teacher_name):
    """Получает расписание из кэша, если оно есть и не устарело"""
    cache_key = f"{file_path}:{date_str}:{teacher_name}"
    with cache_lock:
        if cache_key in schedule_cache:
            cached_data = schedule_cache[cache_key]
            # Проверяем, не устарел ли кэш (30 минут)
            if (datetime.now() - cached_data['timestamp']).total_seconds() < 1800:
                return cached_data['schedule']
    return None

# Функция для загрузки списка подписчиков
def load_subscribers():
    try:
        if os.path.exists(SUBSCRIBERS_FILE):
            with open(SUBSCRIBERS_FILE, 'r') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f"Ошибка при загрузке списка подписчиков: {e}")
        return {}

# Функция для сохранения списка подписчиков
def save_subscribers(subscribers):
    try:
        with open(SUBSCRIBERS_FILE, 'w') as f:
            json.dump(subscribers, f)
    except Exception as e:
        logger.error(f"Ошибка при сохранении списка подписчиков: {e}")

# Функция для проверки подписки пользователя
def is_subscribed(user_id):
    subscribers = load_subscribers()
    return str(user_id) in subscribers

# Обработчик для команды подписки
async def subscribe_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    username = update.effective_user.username or "Unknown"
    
    # Проверяем, подписан ли уже пользователь
    if is_subscribed(user_id):
        await update.message.reply_text(
            "✅ У вас уже есть активная подписка на уведомления о заменах.\n"
            "Вы будете получать уведомления о новых заменах."
        )
        return
    
    # Добавляем пользователя в список подписчиков
    subscribers = load_subscribers()
    subscribers[str(user_id)] = {
        "username": username,
        "subscribed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "active": True
    }
    save_subscribers(subscribers)
    
    await update.message.reply_text(
        "✅ Подписка успешно активирована!\n\n"
        "Теперь вы будете получать уведомления о новых заменах "
    )
    
    logger.info(f"Новый подписчик: {username} (ID: {user_id})")

# Обработчик для команды отписки
async def unsubscribe_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    
    # Проверяем, подписан ли пользователь
    if not is_subscribed(user_id):
        await update.message.reply_text(
            "❌ У вас нет активной подписки на уведомления о заменах."
        )
        return
    
    # Удаляем пользователя из списка подписчиков
    subscribers = load_subscribers()
    if str(user_id) in subscribers:
        del subscribers[str(user_id)]
        save_subscribers(subscribers)
    
    await update.message.reply_text(
        "✅ Вы успешно отписались от уведомлений о заменах."
    )
    
    logger.info(f"Пользователь отписался: ID: {user_id}")

# Обработчик для проверки и отправки уведомлений
async def check_notifications(context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        if os.path.exists("pending_notifications.json"):
            with open("pending_notifications.json", "r") as f:
                notification_data = json.load(f)
            
            message = notification_data.get("message", "")
            
            if message:
                # Загружаем актуальный список подписчиков
                current_subscribers = load_subscribers()
                
                # Отправляем уведомления всем активным подписчикам
                for user_id in current_subscribers:
                    try:
                        await context.bot.send_message(chat_id=user_id, text=message)
                        logger.info(f"Уведомление отправлено пользователю {user_id}")
                    except Exception as e:
                        logger.error(f"Ошибка при отправке уведомления пользователю {user_id}: {e}")
                
                # Удаляем файл после отправки
                os.remove("pending_notifications.json")
                logger.info("Все уведомления отправлены и файл удален")
    
    except Exception as e:
        logger.error(f"Ошибка при проверке и отправке уведомлений: {e}")

def get_week_type(date_str):
    """Определяет тип недели на основе референсной даты"""
    # Начальная дата семестра (первая неделя - нечетная)
    semester_start = datetime.strptime('29.01.2025', '%d.%m.%Y')
    date_obj = datetime.strptime(date_str, '%d.%m.%Y')

    # Получаем начало недели для проверяемой даты (понедельник)
    start_of_week = date_obj - timedelta(days=date_obj.weekday())

    # Получаем начало первой недели семестра (понедельник)
    start_of_semester_week = semester_start - timedelta(days=semester_start.weekday())

    # Вычисляем номер недели от начала семестра (начиная с 0)
    delta_weeks = (start_of_week - start_of_semester_week).days // 7

    # Добавляем отладочную информацию
    logger.info(f"Date: {date_str}")
    logger.info(f"Start of week: {start_of_week}")
    logger.info(f"Start of semester week: {start_of_semester_week}")
    logger.info(f"Delta weeks: {delta_weeks}")

    # Меняем логику: если delta_weeks четное (включая 0), то неделя нечетная
    week_type = 'нечетная' if delta_weeks % 2 == 0 else 'четная'
    logger.info(f"Week type: {week_type}")

    return week_type


def find_day_column(sheet, day_name, start_row, is_even_week):
    """
    Ищет колонку и строку начала расписания для конкретного дня
    """
    logger.info(f"Searching for day: {day_name}")
    logger.info(f"Is even week: {is_even_week}")

    # Определяем тип недели для поиска в Excel
    week_type_text = "четная неделя" if is_even_week else "нечетная неделя"
    logger.info(f"Looking for week type: {week_type_text}")

    # Ищем все вхождения типа недели
    week_rows = []
    for row in range(1, sheet.max_row):
        cell_value = str(sheet.cell(row=row, column=1).value or '').strip().lower()
        if cell_value == week_type_text:
            week_rows.append(row)
            logger.info(f"Found week type at row: {row}, col: 1, value: {cell_value}")

    if not week_rows:
        logger.warning(f"Week type '{week_type_text}' not found in Excel")
        return None, None

    # Для каждого найденного вхождения недели ищем день
    for week_row in week_rows:
        # Ищем день недели в пределах следующих 30 строк после маркера недели
        for row in range(week_row, min(week_row + 30, sheet.max_row)):
            for col in range(1, sheet.max_column + 1):
                cell_value = str(sheet.cell(row=row, column=col).value or '').strip().lower()
                if cell_value == day_name.lower():
                    logger.info(f"Found day at row: {row}, col: {col}, value: {cell_value}")
                    return col, row

    return None, None


def is_theory_lesson(subject):
    """Проверяет, является ли пара теоретической"""
    if not subject:
        return False
    subject = str(subject).lower()
    # Проверяем на явные указатели теоретической пары
    if '(то)' in subject:
        return True
    # Проверяем на практические/лабораторные/профильные работы
    non_theory_indicators = ['(пр)', '(лаб)', '(кп)', '(проф.)']
    # Специальные предметы, которые всегда по подгруппам
    subgroup_subjects = ['ин.яз']
    # Если предмет в списке подгрупповых, он не теоретический
    if any(subj in subject for subj in subgroup_subjects):
        return False
    # Если есть хотя бы один индикатор не теоретической пары, значит пара не теоретическая
    return not any(indicator in subject for indicator in non_theory_indicators)

def parse_teacher_schedule(schedule_file, date_str, teacher_name):
    """Парсит расписание для преподавателя"""
    try:
        # Remove general info logs
        # logger.info(f"Начало парсинга расписания для преподавателя {teacher_name} на дату {date_str}")
        # logger.info(f"Обрабатываемый файл: {schedule_file}")

        # Функция для проверки, является ли текст отменой
        def is_cancellation_text(text):
            if not text:
                return True
                
            # Remove debug log
            # logger.info(f"Проверка текста на отмену: '{text}'")
            
            text = text.strip()
            
            # Проверка на шаблон с дефисами разной длины
            if any(pattern in text.replace(' ', '') for pattern in ['----', '-----', '------', '-------', '--------', '---------', '----------', '-----------', '------------']):
                # Keep cancellation logs
                logger.info(f"Текст '{text}' является отменой (содержит последовательность дефисов)")
                return True
            
            # Прямая проверка на шаблон "1. ------------" или "1.------------"
            if (text.startswith('1. ') or text.startswith('1.')) and set(text[2:].strip('- ')).issubset({'-', ' '}):
                # Keep cancellation logs
                logger.info(f"Текст '{text}' является отменой (специальный шаблон '1. ------------')")
                return True
            
            if text == '------------' or text == '-' or text == '---' or text == '----':
                # Keep cancellation logs
                logger.info(f"Текст '{text}' является отменой (простой шаблон)")
                return True
                
            return False
            
        # Функция для поиска новых пар преподавателя в файлах замен
        def find_new_lessons_in_replacements():
            new_lessons = {}
            replacement_files = [f for f in os.listdir("downloaded_files") 
                              if f.endswith('.xlsx') and '-' in f]
            
            # Remove general info log
            # logger.info(f"Найдены файлы замен для проверки: {replacement_files}")
            
            for replacement_file in replacement_files:
                try:
                    replacement_path = os.path.join("downloaded_files", replacement_file)
                    dates = replacement_file.replace('.xlsx', '').split('-')
                    if len(dates) != 2:
                        continue
                    
                    # Пробуем разные форматы даты
                    start_date = None
                    end_date = None
                    try:
                        # Пробуем формат DD.MM.YY
                        start_date = datetime.strptime(dates[0], '%d.%m.%y').date()
                        end_date = datetime.strptime(dates[1], '%d.%m.%y').date()
                    except ValueError:
                        try:
                            # Пробуем формат DD.MM.YYYY
                            start_date = datetime.strptime(dates[0], '%d.%m.%Y').date()
                            end_date = datetime.strptime(dates[1], '%d.%m.%Y').date()
                        except ValueError:
                            logger.warning(f"Не удалось распознать формат даты в файле: {replacement_file}")
                            continue
                    
                    check_date = datetime.strptime(date_str, '%d.%m.%Y').date()
                    
                    # Remove file check log
                    # logger.info(f"Проверка файла {replacement_file}: Искомая дата {check_date}, в файле {start_date} - {end_date}")

                    if start_date <= check_date <= end_date:
                        # Remove found file log
                        # logger.info(f"Найден подходящий файл замен: {replacement_file} для даты {date_str}")
                        
                        # Загружаем файл замен
                        wb_replacements = openpyxl.load_workbook(replacement_path)
                        sheet_replacements = wb_replacements.active
                        
                        # Ищем замены для текущей даты
                        current_date = None
                        
                        # Проверяем все колонки и строки
                        for row in range(3, sheet_replacements.max_row + 1):
                            # Проверяем дату
                            date_cell = sheet_replacements.cell(row=row, column=2).value
                            if date_cell:
                                try:
                                    current_date = datetime.strptime(str(date_cell), '%d.%m.%Y').strftime('%d.%m.%Y')
                                except ValueError:
                                    continue
                                    
                            if current_date != date_str:
                                continue
                                
                            # Получаем номер пары
                            lesson_num = sheet_replacements.cell(row=row, column=3).value
                            if not isinstance(lesson_num, (int, float)):
                                continue
                                
                            lesson_num = int(lesson_num)
                            
                            # Проверяем все группы
                            for col in range(4, sheet_replacements.max_column + 1):
                                group_cell = sheet_replacements.cell(row=2, column=col).value
                                replacement = sheet_replacements.cell(row=row, column=col).value
                                
                                if not group_cell or not replacement:
                                    continue
                                    
                                group_name = str(group_cell).strip()
                                replacement_text = str(replacement)
                                
                                # Проверяем, упоминается ли преподаватель в тексте замены
                                if teacher_name.lower() in replacement_text.lower():
                                    # Keep replacement found log
                                    logger.info(f"Найдена замена с упоминанием преподавателя {teacher_name} в группе {group_name}, пара {lesson_num}: {replacement_text}")
                                    
                                    # Определяем подгруппу
                                    subgroup = None
                                    subject = ""
                                    room = ""
                                    
                                    # Разбиваем на строки
                                    lines = replacement_text.split('\n')
                                    
                                    # Проверяем каждую строку на наличие имени преподавателя
                                    teacher_found = False
                                    for i, line in enumerate(lines):
                                        if teacher_name.lower() in line.lower():
                                            teacher_found = True
                                            # Определяем, какой подгруппе принадлежит этот преподаватель
                                            for j in range(i, -1, -1):
                                                if j < len(lines) and lines[j].strip().startswith('1.'):
                                                    subgroup = 1
                                                    subject_line = lines[j]
                                                    # Извлекаем название предмета
                                                    subject_match = re.search(r'1\.\s*(\(.*?\))?\s*(.*?)(?=[АA]\d{3,4}|\n|$)', subject_line)
                                                    if subject_match:
                                                        subject_type = subject_match.group(1) or ''
                                                        subject_name = subject_match.group(2) or ''
                                                        subject = (subject_type + ' ' + subject_name).strip()
                                                    break
                                                elif j < len(lines) and lines[j].strip().startswith('2.'):
                                                    subgroup = 2
                                                    subject_line = lines[j]
                                                    # Извлекаем название предмета
                                                    subject_match = re.search(r'2\.\s*(\(.*?\))?\s*(.*?)(?=[АA]\d{3,4}|\n|$)', subject_line)
                                                    if subject_match:
                                                        subject_type = subject_match.group(1) or ''
                                                        subject_name = subject_match.group(2) or ''
                                                        subject = (subject_type + ' ' + subject_name).strip()
                                                    break
                                            
                                            # Если подгруппа не определена, возможно общая пара
                                            if not subgroup and '.' not in line:
                                                subject_match = re.search(r'(\(.*?\))?\s*(.*?)(?=[АA]\d{3,4}|\n|$)', line)
                                                if subject_match:
                                                    subject_type = subject_match.group(1) or ''
                                                    subject_name = subject_match.group(2) or ''
                                                    subject = (subject_type + ' ' + subject_name).strip()
                                            
                                            # Извлекаем номер аудитории
                                            room_match = re.search(r'[АA]\d{3,4}', line)
                                            if room_match:
                                                room = room_match.group(0)
                                            break
                                    
                                    # Если не нашли номер аудитории в строке с преподавателем, ищем в других строках
                                    if not room and teacher_found:
                                        for line in lines:
                                            room_match = re.search(r'[АA]\d{3,4}', line)
                                            if room_match:
                                                room = room_match.group(0)
                                                break
                                    
                                    # Если нашли замену для преподавателя и извлекли хотя бы какую-то информацию
                                    if teacher_found and (subject or room):
                                        # Очищаем имя предмета, если оно содержит только подгруппу
                                       
                                        # Если нет типа занятия в названии предмета, но оно есть в тексте замены
                                        if '(' not in subject:
                                            type_match = re.search(r'\((Пр|Лаб|КП|ТО)\)', replacement_text)
                                            if type_match:
                                                subject = f"({type_match.group(1)}) {subject}"
                                        
                                        # Добавляем новую пару в расписание
                                        new_lessons[lesson_num] = {
                                            'subject': f"✏️ {subject}",
                                            'teacher': teacher_name,
                                            'room': room,
                                            'is_common': subgroup is None,
                                            'subgroup': subgroup,
                                            'group': group_name,
                                            'is_replacement': True,
                                            'emoji': '✏️'
                                        }
                                        # Keep added lesson log
                                        logger.info(f"Добавлена новая пара {lesson_num} для преподавателя {teacher_name}: {subject}, аудитория {room}, подгруппа {subgroup}")
                
                except Exception as e:
                    logger.error(f"Ошибка при поиске новых пар в файле замен {replacement_file}: {str(e)}")
            
            return new_lessons
            
        wb = openpyxl.load_workbook(schedule_file)
        sheet = wb.active
        schedule = {}
        
        # Словарь для отслеживания пар преподавателя в обычном расписании
        teacher_lessons = {}

        # Получаем название группы из первой строки
        group_name = str(sheet.cell(row=1, column=1).value or '').split('группы ')[-1].strip()
        # Remove group found log
        # logger.info(f"Найдена группа: {group_name}")

        # Определяем тип недели
        is_even_week = get_week_type(date_str) == 'четная'

        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        weekday = date_obj.weekday()
        day_name = days_ru[weekday]

        # Ищем колонку с нужным днем
        day_col, current_start_row = find_day_column(sheet, day_name, 3, is_even_week)

        if not day_col:
            logger.warning(f"Не найден день недели или тип недели в файле {schedule_file}")
            
            # Даже если не нашли в обычном расписании, проверяем замены
            new_lessons = find_new_lessons_in_replacements()
            return new_lessons

        # Remove found row/column log
        # logger.info(f"Найдена начальная строка: {current_start_row}, колонка дня: {day_col}")

        # Сначала собираем все пары преподавателя в обычном расписании
        teacher_lessons = {}
        current_row = current_start_row + 1
        while current_row < current_start_row + 15:
            lesson_num = sheet.cell(row=current_row, column=day_col).value
            if not lesson_num:
                break

            try:
                if isinstance(lesson_num, (int, float)):
                    lesson_num = int(lesson_num)
                else:
                    num_str = ''.join(filter(str.isdigit, str(lesson_num)))
                    if not num_str:
                        current_row += 2
                        continue
                    lesson_num = int(num_str)

                subject_first = sheet.cell(row=current_row, column=day_col + 1).value
                subject_second = sheet.cell(row=current_row, column=day_col + 3).value
                room_first = sheet.cell(row=current_row + 1, column=day_col + 2).value
                room_second = sheet.cell(row=current_row + 1, column=day_col + 4).value
                teacher_first = sheet.cell(row=current_row + 1, column=day_col + 1).value
                teacher_second = sheet.cell(row=current_row + 1, column=day_col + 3).value

                # Проверяем, есть ли преподаватель в этой паре
                if teacher_first and teacher_name.lower() in str(teacher_first).lower():
                    teacher_lessons[lesson_num] = {
                        'subject': subject_first,
                        'teacher': teacher_first,
                        'room': room_first,
                        'subgroup': 1 if not is_theory_lesson(subject_first) else None,
                        'is_common': is_theory_lesson(subject_first),
                        'emoji': None  # Добавляем поле для эмодзи, но оставляем его пустым для обычных пар
                    }
                    # Add log for found lesson in original schedule
                    logger.info(f"Найдена пара №{lesson_num} преподавателя {teacher_name} в группе {group_name}: предмет '{subject_first}', аудитория {room_first}")
                
                if teacher_second and teacher_name.lower() in str(teacher_second).lower():
                    teacher_lessons[lesson_num] = {
                        'subject': subject_second,
                        'teacher': teacher_second,
                        'room': room_second,
                        'subgroup': 2 if not is_theory_lesson(subject_second) else None,
                        'is_common': is_theory_lesson(subject_second),
                        'emoji': None  # Добавляем поле для эмодзи, но оставляем его пустым для обычных пар
                    }
                    # Add log for found lesson in original schedule
                    logger.info(f"Найдена пара №{lesson_num} преподавателя {teacher_name} в группе {group_name}: предмет '{subject_second}', аудитория {room_second}")

            except ValueError as e:
                logger.error(f"Ошибка при обработке строки {current_row}: {str(e)}")

            current_row += 2

        # Keep summary of lessons found, but make it more concise
        if teacher_lessons:
            logger.info(f"Найдено {len(teacher_lessons)} пар преподавателя {teacher_name} в группе {group_name}")
        
        # Теперь обрабатываем обычное расписание
        current_row = current_start_row + 1

        # Получаем замены для текущей даты
        replacement_files = [f for f in os.listdir("downloaded_files")
                            if f.endswith('.xlsx') and '-' in f]

        # Проверяем замены перед парсингом основного расписания
        replacement_applied = set()  # Множество для отслеживания номеров пар, для которых были применены замены
        for replacement_file in replacement_files:
            try:
                replacement_path = os.path.join("downloaded_files", replacement_file)
                dates = replacement_file.replace('.xlsx', '').split('-')
                if len(dates) == 2:
                    start_date = datetime.strptime(dates[0], '%d.%m.%y').date()
                    end_date = datetime.strptime(dates[1], '%d.%m.%y').date()
                    check_date = datetime.strptime(date_str, '%d.%m.%Y').date()

                    if start_date <= check_date <= end_date:
                        # Загружаем файл замен
                        wb_replacements = openpyxl.load_workbook(replacement_path)
                        sheet_replacements = wb_replacements.active

                        # Ищем колонку группы
                        group_col = None
                        for col in range(4, sheet_replacements.max_column + 1):
                            group_cell = sheet_replacements.cell(row=2, column=col).value
                            if group_cell and group_name in str(group_cell):
                                group_col = col
                                break

                        if group_col:
                            # Ищем замены для текущей даты
                            current_date = None
                            for row in range(3, sheet_replacements.max_row + 1):
                                date_cell = sheet_replacements.cell(row=row, column=2).value
                                if date_cell:
                                    try:
                                        current_date = datetime.strptime(str(date_cell), '%d.%m.%Y').strftime(
                                            '%d.%m.%Y')
                                    except ValueError:
                                        continue

                                if current_date == date_str:
                                    lesson_num = sheet_replacements.cell(row=row, column=3).value
                                    replacement = sheet_replacements.cell(row=row, column=group_col).value
                                    
                                    # Если найдена замена для этого урока, добавляем номер урока в множество
                                    if isinstance(lesson_num, (int, float)):
                                        lesson_num = int(lesson_num)
                                        # Отмечаем, что для этого урока есть запись в файле замен
                                        # Но если это КП класс, то нужно проверить, для какой подгруппы применяется замена
                                        if lesson_num in teacher_lessons:
                                            current_lesson = teacher_lessons[lesson_num]
                                            if '(КП)' in str(current_lesson.get('subject', '')):
                                                # Проверяем, для какой подгруппы замена
                                                if replacement:
                                                    replacement_text = str(replacement)
                                                    if '1.' in replacement_text and '2.' in replacement_text:
                                                        # Если есть обе подгруппы, то добавляем как обычно
                                                        replacement_applied.add(lesson_num)
                                                    elif '1.' in replacement_text and current_lesson.get('subgroup') == 1:
                                                        # Если только 1-я подгруппа и преподаватель ведет 1-ю
                                                        replacement_applied.add(lesson_num)
                                                    elif '2.' in replacement_text and current_lesson.get('subgroup') == 2:
                                                        # Если только 2-я подгруппа и преподаватель ведет 2-ю
                                                        replacement_applied.add(lesson_num)
                                            else:
                                                # Для обычных пар просто отмечаем
                                                replacement_applied.add(lesson_num)
                                        else:
                                            replacement_applied.add(lesson_num)
                                    
                                    # Remove general replacement check log
                                    # logger.info(f"Проверка замены для пары {lesson_num} в группе {group_name}: '{replacement}'")
                                    
                                    # Проверяем, есть ли у преподавателя пара в это время
                                    if isinstance(lesson_num, (int, float)):
                                        lesson_num = int(lesson_num)
                                        
                                        # Проверяем, упоминается ли преподаватель в замене
                                        if replacement and teacher_name.lower() in str(replacement).lower():
                                            replacement_text = str(replacement)
                                            # Keep log for replacements with teacher
                                            logger.info(f"Обработка замены с упоминанием преподавателя {teacher_name}: '{replacement_text}'")
                                            
                                            # Специальная проверка на отмену по шаблону "1. ------------"
                                            if (replacement_text.startswith('1. ') or replacement_text.startswith('1.')) and set(replacement_text[2:].strip('- ')).issubset({'-', ' '}):
                                                logger.info(f"!!! ПРЯМАЯ ПРОВЕРКА: Обнаружена отмена по шаблону '1. ------------': {replacement_text}")
                                                schedule[lesson_num] = {
                                                    'subject': '❌ Пара отменена',  # Добавляем эмодзи прямо в subject
                                                    'teacher': teacher_name,
                                                    'room': '',
                                                    'is_common': False,
                                                    'subgroup': 1,
                                                    'group': group_name,
                                                    'is_replacement': True,
                                                    'is_cancelled': True,
                                                    'emoji': '❌'  # Добавляем отдельное поле для эмодзи
                                                }
                                                logger.info(f"!!! ДОБАВЛЕНА ОТМЕНА в расписание для пары {lesson_num}, подгруппа 1")
                                                continue
                                                
                                            # Разбиваем на части по строкам для обработки подгрупп
                                            subgroup_parts = replacement_text.split('\n')
                                            processed = False
                                            
                                            # Собираем информацию для подгрупп
                                            subgroup1_text = ""
                                            subgroup2_text = ""
                                            current_subgroup = None
                                            
                                            # Определяем, к какой подгруппе относится текст
                                            for part in subgroup_parts:
                                                part = part.strip()
                                                if not part:
                                                    continue
                                                
                                                # Если нашли маркер первой подгруппы
                                                if part.startswith('1. ') or '1 п/г' in part:
                                                    current_subgroup = 1
                                                    # Проверяем на отмену для подгруппы 1
                                                    if part.startswith('1. ') and set(part[2:].strip('- ')).issubset({'-', ' '}):
                                                        logger.info(f"Обнаружена отмена пары {lesson_num} для подгруппы 1 преподавателя {teacher_name} в группе {group_name}")
                                                        if teacher_name.lower() in '\n'.join(subgroup_parts).lower():
                                                            # Проверяем, что этот преподаватель ведет именно эту подгруппу
                                                            # Ищем, находится ли имя преподавателя между "1." и "2."
                                                            text_until_subgroup2 = ""
                                                            found_sg1 = False
                                                            for p in subgroup_parts:
                                                                if p.strip().startswith('2. '):
                                                                    break
                                                                if found_sg1:
                                                                    text_until_subgroup2 += p + '\n'
                                                                if p.strip().startswith('1. '):
                                                                    found_sg1 = True
                                                                    text_until_subgroup2 += p + '\n'
                                                            
                                                            if text_until_subgroup2 and teacher_name.lower() in text_until_subgroup2.lower():
                                                                schedule[lesson_num] = {
                                                                    'subject': '❌ Пара отменена',
                                                                    'teacher': teacher_name,
                                                                    'room': '',
                                                                    'is_common': False,
                                                                    'subgroup': 1,
                                                                    'group': group_name,
                                                                    'is_replacement': True,
                                                                    'is_cancelled': True,
                                                                    'emoji': '❌'
                                                                }
                                                                processed = True
                                                                break
                                                    subgroup1_text += part + '\n'
                                                # Если нашли маркер второй подгруппы
                                                elif part.startswith('2. ') or '2 п/г' in part:
                                                    current_subgroup = 2
                                                    # Проверяем на отмену для подгруппы 2
                                                    if part.startswith('2. ') and set(part[2:].strip('- ')).issubset({'-', ' '}):
                                                        logger.info(f"Обнаружена отмена пары {lesson_num} для подгруппы 2 преподавателя {teacher_name} в группе {group_name}")
                                                        if teacher_name.lower() in '\n'.join(subgroup_parts).lower():
                                                            # Проверяем, что этот преподаватель ведет именно эту подгруппу
                                                            # Ищем, находится ли имя преподавателя после "2."
                                                            text_after_subgroup2 = ""
                                                            found_sg2 = False
                                                            for p in subgroup_parts:
                                                                if found_sg2:
                                                                    text_after_subgroup2 += p + '\n'
                                                                if p.strip().startswith('2. '):
                                                                    found_sg2 = True
                                                                    text_after_subgroup2 += p + '\n'
                                                            
                                                            if text_after_subgroup2 and teacher_name.lower() in text_after_subgroup2.lower():
                                                                schedule[lesson_num] = {
                                                                    'subject': '❌ Пара отменена',
                                                                    'teacher': teacher_name,
                                                                    'room': '',
                                                                    'is_common': False,
                                                                    'subgroup': 2,
                                                                    'group': group_name,
                                                                    'is_replacement': True,
                                                                    'is_cancelled': True,
                                                                    'emoji': '❌'
                                                                }
                                                                processed = True
                                                                break
                                                    subgroup2_text += part + '\n'
                                                # Если это дополнительный текст для текущей подгруппы
                                                elif current_subgroup == 1:
                                                    subgroup1_text += part + '\n'
                                                elif current_subgroup == 2:
                                                    subgroup2_text += part + '\n'
                                                else:
                                                    # Если не указана подгруппа, добавляем к обоим
                                                    subgroup1_text += part + '\n'
                                                    subgroup2_text += part + '\n'
                                            
                                            # Если учитель найден после маркера "2."
                                            if not processed and teacher_name.lower() in subgroup2_text.lower():
                                                room_match = re.search(r'[АA]\d{3,4}', subgroup2_text)
                                                room = room_match.group(0) if room_match else ''
                                                
                                                # Очищаем текст для второй подгруппы
                                                clean_text = subgroup2_text.replace(teacher_name, '').strip()
                                                if room:
                                                    clean_text = clean_text.replace(room, '').strip()
                                                
                                                # Извлекаем название предмета для 2-й подгруппы
                                                subject_match = re.search(r'2\.\s+\(.*?\)(.*?)(?=[АA]\d{3,4}|\n|$)', subgroup2_text)
                                                if subject_match:
                                                    clean_text = subject_match.group(1).strip()
                                                else:
                                                    # Если нет скобок с типом занятия, пробуем другой формат
                                                    subject_match = re.search(r'2\.\s+(.*?)(?=[АA]\d{3,4}|\n|$)', subgroup2_text)
                                                    if subject_match:
                                                        clean_text = subject_match.group(1).strip()
                                                    else:
                                                        # Удаляем маркер "2."
                                                        clean_text = re.sub(r'^2\.\s*', '', clean_text)
                                                
                                                # Если текст пустой или содержит только дефисы (отмена), используем запасной вариант
                                                if not clean_text or clean_text.isspace() or set(clean_text.strip('- ')).issubset({'-', ' '}):
                                                    # Проверяем, есть ли название предмета в строке с маркером подгруппы
                                                    for part in subgroup_parts:
                                                        if part.startswith('2. ') and not set(part[2:].strip('- ')).issubset({'-', ' '}):
                                                            clean_text = re.sub(r'^2\.\s*', '', part).strip()
                                                            break
                                                    
                                                    # Если все еще пусто, ищем название предмета в любой строке после "2."
                                                    if not clean_text or clean_text.isspace() or set(clean_text.strip('- ')).issubset({'-', ' '}):
                                                        found_sg2 = False
                                                        for part in subgroup_parts:
                                                            if part.startswith('2. '):
                                                                found_sg2 = True
                                                                continue
                                                            if found_sg2 and '(' in part and ')' in part and not teacher_name.lower() in part.lower():
                                                                clean_text = part.strip()
                                                                break
                                                    
                                                    # Если все еще пусто, используем запасной вариант
                                                    if not clean_text or clean_text.isspace() or set(clean_text.strip('- ')).issubset({'-', ' '}):
                                                        clean_text = 'Ин.яз (проф.)'  # Более осмысленное значение по умолчанию
                                                
                                                logger.info(f"Обработана замена для 2-й подгруппы: предмет='{clean_text}', аудитория='{room}'")
                                                
                                                schedule[lesson_num] = {
                                                    'subject': f"✏️ {clean_text}",
                                                    'teacher': teacher_name,
                                                    'room': room,
                                                    'is_common': False,
                                                    'subgroup': 2,
                                                    'group': group_name,
                                                    'is_replacement': True,
                                                    'emoji': '✏️'
                                                }
                                                processed = True
                                            
                                            # Если учитель найден после маркера "1." и до маркера "2."
                                            elif not processed and teacher_name.lower() in subgroup1_text.lower():
                                                room_match = re.search(r'[АA]\d{3,4}', subgroup1_text)
                                                room = room_match.group(0) if room_match else ''
                                                
                                                # Очищаем текст для первой подгруппы
                                                clean_text = subgroup1_text.replace(teacher_name, '').strip()
                                                if room:
                                                    clean_text = clean_text.replace(room, '').strip()
                                                
                                                # Извлекаем название предмета для 1-й подгруппы
                                                subject_match = re.search(r'1\.\s+\(.*?\)(.*?)(?=[АA]\d{3,4}|\n|$)', subgroup1_text)
                                                if subject_match:
                                                    clean_text = subject_match.group(1).strip()
                                                else:
                                                    # Если нет скобок с типом занятия, пробуем другой формат
                                                    subject_match = re.search(r'1\.\s+(.*?)(?=[АA]\d{3,4}|\n|$)', subgroup1_text)
                                                    if subject_match:
                                                        clean_text = subject_match.group(1).strip()
                                                    else:
                                                        # Удаляем маркер "1."
                                                        clean_text = re.sub(r'^1\.\s*', '', clean_text)
                                                
                                                # Если текст пустой или содержит только дефисы (отмена), используем запасной вариант
                                                if not clean_text or clean_text.isspace() or set(clean_text.strip('- ')).issubset({'-', ' '}):
                                                    # Проверяем, есть ли название предмета в строке с маркером подгруппы
                                                    for part in subgroup_parts:
                                                        if part.startswith('1. ') and not set(part[2:].strip('- ')).issubset({'-', ' '}):
                                                            clean_text = re.sub(r'^1\.\s*', '', part).strip()
                                                            break
                                                    
                                                    # Если все еще пусто, ищем название предмета в любой строке до "2."
                                                    if not clean_text or clean_text.isspace() or set(clean_text.strip('- ')).issubset({'-', ' '}):
                                                        for part in subgroup_parts:
                                                            if part.startswith('2. '):
                                                                break
                                                            if '(' in part and ')' in part and not teacher_name.lower() in part.lower():
                                                                clean_text = part.strip()
                                                                break
                                                    
                                                    # Если все еще пусто, используем запасной вариант
                                                    if not clean_text or clean_text.isspace() or set(clean_text.strip('- ')).issubset({'-', ' '}):
                                                        clean_text = 'Ин.яз (проф.)'  # Более осмысленное значение по умолчанию
                                                
                                                logger.info(f"Обработана замена для 1-й подгруппы: предмет='{clean_text}', аудитория='{room}'")
                                                
                                                schedule[lesson_num] = {
                                                    'subject': f"✏️ {clean_text}",
                                                    'teacher': teacher_name,
                                                    'room': room,
                                                    'is_common': False,
                                                    'subgroup': 1,
                                                    'group': group_name,
                                                    'is_replacement': True,
                                                    'emoji': '✏️'
                                                }
                                                processed = True
                                            
                                            # Если не обработали по частям, обрабатываем всю замену целиком
                                            if not processed:
                                                # Извлекаем номер аудитории из текста замены
                                                room_match = re.search(r'[АA]\d{3,4}', replacement_text)
                                                room = room_match.group(0) if room_match else ''
                                                
                                                # Очищаем замену от имени преподавателя и номера аудитории
                                                clean_text = replacement_text
                                                if room:
                                                    clean_text = clean_text.replace(room, '')
                                                
                                                # Удаляем имя преподавателя
                                                clean_text = re.sub(r'(?i)' + re.escape(teacher_name), '', clean_text).strip()
                                                
                                                # Проверяем, есть ли в тексте что-то кроме служебных символов
                                                if not clean_text or clean_text.isspace():
                                                    # Пытаемся извлечь из оригинального текста
                                                    original_text = replacement_text.split('\n')[0] if '\n' in replacement_text else replacement_text
                                                    clean_text = original_text.replace(room, '').replace(teacher_name, '').strip()
                                                
                                                logger.info(f"Обработана замена целиком: предмет='{clean_text}', аудитория='{room}'")
                                                
                                                # Добавляем замену в расписание
                                                schedule[lesson_num] = {
                                                    'subject': clean_text or '(ТО) МДК.01.03 Разработка мобил.прил.',
                                                    'teacher': teacher_name,
                                                    'room': room,
                                                    'is_common': True,  # Замены обычно общие для всей группы
                                                    'subgroup': None,
                                                    'group': group_name,
                                                    'is_replacement': True,
                                                    'is_cancelled': False,
                                                    'emoji': None  # Добавляем поле для эмодзи, но оставляем его пустым для обычных пар
                                                }
                                            
                                            continue
                                        
                                        # Если у преподавателя есть пара в это время
                                        elif lesson_num in teacher_lessons:
                                            logger.info(f"У преподавателя {teacher_name} есть пара №{lesson_num} в группе {group_name}")
                                            
                                            # Если есть замена, но преподаватель не упоминается - проверяем на отмену
                                            if replacement and teacher_name.lower() not in str(replacement).lower():
                                                replacement_text = str(replacement)
                                                logger.info(f"Проверка замены без упоминания преподавателя: '{replacement_text}'")
                                                
                                                # Проверяем, является ли замена отменой для всей группы
                                                if is_cancellation_text(replacement_text):
                                                    logger.info(f"Обнаружена отмена пары {lesson_num} для преподавателя {teacher_name} в группе {group_name} (без упоминания преподавателя)")
                                                    schedule[lesson_num] = {
                                                        'subject': '❌ Пара отменена',  # Добавляем эмодзи прямо в subject
                                                        'teacher': teacher_name,
                                                        'room': '',
                                                        'is_common': False,
                                                        'subgroup': teacher_lessons[lesson_num].get('subgroup'),
                                                        'group': group_name,
                                                        'is_replacement': True,
                                                        'is_cancelled': True,
                                                        'emoji': '❌'  # Добавляем отдельное поле для эмодзи
                                                    }
                                                    continue
                                                
                                                # Проверяем на специальные форматы отмены для подгрупп
                                                subgroup_parts = replacement_text.split('\n')
                                                cancellation_found = False
                                                
                                                for part in subgroup_parts:
                                                    part = part.strip()
                                                    if not part:
                                                        continue
                                                    
                                                    # Проверка на "1. ------------" для отмены 1-й подгруппы
                                                    if (part.startswith('1. ') or part.startswith('1. ')) and set(part[2:].strip('- ')).issubset({'-', ' '}):
                                                        subgroup = 1
                                                        # Проверяем, занимается ли преподаватель с этой подгруппой
                                                        current_subgroup = teacher_lessons[lesson_num].get('subgroup')
                                                        
                                                        # Важно: если это КП класс и у преподавателя указана 2-я подгруппа,
                                                        # то отмена для 1-й подгруппы не должна применяться ко 2-й
                                                        if '(КП)' in str(teacher_lessons[lesson_num].get('subject', '')) and current_subgroup == 2:
                                                            logger.info(f"Пропускаем отмену для 1-й подгруппы, т.к. преподаватель ведет КП в 2-й подгруппе")
                                                            continue
                                                            
                                                        if current_subgroup == 1 or current_subgroup is None:
                                                            logger.info(f"!!! Обнаружена отмена по шаблону '1. ------------' для пары {lesson_num}, подгруппа 1")
                                                            schedule[lesson_num] = {
                                                                'subject': '❌ Пара отменена',
                                                                'teacher': teacher_name,
                                                                'room': '',
                                                                'is_common': False,
                                                                'subgroup': 1,
                                                                'group': group_name,
                                                                'is_replacement': True,
                                                                'is_cancelled': True,
                                                                'emoji': '❌'
                                                            }
                                                            cancellation_found = True
                                                            break
                                                    
                                                    # Проверка на "2. ------------" для отмены 2-й подгруппы
                                                    elif (part.startswith('2. ') or part.startswith('2. ')) and set(part[2:].strip('- ')).issubset({'-', ' '}):
                                                        subgroup = 2
                                                        # Проверяем, занимается ли преподаватель с этой подгруппой
                                                        current_subgroup = teacher_lessons[lesson_num].get('subgroup')
                                                        
                                                        # Важно: если это КП класс и у преподавателя указана 1-я подгруппа,
                                                        # то отмена для 2-й подгруппы не должна применяться к 1-й
                                                        if '(КП)' in str(teacher_lessons[lesson_num].get('subject', '')) and current_subgroup == 1:
                                                            logger.info(f"Пропускаем отмену для 2-й подгруппы, т.к. преподаватель ведет КП в 1-й подгруппе")
                                                            continue
                                                            
                                                        if current_subgroup == 2 or current_subgroup is None:
                                                            logger.info(f"!!! Обнаружена отмена по шаблону '2. ------------' для пары {lesson_num}, подгруппа 2")
                                                            schedule[lesson_num] = {
                                                                'subject': '❌ Пара отменена',
                                                                'teacher': teacher_name,
                                                                'room': '',
                                                                'is_common': False,
                                                                'subgroup': 2,
                                                                'group': group_name,
                                                                'is_replacement': True,
                                                                'is_cancelled': True,
                                                                'emoji': '❌'
                                                            }
                                                            cancellation_found = True
                                                            break
                                                
                                                if cancellation_found:
                                                    continue
                                                
                                                # Проверяем, указан ли другой преподаватель в этой группе на это время (замена преподавателя)
                                                elif replacement:
                                                    replacement_text = str(replacement)
                                                    logger.info(f"Проверка замены на другого преподавателя: '{replacement_text}'")
                                                    
                                                    # Разбиваем текст замены на строки для анализа
                                                    subgroup_parts = replacement_text.split('\n')
                                                    
                                                    # Собираем информацию для подгрупп
                                                    subgroup1_info = {'text': '', 'has_teacher': False}
                                                    subgroup2_info = {'text': '', 'has_teacher': False}
                                                    current_subgroup = None
                                                    
                                                    # Определяем, к какой подгруппе относится текст
                                                    for part in subgroup_parts:
                                                        part = part.strip()
                                                        if not part:
                                                            continue
                                                        
                                                        # Определяем маркеры подгрупп
                                                        if part.startswith('1. '):
                                                            current_subgroup = 1
                                                            subgroup1_info['text'] += part + '\n'
                                                        elif part.startswith('2. '):
                                                            current_subgroup = 2
                                                            subgroup2_info['text'] += part + '\n'
                                                        elif current_subgroup == 1:
                                                            subgroup1_info['text'] += part + '\n'
                                                            if re.search(r'[А-Я][а-я]+\s+[А-Я]\.[А-Я]\.', part):
                                                                subgroup1_info['has_teacher'] = True
                                                        elif current_subgroup == 2:
                                                            subgroup2_info['text'] += part + '\n'
                                                            if re.search(r'[А-Я][а-я]+\s+[А-Я]\.[А-Я]\.', part):
                                                                subgroup2_info['has_teacher'] = True
                                                    
                                                    # Проверяем совпадение подгрупп
                                                    current_teacher_subgroup = teacher_lessons[lesson_num].get('subgroup')
                                                    
                                                    # Если преподаватель ведет 1-ю подгруппу и в замене указан другой преподаватель для 1-й подгруппы
                                                    if (current_teacher_subgroup == 1 or current_teacher_subgroup is None) and \
                                                       subgroup1_info['has_teacher'] and teacher_name.lower() not in subgroup1_info['text'].lower():
                                                        # Проверяем, есть ли преподаватель во второй подгруппе (переназначение)
                                                        if teacher_name.lower() in subgroup2_info['text'].lower():
                                                            # Преподаватель переназначен на другую подгруппу - не отменяем
                                                            pass
                                                        else:
                                                            # Занятие отменено или заменено другим преподавателем
                                                            schedule[lesson_num] = {
                                                                'subject': '❌ Пара отменена',
                                                                'teacher': teacher_name,
                                                                'room': '',
                                                                'is_common': False,
                                                                'subgroup': 1,
                                                                'group': group_name,
                                                                'is_replacement': True,
                                                                'is_cancelled': True,
                                                                'emoji': '❌'
                                                            }
                                                            logger.info(f"Пара отменена для преподавателя {teacher_name}, занятие ведет другой преподаватель")
                                                            continue
                                                    
                                                    # Если преподаватель ведет 2-ю подгруппу и в замене указан другой преподаватель для 2-й подгруппы
                                                    if (current_teacher_subgroup == 2 or current_teacher_subgroup is None) and \
                                                       subgroup2_info['has_teacher'] and teacher_name.lower() not in subgroup2_info['text'].lower():
                                                        # Проверяем, есть ли преподаватель в первой подгруппе (переназначение)
                                                        if teacher_name.lower() in subgroup1_info['text'].lower():
                                                            # Преподаватель переназначен на другую подгруппу - не отменяем
                                                            pass
                                                        else:
                                                            # Занятие отменено или заменено другим преподавателем
                                                            schedule[lesson_num] = {
                                                                'subject': '❌ Пара отменена',
                                                                'teacher': teacher_name,
                                                                'room': '',
                                                                'is_common': False,
                                                                'subgroup': 2,
                                                                'group': group_name,
                                                                'is_replacement': True,
                                                                'is_cancelled': True,
                                                                'emoji': '❌'
                                                            }
                                                            logger.info(f"Пара отменена для преподавателя {teacher_name}, занятие ведет другой преподаватель")
                                                            continue
            except Exception as e:
                logger.error(f"Ошибка при обработке файла замен {replacement_file}: {str(e)}")

        # Добавляем обычные пары преподавателя, если они не были заменены или отменены
        for lesson_num, lesson_info in teacher_lessons.items():
            # Проверяем, была ли пара обработана в файле замен
            lesson_in_replacements = lesson_num in replacement_applied
            
            # Если пара не в расписании и либо не было файла замен для этой группы, либо для этой пары не применялась замена
            if lesson_num not in schedule:
                schedule[lesson_num] = {
                    'subject': lesson_info['subject'],
                    'teacher': lesson_info['teacher'],
                    'room': lesson_info['room'],
                    'is_common': lesson_info['is_common'],
                    'subgroup': lesson_info['subgroup'],
                    'group': group_name,
                    'is_replacement': False
                }
            # Если пара в расписании и она помечена как отмененная, но в файле замен не было информации для этой пары - убираем отмену
            elif lesson_num in schedule and schedule[lesson_num].get('is_cancelled') and not lesson_in_replacements:
                schedule[lesson_num] = {
                    'subject': lesson_info['subject'],
                    'teacher': lesson_info['teacher'],
                    'room': lesson_info['room'],
                    'is_common': lesson_info['is_common'],
                    'subgroup': lesson_info['subgroup'],
                    'group': group_name,
                    'is_replacement': False
                }
        
        # Ищем новые пары в замене, которых нет в обычном расписании
        new_lessons = find_new_lessons_in_replacements()
        
        # Добавляем новые пары из замен
        for lesson_num, lesson_info in new_lessons.items():
            if lesson_num not in schedule:
                schedule[lesson_num] = lesson_info
                logger.info(f"Добавлена новая пара из замены: {lesson_num}, {lesson_info['subject']}")
        
        logger.info(f"Итоговое расписание для {teacher_name} на {date_str}: {schedule}")
        return schedule

    except Exception as e:
        logger.error(f"Ошибка при парсинге расписания: {str(e)}")
        return {}


def format_teacher_schedule(schedule_data, teacher_name, start_date, end_date):
    """Форматирует расписание преподавателя"""
    try:
        # Заголовок с эмодзи и именем преподавателя
        formatted = [f"📅 Расписание преподавателя {teacher_name}"]
        
        # Добавляем период расписания
        formatted.append(f"Период: с {start_date} по {end_date}")
        
        # Если нет данных расписания, возвращаем сообщение
        if not schedule_data:
            return "\n".join(formatted) + "\n\nРасписание на указанный период не найдено"
        
        # Sort dates chronologically
        sorted_dates = sorted(schedule_data.keys(), 
                             key=lambda date_str: datetime.strptime(date_str, '%d.%m.%Y'))
        
        # Use the sorted dates when iterating
        for date_str in sorted_dates:
            day_schedule = schedule_data[date_str]
            
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            weekday = days_ru[date_obj.weekday()]
            week_type = get_week_type(date_str)
            
            # Добавляем день недели с датой и типом недели
            formatted.append(f"\nРасписание на {weekday} {date_str} ({week_type} неделя):")
            
            # Если на этот день нет пар
            if not day_schedule:
                formatted.append("В этот день пар нет")
                continue
            
            # Сортируем пары по номеру
            for lesson_num in sorted(day_schedule.keys(), key=lambda x: int(str(x).split('.')[0])):
                lesson = day_schedule[lesson_num]
                
                # Базовый номер пары с эмодзи
                lesson_str = f"{str(lesson_num)}️⃣"
                
                # Получаем текст предмета
                subject_text = lesson.get('subject', '')
                # Удаляем маркер замены если он есть
                if subject_text.startswith('✏️ '):
                    subject_text = subject_text[2:].strip()
                
                # Удаляем маркеры подгрупп из текста
                subject_text = subject_text.replace('1. ', '').replace('2. ', '').strip()
                
                # Информация о группе и подгруппе
                group_name = lesson.get('group', '')
                subgroup_info = ""
                if lesson.get('subgroup') and not lesson.get('is_common'):
                    subgroup_info = f", {lesson['subgroup']}-я подгруппа"
                
                # Обрабатываем отмененные пары
                if lesson.get('is_cancelled') or subject_text.startswith('❌'):
                    lesson_str += " ❌ Пара отменена"
                    if group_name:
                        lesson_str += f" 🚪 [{group_name}{subgroup_info}]"
                else:
                    # Добавляем маркер замены если нужно
                    if lesson.get('is_replacement'):
                        lesson_str += " ✏️"
                    
                    # Добавляем тип пары и название предмета
                    lesson_str += f" ({subject_text})"
                    
                    # Добавляем аудиторию
                    room = lesson.get('room')
                    if room and room not in ['None', None, '']:
                        lesson_str += f" 🚪{room}"
                    
                    # Добавляем группу и подгруппу
                    if group_name:
                        # Преобразуем название группы к единому формату
                        group_name = group_name[0].upper() + group_name[1:] if group_name else ""
                        lesson_str += f" [{group_name}{subgroup_info}]"
                
                formatted.append(lesson_str)

        return "\n".join(formatted)

    except Exception as e:
        logger.error(f"Ошибка при форматировании расписания преподавателя: {str(e)}")
        return "Ошибка при форматировании расписания"

def parse_schedule(file_path, date_str, selected_subgroup=None):
    try:
        logger.info(f"Parsing schedule for date: {date_str}, subgroup: {selected_subgroup}")

        local_path = os.path.join("downloaded_files", os.path.basename(file_path))
        if not os.path.exists(local_path):
            return {}

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        is_even_week = get_week_type(date_str) == 'четная'

        days = {
            0: 'понедельник',
            1: 'вторник',
            2: 'среда',
            3: 'четверг',
            4: 'пятница',
            5: 'суббота'
        }

        day_name = days[datetime.strptime(date_str, '%d.%m.%Y').weekday()]
        day_col, current_start_row = find_day_column(sheet, day_name, 3, is_even_week)

        if not day_col:
            return {}

        schedule = {}
        current_row = current_start_row + 1

        while current_row < current_start_row + 20:
            lesson_num = sheet.cell(row=current_row, column=day_col).value
            if not lesson_num:
                break

            subject_first = sheet.cell(row=current_row, column=day_col + 1).value
            subject_second = sheet.cell(row=current_row, column=day_col + 3).value
            room_first = sheet.cell(row=current_row + 1, column=day_col + 2).value
            room_second = sheet.cell(row=current_row + 1, column=day_col + 4).value

            if is_theory_lesson(subject_first):
                # Общая теоретическая пара для всей группы
                schedule[lesson_num] = {
                    'subject': f"{subject_first}" if subject_first else subject_first,
                    'teacher': sheet.cell(row=current_row + 1, column=day_col + 1).value,
                    'room': room_second,  # Для теории всегда берем правую аудиторию
                    'is_common': True,
                    'subgroup': None
                }
            else:
                # Проверяем практические занятия
                # Проверяем практические занятия и КП
                is_first_practical = subject_first and '(пр)' in str(subject_first).lower()
                is_second_practical = subject_second and '(пр)' in str(subject_second).lower()
                is_first_kp = subject_first and '(кп)' in str(subject_first).lower()
                is_second_kp = subject_second and '(кп)' in str(subject_second).lower()

                # Проверка на общую практическую пару
                # Проверка на общую практическую пару
                is_common_practice = False
                if is_first_practical:
                    if not subject_second:  # Нет пары во второй подгруппе
                        is_common_practice = bool(room_second)  # Общая, если есть правая аудитория
                    elif subject_second == subject_first:  # Одинаковые предметы в обеих подгруппах
                        is_common_practice = True

                # Проверка на общую пару КП
                is_common_kp = False
                if is_first_kp:
                    if not subject_second:  # Нет пары во второй подгруппе
                        is_common_kp = bool(room_second)  # Общая, если есть правая аудитория
                    elif subject_second == subject_first:  # Одинаковые предметы в обеих подгруппах
                        is_common_kp = True        

                if is_common_practice:
                    # Общая практическая пара
                    schedule[lesson_num] = {
                        'subject': subject_first,
                        'teacher': sheet.cell(row=current_row + 1, column=day_col + 1).value,
                        'room': room_second,  # Для общей пары берем правую аудиторию
                        'is_common': True,
                        'subgroup': None
                    }
                else:
                    # Пары по подгруппам
                    if (selected_subgroup is None or selected_subgroup == 1) and subject_first:
                        # Используем уникальный ключ для первой подгруппы
                        lesson_key = f"{lesson_num}_1" if selected_subgroup is None else lesson_num
                        schedule[lesson_key] = {
                            'subject': subject_first,
                            'teacher': sheet.cell(row=current_row + 1, column=day_col + 1).value,
                            'room': room_first if room_first else room_second,
                            'is_common': False,
                            'subgroup': 1,
                            'original_num': lesson_num  # Сохраняем оригинальный номер пары
                        }

                    if (selected_subgroup is None or selected_subgroup == 2) and subject_second:
                        # Используем уникальный ключ для второй подгруппы
                        lesson_key = f"{lesson_num}_2" if selected_subgroup is None else lesson_num
                        schedule[lesson_key] = {
                            'subject': subject_second,
                            'teacher': sheet.cell(row=current_row + 1, column=day_col + 3).value,
                            'room': room_second,
                            'is_common': False,
                            'subgroup': 2,
                            'original_num': lesson_num  # Сохраняем оригинальный номер пары
                        }

            current_row += 2
            logger.info(f"Processed lesson {lesson_num}: {schedule.get(lesson_num)}")

        return schedule

    except Exception as e:
        logger.error(f"Ошибка при парсинге расписания: {str(e)}")
        return {}


def process_schedule_files(group: str, subgroup: int = None) -> str:
    """Обрабатывает файлы расписания для группы"""
    try:
        # Ищем файл расписания для конкретной группы
        schedule_file = None
        for file in os.listdir("downloaded_files"):
            if file.startswith(group) and file.endswith('.xlsx'):
                schedule_file = file
                break

        if not schedule_file:
            return f"Расписание для группы {group} не найдено."

        all_schedules = {}
        dates_to_check = set()
        today = datetime.now().date()

        # Находим все файлы замен
        replacement_files = [f for f in os.listdir("downloaded_files")
                             if f.endswith('.xlsx') and '-' in f]

        # Собираем все даты из файлов замен
        for replacement_file in replacement_files:
            try:
                dates = replacement_file.replace('.xlsx', '').split('-')
                if len(dates) == 2:
                    start_date = datetime.strptime(dates[0], '%d.%m.%y').date()
                    end_date = datetime.strptime(dates[1], '%d.%m.%y').date()

                    current_date = max(today, start_date)
                    while current_date <= end_date:
                        if current_date.weekday() != 6:  # Пропускаем воскресенье
                            dates_to_check.add(current_date.strftime('%d.%m.%Y'))
                        current_date += timedelta(days=1)
            except Exception as e:
                logger.error(f"Ошибка при обработке файла замен {replacement_file}: {str(e)}")
                continue

        if not dates_to_check:
            return "Не найдены актуальные даты в файлах замен."

        # Обрабатываем расписание для каждой даты
        local_path = os.path.join("downloaded_files", schedule_file)

        # Проверяем кэш
        cache_key = f"{schedule_file}:{group}_{subgroup}"
        cached_schedule = get_cached_schedule(local_path, cache_key, group)

        if cached_schedule is not None:
            all_schedules.update(cached_schedule)
        else:
            # Если нет в кэше, парсим расписание для каждой даты
            for date_str in sorted(dates_to_check):
                schedule = parse_schedule(local_path, date_str, subgroup)

                # Проверяем замены
                for replacement_file in replacement_files:
                    try:
                        replacement_path = os.path.join("downloaded_files", replacement_file)
                        temp_schedule = process_schedule_with_replacements(
                            local_path,
                            replacement_path,
                            date_str,
                            subgroup
                        )
                        if temp_schedule:
                            schedule = temp_schedule
                    except Exception as e:
                        logger.error(f"Ошибка при обработке замен {replacement_file}: {str(e)}")
                        continue

                if schedule:
                    all_schedules[date_str] = schedule

            # Кэшируем результат
            cache_schedule(local_path, cache_key, group, all_schedules)

        if not all_schedules:
            return f"Расписание для группы {group} не найдено."

        # Форматируем расписание
        formatted_schedule = []
        for date_str in sorted(all_schedules.keys()):
            schedule = all_schedules[date_str]
            if schedule:
                date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                week_type = get_week_type(date_str)
                weekday_ru = days_ru.get(date_obj.weekday(), str(date_obj.weekday()))

                schedule_text = f"\n📅 Расписание на {weekday_ru} {date_str} ({week_type} неделя):"

                if isinstance(schedule, dict) and schedule:
                    for lesson_num in sorted(schedule.keys(), key=lambda x: float(str(x).replace(',', '.'))):
                        lesson = schedule[lesson_num]
                        schedule_text += f"\n{lesson_num}️⃣ "

                        if isinstance(lesson, dict):
                            if lesson.get('status') == 'cancelled':
                                schedule_text += "ОТМЕНЕНА"
                            else:
                                schedule_text += f"{lesson.get('subject', '')} "
                                if lesson.get('room'):
                                    schedule_text += f"🚪{lesson['room']} "
                                if not lesson.get('is_common', False):
                                    schedule_text += f" [{subgroup}-я подгруппа]"
                        else:
                            schedule_text += str(lesson)

                    formatted_schedule.append(schedule_text)

        return "\n\n".join(formatted_schedule) if formatted_schedule else "Расписание не найдено"

    except Exception as e:
        logger.error(f"Ошибка при обработке расписания: {str(e)}")
        return f"Произошла ошибка при получении расписания: {str(e)}"


def get_replacements_file(date_str):
    """Определяет файл с заменами для указанной даты"""
    try:
        target_date = datetime.strptime(date_str, '%d.%m.%Y')
        files = [f for f in os.listdir("downloaded_files") if f.endswith('.xlsx') and f[0].isdigit() and '-' in f]

        for file in files:
            try:
                # Извлекаем даты из имени файла (форматы: dd.mm.yy-dd.mm.yy.xlsx или dd.mm.yyyy-dd.mm.yyyy.xlsx)
                dates = file.replace('.xlsx', '').split('-')
                if len(dates) != 2:
                    continue

                start_str = dates[0]
                end_str = dates[1]
                
                # Определяем формат даты и преобразуем строки в объекты datetime
                try:
                    # Пробуем сначала формат с двузначным годом
                    start_date = datetime.strptime(start_str, '%d.%m.%y')
                    end_date = datetime.strptime(end_str, '%d.%m.%y')
                except ValueError:
                    try:
                        # Пробуем формат с четырехзначным годом
                        start_date = datetime.strptime(start_str, '%d.%m.%Y')
                        end_date = datetime.strptime(end_str, '%d.%m.%Y')
                    except ValueError:
                        # Если не удалось распознать формат, пропускаем файл
                        logger.warning(f"Не удалось распознать формат даты в файле: {file}")
                        continue

                # Проверяем, входит ли целевая дата в диапазон
                if start_date <= target_date <= end_date:
                    logger.info(f"Найден файл замен для даты {date_str}: {file}")
                    return os.path.join("downloaded_files", file)
            except Exception as e:
                logger.error(f"Ошибка при обработке файла {file}: {str(e)}")
                continue

        logger.warning(f"Файл с заменами для даты {date_str} не найден")
        return None

    except Exception as e:
        logger.error(f"Ошибка при поиске файла замен: {str(e)}")
        return None


def process_schedule_with_replacements(schedule_file, replacements_file, date_str, selected_subgroup=None):
    try:
        logger.info(f"Обработка замен для даты {date_str}, подгруппа {selected_subgroup}")

        # Определяем правильный файл с заменами для даты
        actual_replacements_file = get_replacements_file(date_str)
        if actual_replacements_file:
            replacements = load_replacements(actual_replacements_file)
        else:
            replacements = {}
        logger.info(f"Загруженные замены: {replacements.get(date_str, {})}")

        schedule = parse_schedule(schedule_file, date_str, selected_subgroup)
        logger.info(f"Базовое расписание: {schedule}")

        wb = openpyxl.load_workbook(schedule_file)
        sheet = wb.active
        group_name = str(sheet.cell(row=1, column=1).value or '').split('группы ')[-1].strip()

        if date_str in replacements and group_name in replacements[date_str]:
            group_replacements = replacements[date_str][group_name]
            logger.info(f"Найдены замены для группы {group_name}: {group_replacements}")

            for lesson_num, replacement_data in group_replacements.items():
                logger.info(f"Обработка замены для пары {lesson_num}: {replacement_data}")

                # Если есть замены для разных подгрупп
                if isinstance(replacement_data, dict) and any(isinstance(k, int) for k in replacement_data.keys()):
                    # Для отладки пары 5
                    if int(lesson_num) == 5:
                        logger.info(f"Обработка замены для пары 5. Данные: {replacement_data}")
                    
                    for subgroup, replacement in replacement_data.items():
                        if selected_subgroup is None or selected_subgroup == subgroup:
                            # Создаем уникальный ключ для пары с учетом подгруппы
                            lesson_key = f"{lesson_num}_{subgroup}" if selected_subgroup is None else lesson_num

                            # Если это пара 5, специальная обработка 
                            if int(lesson_num) == 5:
                                # Добавляем необходимые поля
                                replacement['is_replacement'] = True
                                replacement['original_num'] = lesson_num
                                replacement['group'] = group_name
                                
                                # Отметка если это отмена
                                if replacement.get('status') == 'cancelled' or replacement.get('is_cancelled'):
                                    replacement['is_cancelled'] = True
                                    replacement['emoji'] = '❌'
                                
                                logger.info(f"Пара 5, подгруппа {subgroup}, ключ {lesson_key}, итоговая замена: {replacement}")
                            else:
                                # Обычная обработка для других пар
                                # Сохраняем оригинальный номер пары
                                replacement['original_num'] = lesson_num
                                replacement['is_replacement'] = True
                            
                            schedule[lesson_key] = replacement
                            logger.info(f"Применена замена для пары {lesson_num} подгруппы {subgroup}, ключ {lesson_key}")
                else:
                    # Если замена для всей группы
                    replacement_subgroup = replacement_data.get('subgroup')
                    if selected_subgroup is None or replacement_subgroup is None or replacement_subgroup == selected_subgroup:
                        schedule[lesson_num] = replacement_data
                        logger.info(f"Применена замена для пары {lesson_num}")

        return schedule

    except Exception as e:
        logger.error(f"Ошибка при обработке замен: {str(e)}")
        return schedule


def load_replacements(replacements_file):
    """Загружает замены из файла"""
    try:
        wb = openpyxl.load_workbook(replacements_file)
        sheet = wb.active
        replacements = {}

        logger.info(f"Начало загрузки замен из файла {replacements_file}")

        # Получаем словарь групп из строки 2
        groups = {}
        for col in range(4, sheet.max_column + 1):
            group_name = sheet.cell(row=2, column=col).value
            if group_name:
                groups[col] = group_name.strip()

        logger.info(f"Найденные группы: {groups}")

        current_date = None
        for row in range(3, sheet.max_row + 1):
            date_cell = sheet.cell(row=row, column=2).value
            if date_cell:
                try:
                    current_date = datetime.strptime(str(date_cell), '%d.%m.%Y').strftime('%d.%m.%Y')
                    logger.info(f"Обработка даты: {current_date}")
                except ValueError:
                    continue

            if not current_date:
                continue

            lesson_num = sheet.cell(row=row, column=3).value
            if not lesson_num:
                continue

            for col, group in groups.items():
                cell_value = sheet.cell(row=row, column=col).value
                if not cell_value:
                    continue

                if current_date not in replacements:
                    replacements[current_date] = {}
                if group not in replacements[current_date]:
                    replacements[current_date][group] = {}

                cell_value = str(cell_value).strip()

                # Обработка случая, когда замены для разных подгрупп находятся в одной строке
                # Пример: "1. ------------                 2. (Лаб) МДК.01.04 Систем.программ. А207 Тутарова В.Д."
                if ('1. ' in cell_value and '2. ' in cell_value):
                    # Добавляем дополнительную проверку для лога
                    if lesson_num == 5:
                        logger.info(f"Обрабатываем замену для пары 5: {cell_value}")
                    
                    # Разбиваем по маркерам подгрупп и обрабатываем каждую часть
                    import re
                    
                    # Сначала пробуем найти точное расположение 1. и 2. в тексте
                    pos_1 = cell_value.find('1. ')
                    pos_2 = cell_value.find('2. ')
                    
                    if pos_1 != -1 and pos_2 != -1 and pos_1 < pos_2:
                        # Разделяем текст на две части - для подгруппы 1 и подгруппы 2
                        subgroup1_text = cell_value[pos_1+2:pos_2].strip()
                        subgroup2_text = cell_value[pos_2+2:].strip()
                        
                        # Проверка для дополнительной отладки
                        if lesson_num == 5:
                            logger.info(f"Пара 5, разделение строки: '{cell_value}'")
                            logger.info(f"Пара 5, подгруппа 1: '{subgroup1_text}'")
                            logger.info(f"Пара 5, подгруппа 2: '{subgroup2_text}'")
                        
                        if lesson_num not in replacements[current_date][group]:
                            replacements[current_date][group][lesson_num] = {}
                        
                        # Проверяем, является ли текст для подгруппы 1 отменой
                        is_sg1_cancellation = bool(re.match(r'^[-\s]*$', subgroup1_text)) or not subgroup1_text
                        if is_sg1_cancellation:
                            replacements[current_date][group][lesson_num][1] = {
                                'status': 'cancelled',
                                'new_data': 'Пара отменена',
                                'subgroup': 1,
                                'is_common': False,
                                'is_cancelled': True,
                                'emoji': '❌'
                            }
                        else:
                            # Удаляем возможные переносы строк в тексте замены
                            clean_text = ' '.join(subgroup1_text.replace('\n', ' ').split())
                            replacements[current_date][group][lesson_num][1] = {
                                'status': 'replaced',
                                'new_data': clean_text,
                                'subgroup': 1,
                                'is_common': False
                            }
                        
                        # Проверяем, является ли текст для подгруппы 2 отменой
                        is_sg2_cancellation = bool(re.match(r'^[-\s]*$', subgroup2_text)) or not subgroup2_text
                        if is_sg2_cancellation:
                            replacements[current_date][group][lesson_num][2] = {
                                'status': 'cancelled',
                                'new_data': 'Пара отменена',
                                'subgroup': 2,
                                'is_common': False,
                                'is_cancelled': True,
                                'emoji': '❌'
                            }
                        else:
                            # Удаляем возможные переносы строк в тексте замены
                            clean_text = ' '.join(subgroup2_text.replace('\n', ' ').split())
                            replacements[current_date][group][lesson_num][2] = {
                                'status': 'replaced',
                                'new_data': clean_text,
                                'subgroup': 2,
                                'is_common': False
                            }
                        
                        if lesson_num == 5:
                            logger.info(f"Пара 5, итоговая замена: {replacements[current_date][group][lesson_num]}")
                        
                        continue
                    
                    # Если предыдущий подход не сработал, используем запасной с разделением
                    # Разделяем строку на части, начинающиеся с "1." и "2."
                    parts = re.split(r'(\d+\.)', cell_value)
                    if len(parts) >= 3:  # Убедимся, что у нас есть хотя бы "1." и текст после него
                        if lesson_num not in replacements[current_date][group]:
                            replacements[current_date][group][lesson_num] = {}
                        
                        current_subgroup = None
                        current_text = ""
                        
                        # Пропускаем первый элемент, если он пустой (перед "1.")
                        start_idx = 0 if parts[0].strip() else 1
                        
                        # Для отладки
                        logger.info(f"Разбор данных замены в ячейке: {cell_value}")
                        logger.info(f"Разбитая строка: {parts}")
                        
                        for i in range(start_idx, len(parts), 2):
                            if i + 1 < len(parts):
                                # Это маркер подгруппы (например, "1." или "2.")
                                if parts[i].strip() in ("1. ", "2. "):
                                    # Если у нас был предыдущий subgroup, сохраним его данные
                                    if current_subgroup is not None and current_text:
                                        # Обработка предыдущей подгруппы
                                        text = current_text.strip()
                                        is_cancellation = bool(re.match(r'^[-\s]*$', text)) or not text
                                        
                                        logger.info(f"Подгруппа {current_subgroup}: текст = '{text}', отмена = {is_cancellation}")
                                        
                                        if is_cancellation:
                                            replacements[current_date][group][lesson_num][current_subgroup] = {
                                                'status': 'cancelled',
                                                'new_data': 'Пара отменена',
                                                'subgroup': current_subgroup,
                                                'is_common': False,
                                                'is_cancelled': True,
                                                'emoji': '❌'
                                            }
                                        else:
                                            # Удаляем возможные переносы строк
                                            clean_text = ' '.join(text.replace('\n', ' ').split())
                                            replacements[current_date][group][lesson_num][current_subgroup] = {
                                                'status': 'replaced',
                                                'new_data': clean_text,
                                                'subgroup': current_subgroup,
                                                'is_common': False
                                            }
                                    
                                    # Установка текущей подгруппы
                                    current_subgroup = int(parts[i].strip()[0])
                                    current_text = parts[i+1].strip()
                            
                        # Обработка последней подгруппы
                        if current_subgroup is not None and current_text:
                            text = current_text.strip()
                            is_cancellation = bool(re.match(r'^[-\s]*$', text)) or not text
                            
                            logger.info(f"Последняя подгруппа {current_subgroup}: текст = '{text}', отмена = {is_cancellation}")
                            
                            if is_cancellation:
                                replacements[current_date][group][lesson_num][current_subgroup] = {
                                    'status': 'cancelled',
                                    'new_data': 'Пара отменена',
                                    'subgroup': current_subgroup,
                                    'is_common': False,
                                    'is_cancelled': True,
                                    'emoji': '❌'
                                }
                            else:
                                # Удаляем возможные переносы строк
                                clean_text = ' '.join(text.replace('\n', ' ').split())
                                replacements[current_date][group][lesson_num][current_subgroup] = {
                                    'status': 'replaced',
                                    'new_data': clean_text,
                                    'subgroup': current_subgroup,
                                    'is_common': False
                                }
                        continue
                
                # Обработка обычных случаев с переносами строк
                # Разделяем замены для разных подгрупп
                subgroup_replacements = [r.strip() for r in cell_value.split('\n') if r.strip()]

                # Проверяем наличие маркеров подгрупп "1." и "2."
                has_subgroup_1 = any(line.startswith('1. ') for line in subgroup_replacements)
                has_subgroup_2 = any(line.startswith('2. ') for line in subgroup_replacements)

                # Если нет маркеров подгрупп или есть (ТО), значит это общая пара
                if (not (has_subgroup_1 or has_subgroup_2)) or cell_value.strip().upper().startswith('(ТО)'):
                    # Это общая замена
                    new_data = ' '.join(subgroup_replacements)
                    replacements[current_date][group][lesson_num] = {
                        'status': 'replaced',
                        'subgroup': None,  # None означает общую пару
                        'new_data': new_data,
                        'is_common': True
                    }
                    continue

                # Обрабатываем замены для подгрупп
                current_subgroup = None
                current_data = []

                for line in subgroup_replacements:
                    if line.startswith('1. ') or line.startswith('2. '):
                        # Сохраняем предыдущую подгруппу
                        if current_subgroup and current_data:
                            if lesson_num not in replacements[current_date][group]:
                                replacements[current_date][group][lesson_num] = {}
                            replacements[current_date][group][lesson_num][current_subgroup] = {
                                'status': 'replaced',
                                'new_data': ' '.join(current_data),
                                'subgroup': current_subgroup,
                                'is_common': False
                            }
                            current_data = []
                        current_subgroup = int(line[0])
                        current_data.append(line[2:].strip())
                    else:
                        current_data.append(line.strip())

                # Сохраняем последнюю подгруппу
                if current_subgroup and current_data:
                    if lesson_num not in replacements[current_date][group]:
                        replacements[current_date][group][lesson_num] = {}
                    replacements[current_date][group][lesson_num][current_subgroup] = {
                        'status': 'replaced',
                        'new_data': ' '.join(current_data),
                        'subgroup': current_subgroup,
                        'is_common': False
                    }

        logger.info(f"Загружены замены: {replacements}")
        return replacements

    except Exception as e:
        logger.error(f"Ошибка при загрузке замен: {str(e)}")
        return {}


async def get_teacher_schedule(teacher_name: str, start_date: str, end_date: str) -> str:
    """Gets schedule for teacher (async version)"""
    try:
        # Check if we have this schedule in cache
        cached_schedule = get_cached_teacher_schedule(teacher_name, start_date, end_date)
        if cached_schedule:
            logger.info(f"Using cached schedule for teacher {teacher_name}")
            return cached_schedule
            
        schedule_files = await run_blocking(lambda: [
            f for f in os.listdir("downloaded_files")
            if f.endswith('.xlsx') and not '-' in f
        ])

        all_schedules = {}
        start_date_obj = datetime.strptime(start_date, '%d.%m.%Y').date()
        end_date_obj = datetime.strptime(end_date, '%d.%m.%Y').date()

        async def process_file_and_date(file, date_str):
            try:
                file_path = os.path.join("downloaded_files", file)
                schedule = await run_blocking(parse_teacher_schedule, file_path, date_str, teacher_name)
                return date_str, schedule
            except Exception as e:
                logger.error(f"Error processing {file} for {date_str}: {e}")
                return date_str, {}

        tasks = []
        current_date = start_date_obj
        while current_date <= end_date_obj:
            if current_date.weekday() != 6:  # Skip Sundays
                date_str = current_date.strftime('%d.%m.%Y')
                for file in schedule_files:
                    tasks.append(process_file_and_date(file, date_str))
            current_date += timedelta(days=1)

        results = await asyncio.gather(*tasks)

        for date_str, schedule in results:
            if schedule:
                if date_str not in all_schedules:
                    all_schedules[date_str] = {}
                all_schedules[date_str].update(schedule)

        formatted_schedule = await run_blocking(
            format_teacher_schedule,
            all_schedules,
            teacher_name,
            start_date,
            end_date
        )
        
        # Cache the formatted schedule
        cache_teacher_schedule(teacher_name, start_date, end_date, formatted_schedule)
        
        return formatted_schedule

    except Exception as e:
        logger.error(f"Error getting teacher schedule: {e}")
        return f"Ошибка при получении расписания преподавателя: {str(e)}"


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Начало разговора и показ главного меню"""
    if not update or not update.message:
        logger.error("Объект update или update.message отсутствует.")
        return ConversationHandler.END

    # Проверяем, не идет ли процесс обновления файлов
    if is_update_in_progress():
        status_message = get_update_status_message()
        await update.message.reply_text(
            status_message,
            parse_mode='Markdown'
        )
        # Продолжаем отображать меню, даже если идет обновление
        
    keyboard = [
        ['⏰ Расписание звонков','👥 Расписание группы'],
        ['🎓 Расписание преподавателя','🚪 Расписание кабинета'],
        ['Подписаться на замены', 'Отписаться от замен']
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(
        'Выберите действие:',
        reply_markup=reply_markup
    )
    return ConversationHandler.END

async def handle_classroom_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка запроса расписания кабинета"""
    keyboard = ReplyKeyboardMarkup([
        ['Отмена']
    ], resize_keyboard=True)
    await update.message.reply_text(
        "Введите номер кабинета (например: А403):",
        reply_markup=keyboard
    )
    return ENTER_CLASSROOM

async def enter_classroom(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка ввода номера кабинета"""
    # Проверяем, не идет ли процесс обновления файлов
    if is_update_in_progress():
        status_message = get_update_status_message()
        await update.message.reply_text(
            status_message,
            parse_mode='Markdown'
        )
        # Возвращаем тот же статус, чтобы бот продолжал слушать команды
        return CHOOSE_ACTION
        
    classroom = update.message.text.strip()
    
    # Check if user clicked Cancel
    if classroom == "Отмена":
        keyboard = [
            ['⏰ Расписание звонков','👥 Расписание группы'],
            ['🎓 Расписание преподавателя','🚪 Расписание кабинета'],
            ['Подписаться на замены', 'Отписаться от замен']
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(
            'Выберите действие:',
            reply_markup=reply_markup
        )
        return ConversationHandler.END
    
    # Validate classroom format (should be like А403)
    if not re.match(r'^[А-ЯA-Z]\d{2,}$', classroom, re.IGNORECASE):
        await update.message.reply_text(
            "❌ Неверный формат номера кабинета.\n"
            "Введите в формате: Букваxx, где xx - номер\n"
            "Примеры: А403, В12, К25"
        )
        return ENTER_CLASSROOM
    
    # Store classroom in user data
    context.user_data['classroom'] = classroom
    
    # Get available dates from replacement files
    available_dates = []
    today = datetime.now().date()
    
    for file in os.listdir("downloaded_files"):
        if file.endswith('.xlsx') and '-' in file:
            try:
                dates = file.replace('.xlsx', '').split('-')
                if len(dates) == 2:
                    # Пробуем разные форматы даты
                    start_date = None
                    end_date = None
                    
                    # Пробуем формат DD.MM.YY
                    try:
                        start_date = datetime.strptime(dates[0], '%d.%m.%y').date()
                        end_date = datetime.strptime(dates[1], '%d.%m.%y').date()
                    except ValueError:
                        # Пробуем формат DD.MM.YYYY
                        try:
                            start_date = datetime.strptime(dates[0], '%d.%m.%Y').date()
                            end_date = datetime.strptime(dates[1], '%d.%m.%Y').date()
                        except ValueError:
                            logger.warning(f"Не удалось распознать формат даты в файле: {file}")
                            continue
                    
                    if start_date and end_date:
                        # Add all dates in range that are today or later
                        current_date = max(today, start_date)
                        while current_date <= end_date:
                            if current_date.weekday() != 6:  # Skip Sundays
                                date_str = current_date.strftime('%d.%m.%Y')
                                available_dates.append(date_str)
                            current_date += timedelta(days=1)
            except Exception as e:
                logger.error(f"Error processing file {file}: {e}")
                logger.error(traceback.format_exc())
    
    if not available_dates:
        # Log the files in the directory for debugging
        all_files = [f for f in os.listdir("downloaded_files") if f.endswith('.xlsx')]
        logger.info(f"All Excel files in directory: {all_files}")
        logger.info(f"No available dates found for classroom {classroom}")
        
        await update.message.reply_text(
            "❌ Не найдены актуальные файлы замен."
        )
        return ConversationHandler.END
    
    # Sort dates and create keyboard
    available_dates = sorted(set(available_dates), key=lambda x: datetime.strptime(x, '%d.%m.%Y'))
    
    # Create keyboard with dates (3 dates per row)
    keyboard = []
    row = []
    
    # Create a mapping between display format and actual date for reference
    date_mapping = {}
    
    for date_str in available_dates:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        formatted_date = f"{date_obj.strftime('%d.%m')} ({days_ru[date_obj.weekday()]})"
        date_mapping[formatted_date] = date_str  # Store mapping
        row.append(formatted_date)
        if len(row) == 3:
            keyboard.append(row)
            row = []
    
    if row:  # Add remaining dates
        keyboard.append(row)
    
    keyboard.append(['Отмена'])
    
    # Store the mapping in user_data
    context.user_data['available_dates'] = date_mapping

    # Log available dates for debugging
    logger.info(f"Available dates for classroom {classroom}: {available_dates}")
    logger.info(f"Date mapping created: {date_mapping}")
    
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "Выберите дату:",
        reply_markup=reply_markup
    )
    return CHOOSE_DATE_FOR_CLASSROOM

async def choose_date_for_classroom(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle classroom date selection."""
    # Проверяем, не идет ли процесс обновления файлов
    if is_update_in_progress():
        status_message = get_update_status_message()
        await update.message.reply_text(
            status_message,
            parse_mode='Markdown'
        )
        # Возвращаем тот же статус, чтобы бот продолжал слушать команды
        return CHOOSE_ACTION
        
    selected_date_display = update.message.text.strip()
    logger.info(f"User selected date: {selected_date_display}")
    
    # Check if user clicked Cancel
    if selected_date_display == "Отмена":
        keyboard = [
            ['⏰ Расписание звонков','👥 Расписание группы'],
            ['🎓 Расписание преподавателя','🚪 Расписание кабинета'],
            ['Подписаться на замены', 'Отписаться от замен']
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(
            'Выберите действие:',
            reply_markup=reply_markup
        )
        return ConversationHandler.END
    
    # Handle "Другой кабинет" option
    if selected_date_display == "Другой кабинет":
        keyboard = ReplyKeyboardMarkup([
            ['Отмена']
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Введите номер кабинета (например: А403):",
            reply_markup=keyboard
        )
        return ENTER_CLASSROOM
    
    # Handle "Другая дата" option
    if selected_date_display == "Другая дата":
        # Re-display the dates keyboard
        date_mapping = context.user_data.get('available_dates', {})
        if not date_mapping:
            await update.message.reply_text(
                "❌ Не найдены доступные даты. Пожалуйста, начните сначала."
            )
            return ConversationHandler.END
        
        # Convert the mapping back to a keyboard
        keyboard = []
        row = []
        # Sort dates by actual date
        sorted_dates = sorted(date_mapping.items(), 
                              key=lambda x: datetime.strptime(x[1], '%d.%m.%Y'))
        
        # Create keyboard with sorted dates (3 per row)
        for display_date, _ in sorted_dates:
            row.append(display_date)
            if len(row) == 3:
                keyboard.append(row)
                row = []
        
        if row:  # Add remaining dates
            keyboard.append(row)
        
        keyboard.append(['Отмена'])
        
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(
            "Выберите дату:",
            reply_markup=reply_markup
        )
        return CHOOSE_DATE_FOR_CLASSROOM
    
    # Get the date mapping
    date_mapping = context.user_data.get('available_dates', {})
    logger.info(f"Available date mapping: {date_mapping}")
    
    # Handle different behaviors
    selected_date = None
    
    # Check if the selected date is directly in the mapping
    if selected_date_display in date_mapping:
        selected_date = date_mapping[selected_date_display]
        logger.info(f"Found direct match in mapping: {selected_date}")
    else:
        # Try to extract date from format like "06.05 (вторник)"
        match = re.match(r'(\d{2})\.(\d{2}) \([а-яА-Я]+\)', selected_date_display)
        if match:
            day, month = match.groups()
            # Try each mapped date to find a match
            for displayed_date, actual_date in date_mapping.items():
                if displayed_date.startswith(f"{day}.{month}"):
                    selected_date = actual_date
                    logger.info(f"Found match by pattern: {selected_date}")
                    break
    
    # If still no match found
    if not selected_date:
        logger.warning(f"No date match found for: {selected_date_display}")
        await update.message.reply_text(
            "❌ Выбрана некорректная дата. Пожалуйста, выберите дату из списка."
        )
        return CHOOSE_DATE_FOR_CLASSROOM
    
    classroom = context.user_data.get('classroom')
    
    # Send a waiting message
    wait_message = await update.message.reply_text(
        f"🔍 Ищу расписание для кабинета {classroom} на {selected_date}..."
    )
    
    try:
        # Get classroom schedule
        schedule = await get_classroom_schedule(classroom, selected_date)
        
        # Try to delete the wait message
        try:
            await wait_message.delete()
        except Exception as e:
            logger.warning(f"Could not delete wait message: {e}")
        
        # Create keyboard for returning to menu
        keyboard = ReplyKeyboardMarkup([
            ['Другой кабинет', 'Другая дата'],
            ['Отмена']
        ], resize_keyboard=True)
        
        # Send the schedule
        await update.message.reply_text(schedule, reply_markup=keyboard)
        
        # Store date in context for potential "Other date" selection
        context.user_data['last_checked_date'] = selected_date
        
        # Return to the same state to allow selecting another date
        return CHOOSE_DATE_FOR_CLASSROOM
        
    except Exception as e:
        logger.error(f"Error getting classroom schedule: {e}")
        logger.error(traceback.format_exc())
        
        try:
            await wait_message.delete()
        except Exception as delete_error:
            logger.warning(f"Failed to delete wait message: {delete_error}")
            
        await update.message.reply_text(
            "Произошла ошибка при обработке вашего запроса. Пожалуйста, попробуйте еще раз."
        )
        return ConversationHandler.END

async def handle_group_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка запроса расписания группы"""
    keyboard = ReplyKeyboardMarkup([
        ['Отмена']
    ], resize_keyboard=True)
    await update.message.reply_text(
        "Введите номер группы (например: ИИпП-26-1):",
        reply_markup=keyboard
    )
    return ENTER_GROUP


async def handle_teacher_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка запроса расписания преподавателя"""
    keyboard = ReplyKeyboardMarkup([
        ['Отмена']
    ], resize_keyboard=True)
    await update.message.reply_text(
        "Введите фамилию преподавателя Введите в формате: Фамилия И.О.:",
        reply_markup=keyboard
    )
    return ENTER_TEACHER


async def process_date(context, teacher_name, date_str):
    """Обрабатывает расписание для конкретной даты"""
    try:
        schedule_files = [f for f in os.listdir("downloaded_files") if f.endswith('.xlsx')]

        # Получаем день недели
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        weekday = date_obj.weekday()
        weekday_ru = days_ru[weekday]
        week_type = get_week_type(date_str)

        # Проверяем каждый файл на наличие расписания
        found_schedule = {}
        for file in schedule_files:
            file_path = os.path.join("downloaded_files", file)
            schedule = parse_teacher_schedule(file_path, date_str, teacher_name)
            if schedule:
                found_schedule.update(schedule)

        # Всегда возвращаем информацию о дне, даже если расписания нет
        formatted = [f"\nРасписание на {weekday_ru} {date_str} ({week_type} неделя):"]

        if found_schedule:
            # Сортируем пары по номеру
            for lesson_num in sorted(found_schedule.keys(),
                                     key=lambda x: int(str(x).split('.')[0]) if isinstance(x, (int, str)) else 0):
                lesson = found_schedule[lesson_num]

                # Формируем строку с информацией о паре
                lesson_str = f"{str(lesson_num)}️⃣ {lesson['subject']} 🚪{lesson['room']}"

                # Добавляем информацию о группе и подгруппе
                if lesson.get('group'):
                    subgroup_info = f", {lesson['subgroup']}-я подгруппа" if lesson.get('subgroup') else ""
                    lesson_str += f" [{lesson['group']}{subgroup_info}]"

                formatted.append(lesson_str)
        else:
            formatted.append("В этот день пар нет")

        return "\n".join(formatted)

    except Exception as e:
        logger.error(f"Ошибка при обработке даты {date_str}: {str(e)}")
        return None


async def enter_teacher(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка ввода имени преподавателя"""
    try:
        # Проверяем, не идет ли процесс обновления файлов
        if is_update_in_progress():
            status_message = get_update_status_message()
            await update.message.reply_text(
                status_message,
                parse_mode='Markdown'
            )
            # Возвращаем тот же статус, чтобы бот продолжал слушать команды
            return CHOOSE_ACTION
            
        text = update.message.text.strip()

        # Check if user clicked Cancel
        if text == "Отмена":
            keyboard = ReplyKeyboardMarkup([
                [KeyboardButton("⏰ Расписание звонков"), KeyboardButton("👥 Расписание группы")],
                [KeyboardButton("🎓 Расписание преподавателя"), KeyboardButton("🚪 Расписание кабинета")],
                [KeyboardButton("Подписаться на замены"), KeyboardButton("Отписаться от замен")]
            ], resize_keyboard=True)
            await update.message.reply_text(
                "Выберите действие:",
                reply_markup=keyboard
            )
            return CHOOSE_ACTION

        # Validate teacher name format (Фамилия И.И.)
        name_parts = text.split()
        if len(name_parts) < 2:
            keyboard = ReplyKeyboardMarkup([
                ['Отмена']
            ], resize_keyboard=True)
            await update.message.reply_text(
                "❌ Неверный формат ФИО преподавателя.\n"
                "Введите в формате: Фамилия И.О.\n"
                "Пример: Иванов И.И.",
                reply_markup=keyboard
            )
            return ENTER_TEACHER

        context.user_data['teacher_name'] = text

        # Отправляем сообщение о начале поиска
        wait_message = await update.message.reply_text("Получаю расписание, может потребоваться около 2 минут...")

        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        # Получаем все файлы с заменами
        results = []
        for file in os.listdir("downloaded_files"):
            if file.endswith(".xlsx") and '-' in file:
                try:
                    # Используем регулярное выражение для проверки формата даты
                    import re
                    date_pattern = re.compile(r'^(\d{2}\.\d{2}\.\d{2,4})-(\d{2}\.\d{2}\.\d{2,4})\.xlsx$')
                    match = date_pattern.match(file)
                    
                    if match:
                        # Если файл соответствует формату даты
                        dates = match.groups()
                        if len(dates) == 2:
                            # Пробуем разные форматы даты
                            try:
                                # Сначала пробуем формат с двузначным годом
                                start_date = datetime.strptime(dates[0], '%d.%m.%y')
                                end_date = datetime.strptime(dates[1], '%d.%m.%y')
                            except ValueError:
                                try:
                                    # Затем пробуем формат с четырехзначным годом
                                    start_date = datetime.strptime(dates[0], '%d.%m.%Y')
                                    end_date = datetime.strptime(dates[1], '%d.%m.%Y')
                                except ValueError:
                                    logger.warning(f"Не удалось распознать формат даты в файле: {file}")
                                    continue
                            
                            results.append((start_date, end_date, file))
                except Exception as e:
                    logger.error(f"Ошибка при обработке файла {file}: {str(e)}")
                    continue

        # Находим самую раннюю и самую позднюю даты для определения диапазона
        latest_end_date = None
        earliest_start_date = None

        for result in results:
            if result:
                start_date, end_date, file_name = result
                if not latest_end_date or end_date > latest_end_date:
                    latest_end_date = end_date
                if not earliest_start_date or start_date < earliest_start_date:
                    earliest_start_date = start_date

        # Устанавливаем earliest_start_date как максимум между сегодняшней датой и самой ранней датой из файлов
        earliest_start_date = max(today, earliest_start_date) if earliest_start_date else today
        
        # Если не нашли файлы с заменами, показываем ошибку
        if not results:
            try:
                await wait_message.delete()
            except Exception as e:
                logger.warning(f"Could not delete wait message: {e}")
            
            # Выводим все файлы для отладки
            all_files = [f for f in os.listdir("downloaded_files") if f.endswith('.xlsx')]
            logger.info(f"Все файлы Excel в директории: {all_files}")
            
            await update.message.reply_text("Не найдены актуальные файлы замен.")
            return ConversationHandler.END
            
        # Форматируем даты для вызова оптимизированной функции
        start_date_str = earliest_start_date.strftime('%d.%m.%Y') if earliest_start_date else today.strftime('%d.%m.%Y')
        end_date_str = latest_end_date.strftime('%d.%m.%Y') if latest_end_date else today.strftime('%d.%m.%Y')
        
        # Используем оптимизированную функцию get_teacher_schedule, которая теперь использует индекс
        logger.info(f"Запрашиваем расписание для {text} с {start_date_str} по {end_date_str}")
        schedule_result = await get_teacher_schedule(text, start_date_str, end_date_str)
        
        try:
            await wait_message.delete()
        except Exception as e:
            logger.warning(f"Could not delete wait message: {e}")

        # Создаем клавиатуру для ввода другого преподавателя
        keyboard = ReplyKeyboardMarkup([
            ['Ввести другого преподавателя'],
            ['Расписание звонков'],
            ['Отмена']
        ], resize_keyboard=True)
        
        await update.message.reply_text(schedule_result, reply_markup=keyboard)
        return CHOOSE_ACTION
    
    except Exception as e:
        logger.error(f"Ошибка в enter_teacher: {str(e)}")
        logger.error(traceback.format_exc())
        
        await update.message.reply_text(
            "Произошла ошибка при обработке запроса. Проверьте меню.",
            reply_markup=ReplyKeyboardMarkup([['Отмена']], resize_keyboard=True)
        )
        return CHOOSE_ACTION

async def enter_group(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка ввода номера группы"""
    # Check if user clicked Cancel
    if update.message.text == "Отмена":
        keyboard = ReplyKeyboardMarkup([
            [KeyboardButton("⏰ Расписание звонков"), KeyboardButton("👥 Расписание группы")],
            [KeyboardButton("🎓 Расписание преподавателя"), KeyboardButton("🚪 Расписание кабинета")],
            [KeyboardButton("Подписаться на замены"), KeyboardButton("Отписаться от замен")]
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Выберите действие:",
            reply_markup=keyboard
        )
        return CHOOSE_ACTION

    group = update.message.text
    context.user_data['group'] = group
    # Store the group in persistent data
    if 'user_groups' not in context.application.user_data:
        context.application.user_data['user_groups'] = {}
    context.application.user_data['user_groups'][update.effective_user.id] = group

    keyboard = [['1-я подгруппа', '2-я подгруппа'], ['Общее расписание']]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(
        'Выберите подгруппу:',
        reply_markup=reply_markup
    )
    return CHOOSE_SUBGROUP


async def choose_subgroup(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка выбора подгруппы"""
    choice = update.message.text
    group = context.user_data.get('group')

    if choice == 'Общее расписание':
        subgroup = None
    else:
        subgroup = 1 if '1' in choice else 2

    schedule = await get_schedule_for_days(group, subgroup, update)
    await update.message.reply_text(schedule)

    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update and update.effective_message:
        keyboard = ReplyKeyboardMarkup([
            [KeyboardButton("⏰ Расписание звонков"), KeyboardButton("👥 Расписание группы")],
            [KeyboardButton("🎓 Расписание преподавателя") , KeyboardButton("🚪 Расписание кабинета")],
            [KeyboardButton("Подписаться на замены"), KeyboardButton("Отписаться от замен")]
        ], resize_keyboard=True)
        await update.effective_message.reply_text("Операция отменена. Выберите действие заново:", reply_markup=keyboard)
    else:
        logger.error("Не удалось отправить сообщение об отмене.")
    return CHOOSE_ACTION


async def set_commands(application: Application):
    commands = [
        BotCommand(command="start", description="Начать работу с ботом"),
        BotCommand(command="help", description="Показать помощь"),
        BotCommand(command="cancel", description="Отменить текущую операцию"),
        BotCommand(command="clear_cache", description="Очистить кэш (только для администраторов)"),
        BotCommand("classroom", "Расписание кабинета (например, /classroom А403 02.05.2023)"),
        BotCommand(command="myid", description="Узнать свой ID пользователя")
    ]
    await application.bot.set_my_commands(commands)


# Добавьте новую функцию для поиска файла замен
def find_applicable_replacement_files(dbx, target_date):
    """
    Находит все подходящие файлы замен для указанной даты
    """
    try:
        result = dbx.files_list_folder('')
        target_date_obj = datetime.strptime(target_date, '%d.%m.%Y')
        matching_files = []

        for entry in result.entries:
            if not entry.name.endswith('.xlsx'):
                continue

            try:
                # Извлекаем даты из имени файла (формат DD.MM.YY-DD.MM.YY.xlsx)
                dates = entry.name.replace('.xlsx', '').split('-')
                if len(dates) != 2:
                    continue

                start_date = datetime.strptime(dates[0], '%d.%m.%y')
                end_date = datetime.strptime(dates[1], '%d.%m.%y')

                # Проверяем, попадает ли целевая дата в диапазон
                if start_date <= target_date_obj <= end_date:
                    matching_files.append(entry.name)

            except ValueError:
                continue

        return matching_files

    except Exception as e:
        logger.error(f"Ошибка при поиске файлов замен: {str(e)}")
        return []


def format_schedule(schedule, group_name, date_str, selected_subgroup=None):
    """Форматирует расписание с учетом замен"""
    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
    week_type = get_week_type(date_str)
    weekday_ru = days_ru.get(date_obj.weekday(), str(date_obj.weekday()))

    # Добавляем информацию о подгруппе в заголовок
    subgroup_info = f" ({selected_subgroup}-я подгруппа)" if selected_subgroup else ""
    formatted = [f"Расписание группы {group_name}{subgroup_info} на {weekday_ru} {date_str} ({week_type} неделя)"]

    # Сортируем уроки по их оригинальным номерам
    def get_sort_key(key):
        if isinstance(key, str) and '_' in key:
            num = key.split('_')[0]
        else:
            num = str(key)
        return int(float(num.replace(',', '.')))

    sorted_lessons = sorted(schedule.keys(), key=get_sort_key)

    for lesson_key in sorted_lessons:
        lesson_info = schedule[lesson_key]
        
        # Добавляем отладку для пары 5
        if str(lesson_key).startswith('5') or (isinstance(lesson_info.get('original_num'), (int, str)) and str(lesson_info.get('original_num')).startswith('5')):
            logger.info(f"Форматирование пары 5: ключ={lesson_key}, данные={lesson_info}")
            
            # Специальная обработка для пары 5
            display_num = 5
            lesson_str = f"{str(display_num)}️⃣"
            
            # Если это отмена (для подгруппы 2)
            if lesson_info.get('is_cancelled') or lesson_info.get('status') == 'cancelled' or (
                lesson_info.get('subgroup') == 2 and lesson_info.get('new_data') and (
                    bool(re.match(r'^[-\s]*$', lesson_info['new_data'])) or 
                    '--------' in lesson_info['new_data']
                )
            ):
                lesson_str += "➡️ Пара отменена ❌"
            elif lesson_info.get('status') == 'replaced':
                # Проверяем на особый случай урока 5
                new_data = lesson_info.get('new_data', '')
                
                # Специальный случай для урока 5 "Ин.яз (проф.) Л725м Грипкова Г.И."
                if "Ин.яз (проф.)" in new_data and "Грипкова Г.И." in new_data and "Л725м" in new_data:
                    lesson_str += f"➡️✏️ Ин.яз (проф.) 🎓Грипкова Г.И. 🚪Л725м"
                    formatted.append(lesson_str)
                    continue
                
                # Если это замена для подгруппы 1 - показываем урок
                if lesson_info.get('subgroup') == 1:
                    new_data = lesson_info.get('new_data', '')
                    # Очищаем от упоминаний второй подгруппы, если они есть
                    if '2. ' in new_data:
                        new_data = new_data.split('2. ')[0].strip()
                    # Форматируем в одну строку и добавляем карандаш
                    new_data = ' '.join(new_data.replace('\n', ' ').split())
                    
                    # Проверяем наличие и добавляем эмодзи для преподавателя и аудитории
                    room_match = re.search(r'\b[ЛАл]\d{3,}[кКтТ]?\b', new_data)
                    teacher_match = re.search(r'([А-Яа-я]+\s+[А-Я]\.[А-Я]\.)', new_data)
                    
                    if teacher_match and room_match:
                        # Извлекаем информацию о преподавателе и аудитории
                        teacher = teacher_match.group(0)
                        room = room_match.group(0)
                        # Удаляем эту информацию из исходного текста
                        subject_text = new_data.replace(teacher, '').replace(room, '').strip()
                        # Формируем новую строку с эмодзи
                        lesson_str += f"➡️✏️ {subject_text} 🎓{teacher} 🚪{room}"
                    else:
                        # Улучшенный поиск преподавателя и аудитории
                        teacher_pattern = r'([А-Яа-я]+)\s+([А-Я])\.\s*([А-Я])\.?'
                        classroom_pattern = r'\b[ЛАл]\d{3,}[кКтТ]?\b'
                        
                        teacher_match = re.search(teacher_pattern, new_data)
                        room_match = re.search(classroom_pattern, new_data)
                        
                        if teacher_match and room_match:
                            teacher = teacher_match.group(0)
                            room = room_match.group(0)
                            
                            # Удаляем из текста
                            subject_text = new_data
                            subject_text = re.sub(teacher_pattern, '', subject_text)
                            subject_text = re.sub(classroom_pattern, '', subject_text)
                            subject_text = subject_text.strip()
                            
                            lesson_str += f"➡️✏️ {subject_text} 🎓{teacher} 🚪{room}"
                        else:
                            lesson_str += f"➡️✏️ {new_data}"
                else:
                    # Для подгруппы 2, проверяем данные
                    new_data = lesson_info.get('new_data', '')
                    if not new_data or new_data.strip().startswith('-'):
                        lesson_str += "➡️ Пара отменена ❌"
                    else:
                        # Форматируем в одну строку и добавляем карандаш
                        new_data = ' '.join(new_data.replace('\n', ' ').split())
                        
                        # Проверяем наличие и добавляем эмодзи для преподавателя и аудитории
                        room_match = re.search(r'\b[ЛАл]\d{3,}[кКтТ]?\b', new_data)
                        teacher_match = re.search(r'([А-Яа-я]+\s+[А-Я]\.[А-Я]\.)', new_data)
                        
                        if teacher_match and room_match:
                            # Извлекаем информацию о преподавателе и аудитории
                            teacher = teacher_match.group(0)
                            room = room_match.group(0)
                            # Удаляем эту информацию из исходного текста
                            subject_text = new_data.replace(teacher, '').replace(room, '').strip()
                            # Формируем новую строку с эмодзи
                            lesson_str += f"➡️✏️ {subject_text} 🎓{teacher} 🚪{room}"
                        else:
                            lesson_str += f"➡️✏️ {new_data}"
            else:
                # Обычный урок
                subject = lesson_info.get('subject', '')
                teacher = lesson_info.get('teacher', '')
                room = lesson_info.get('room', '')

                if subject:
                    lesson_str += f" {subject}"
                if teacher:
                    lesson_str += f" 🎓{teacher}"
                if room:
                    lesson_str += f" 🚪{room}"
                    
            formatted.append(lesson_str)
            continue
        
        # Для всех остальных пар - стандартная обработка
        # Пропускаем уроки не выбранной подгруппы
        lesson_subgroup = lesson_info.get('subgroup')
        if selected_subgroup is not None and lesson_subgroup is not None and lesson_subgroup != selected_subgroup:
            logger.info(f"Пропускаем урок {lesson_key} с подгруппой {lesson_subgroup}, т.к. выбрана подгруппа {selected_subgroup}")
            continue

        # Получаем оригинальный номер урока для отображения
        display_num = lesson_info.get('original_num', lesson_key)
        if isinstance(display_num, str) and '_' in display_num:
            display_num = display_num.split('_')[0]

        lesson_str = f"{str(display_num)}️⃣"
        
        # Добавляем индикатор для замен
        if lesson_info.get('is_replacement'):
            lesson_str += "➡️"

        # Обработка отмены пары
        if lesson_info.get('is_cancelled') == True:
            lesson_str += " Пара отменена ❌"
        elif lesson_info.get('status') == 'cancelled':
            lesson_str += " Пара отменена ❌"
        elif lesson_info.get('status') == 'replaced':
            lesson_str += "✏️ "
            new_data = lesson_info.get('new_data', '')
            
            # Проверяем, является ли замена фактически отменой (строка с дефисами)
            is_cancellation = False
            if new_data:
                new_data_clean = new_data.replace(' ', '')
                if (set(new_data_clean) == {'-'} or 
                    any(pattern in new_data_clean for pattern in ['----', '-----', '------', '-------', '--------', '---------', '----------', '-----------', '------------'])):
                    is_cancellation = True
            
            if is_cancellation:
                lesson_str = f"{str(display_num)}️⃣➡️ Пара отменена ❌"
            else:
                # Форматируем текст замены в одну строку
                new_data = ' '.join(new_data.replace('\n', ' ').split())
                
                # Проверяем наличие и добавляем эмодзи для преподавателя и аудитории
                room_match = re.search(r'\b[ЛАл]\d{3,}[кКтТ]?\b', new_data)
                teacher_match = re.search(r'([А-Яа-я]+\s+[А-Я]\.[А-Я]\.)', new_data)
                
                if teacher_match and room_match:
                    # Извлекаем информацию о преподавателе и аудитории
                    teacher = teacher_match.group(0)
                    room = room_match.group(0)
                    # Удаляем эту информацию из исходного текста
                    subject_text = new_data.replace(teacher, '').replace(room, '').strip()
                    # Формируем новую строку с эмодзи
                    lesson_str += f"{subject_text} 🎓{teacher} 🚪{room}"
                else:
                    # Улучшенный поиск преподавателя и аудитории
                    teacher_pattern = r'([А-Яа-я]+)\s+([А-Я])\.\s*([А-Я])\.?'
                    classroom_pattern = r'\b[ЛАл]\d{3,}[кКтТ]?\b'
                    
                    teacher_match = re.search(teacher_pattern, new_data)
                    room_match = re.search(classroom_pattern, new_data)
                    
                    if teacher_match and room_match:
                        teacher = teacher_match.group(0)
                        room = room_match.group(0)
                        
                        # Удаляем из текста
                        subject_text = new_data
                        subject_text = re.sub(teacher_pattern, '', subject_text)
                        subject_text = re.sub(classroom_pattern, '', subject_text)
                        subject_text = subject_text.strip()
                        
                        lesson_str += f"{subject_text} 🎓{teacher} 🚪{room}"
                    else:
                        lesson_str += f"{new_data}"
        else:
            if not lesson_info.get('is_common'):
                lesson_str += "➡️"

            subject = lesson_info.get('subject', '')
            teacher = lesson_info.get('teacher', '')
            room = lesson_info.get('room', '')

            if subject:
                lesson_str += f" {subject}"
            if teacher:
                lesson_str += f" 🎓{teacher}"
            if room:
                lesson_str += f" 🚪{room}"

        formatted.append(lesson_str)

    return "\n".join(formatted) if formatted else "Расписание на этот день не найдено"


# Добавьте функцию обработки ошибок
async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Обработчик ошибок для бота"""
    logger.error(f"Произошла ошибка: {context.error}")

    try:
        if update and update.effective_message:
            # Проверяем, не связана ли ошибка с обновлением файлов
            if is_update_in_progress():
                status_message = get_update_status_message()
                await update.effective_message.reply_text(
                    status_message,
                    parse_mode='Markdown'
                )
            else:
                await update.effective_message.reply_text(
                    "Произошла ошибка при обработке запроса. Проверьте меню."
                )
    except Exception as e:
        logger.error(f"Ошибка при отправке сообщения об ошибке: {str(e)}")

    # Если произошел таймаут, логируем это событие
    if isinstance(context.error, TimeoutError):
        logger.warning("Таймаут при обращении к Telegram API. Возможно, бот перегружен операциями с файлами.")
    
    # Для других типов ошибок
    elif isinstance(context.error, (IOError, ConnectionError)):
        logger.error(f"Ошибка ввода/вывода или соединения: {context.error}")

    # Для неизвестных ошибок выводим полный стек-трейс
    else:
        import traceback
        traceback.print_exception(None, context.error, context.error.__traceback__)
        logger.error(f"Необработанная ошибка: {traceback.format_exc()}")


# Добавьте функцию для корректного завершения бота
def signal_handler(signum, frame):
    print('Завершение работы бота...')
    sys.exit(0)


async def choose_date_for_teacher(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    date_str = update.message.text.strip()
    try:
        datetime.strptime(date_str, '%d.%m.%Y')
    except ValueError:
        await update.message.reply_text("Неверный формат даты. Попробуйте снова (формат: DD.MM.YYYY):")
        return CHOOSE_DATE_FOR_TEACHER

    context.user_data['date_str'] = date_str
    await update.message.reply_text("Получаю расписание, может потребовать от 30 секунд до 4 минут...")
    teacher_name = context.user_data.get('teacher_name')
    schedule_text = await get_teacher_schedule(teacher_name, date_str, date_str)
    await update.message.reply_text(schedule_text)
    return ConversationHandler.END


async def subgroup_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle subgroup choice"""
    # Проверяем, не идет ли процесс обновления файлов
    if is_update_in_progress():
        status_message = get_update_status_message()
        await update.message.reply_text(
            status_message,
            parse_mode='Markdown'
        )
        # Возвращаем тот же статус, чтобы бот продолжал слушать команды
        return CHOOSE_ACTION
        
    choice = update.message.text
    group = context.user_data.get('group')

    # If group is not in context, try to get it from persistent storage
    if not group and 'user_groups' in context.application.user_data:
        group = context.application.user_data['user_groups'].get(update.effective_user.id)
        if group:
            context.user_data['group'] = group

    if not group:
        keyboard = ReplyKeyboardMarkup([
            ['👥 Расписание группы'],
            ['Отмена']
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Пожалуйста, выберите группу заново:",
            reply_markup=keyboard
        )
        return CHOOSE_ACTION

    subgroup = None
    if choice == "Первая подгруппа":
        subgroup = 1
    elif choice == "Вторая подгруппа":
        subgroup = 2
    elif choice == "Отмена":
        return await cancel(update, context)
    else:
        keyboard = ReplyKeyboardMarkup([
            [KeyboardButton("Первая подгруппа"), KeyboardButton("Вторая подгруппа")],
            [KeyboardButton("Отмена")]
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Пожалуйста, выберите подгруппу, используя кнопки.",
            reply_markup=keyboard
        )
        return CHOOSE_SUBGROUP

    schedule_text = await get_schedule_for_days(group, subgroup, update)
    await update.message.reply_text(schedule_text)

    keyboard = ReplyKeyboardMarkup([
        [KeyboardButton("Первая подгруппа"), KeyboardButton("Вторая подгруппа")],
        [KeyboardButton("Отмена")]
    ], resize_keyboard=True)

    await update.message.reply_text(
        f"Группа: {group}\nВыберите другую подгруппу или нажмите 'Отмена':",
        reply_markup=keyboard
    )
    return CHOOSE_SUBGROUP


async def group_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обработка ввода номера группы"""
    if update.message is None:
        return ConversationHandler.END

    # Проверяем, не идет ли процесс обновления файлов
    if is_update_in_progress():
        status_message = get_update_status_message()
        await update.message.reply_text(
            status_message,
            parse_mode='Markdown'
        )
        # Возвращаем тот же статус, чтобы бот продолжал слушать команды
        return CHOOSE_ACTION
        
    group = update.message.text.strip()

    # Check if user clicked Cancel
    if group == "Отмена":
        keyboard = ReplyKeyboardMarkup([
                [KeyboardButton("⏰ Расписание звонков"), KeyboardButton("👥 Расписание группы")],
                [KeyboardButton("🎓 Расписание преподавателя"), KeyboardButton("🚪 Расписание кабинета")],
                [KeyboardButton("Подписаться на замены"), KeyboardButton("Отписаться от замен")]
            ], resize_keyboard=True)
        await update.message.reply_text(
                "Выберите действие:",
                reply_markup=keyboard
            )
        return CHOOSE_ACTION

    context.user_data['group'] = group

    try:
        file_name = f"{group}.xlsx"
        local_path = os.path.join("downloaded_files", file_name)

        # Проверяем существование файла локально
        if not os.path.exists(local_path):
            keyboard = ReplyKeyboardMarkup([
                ['Отмена']
            ], resize_keyboard=True)
            await update.message.reply_text(
                "❌ Группа не найдена. Проверьте правильность написания и попробуйте снова.\n"
                "Пример: ИСпП-22-1",
                reply_markup=keyboard
            )
            return ENTER_GROUP

        # Если файл существует, показываем кнопки выбора подгруппы
        keyboard = ReplyKeyboardMarkup([
            [KeyboardButton("Первая подгруппа"), KeyboardButton("Вторая подгруппа")],
            [KeyboardButton("Отмена")]
        ], resize_keyboard=True)

        await update.message.reply_text(
            f"✅ Группа {group} найдена\nВыберите подгруппу:",
            reply_markup=keyboard
        )
        return CHOOSE_SUBGROUP

    except Exception as e:
        logger.error(f"Ошибка при проверке группы: {str(e)}")
        keyboard = ReplyKeyboardMarkup([
            ['Отмена']
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Произошла ошибка. Попробуйте еще раз или обратитесь к администратору.",
            reply_markup=keyboard
        )
        return ENTER_GROUP


# Измените функцию start:
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not update or not update.message:
        logger.error("Объект update или update.message отсутствует.")
        return ConversationHandler.END

    # Проверяем, не идет ли процесс обновления файлов
    if is_update_in_progress():
        status_message = get_update_status_message()
        await update.message.reply_text(
            status_message,
            parse_mode='Markdown'
        )

    keyboard = ReplyKeyboardMarkup([
        [KeyboardButton("⏰ Расписание звонков"), KeyboardButton("👥 Расписание группы")],
        [KeyboardButton("🎓 Расписание преподавателя"), KeyboardButton("🚪 Расписание кабинета")],
        [KeyboardButton("Подписаться на замены"), KeyboardButton("Отписаться от замен")]
    ], resize_keyboard=True)

    welcome_text = """
👋 Добро пожаловать!
"""
    await update.message.reply_text(welcome_text, reply_markup=keyboard)
    return CHOOSE_ACTION


async def choice_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    if not query:
        logger.error("Объект callback_query отсутствует.")
        return ConversationHandler.END

    await query.answer()

    if query.data == "timetable_calls":
        now = datetime.now()
        weekday = now.weekday()
        if weekday == 5:  # Суббота
            schedule = {
                1: "8:30 - 10:00",
                2: "10:10 - 11:40",
                3: "11:50 - 13:20",
                4: "13:30 - 15:00",
                5: "15:10 - 16:40"
            }
        else:  # ПН-ПТ
            schedule = {
                1: "8:30 - 10:00",
                2: "10:10 - 11:40",
                3: "12:20 - 13:50",
                4: "14:20 - 15:50",
                5: "16:00 - 17:30",
                6: "17:40 - 19:10"
            }

        current_time = now.strftime("%H:%M")
        formatted_schedule = []
        next_lesson = None

        for lesson_num, time_range in schedule.items():
            start_time, end_time = time_range.split(" - ")
            lesson_start = datetime.strptime(start_time, "%H:%M")
            lesson_end = datetime.strptime(end_time, "%H:%M")
            current_hour_minute = datetime.strptime(current_time, "%H:%M")

            if lesson_start <= current_hour_minute < lesson_end:
                formatted_schedule.append(f"{lesson_num}️⃣ {time_range} ➡️ Сейчас идет пара")
                next_lesson = None
            elif current_hour_minute < lesson_start:
                time_until_start = (lesson_start - current_hour_minute).seconds // 60
                formatted_schedule.append(f"{lesson_num}️⃣ {time_range} ({time_until_start} мин. до начала)")
                if not next_lesson:
                    next_lesson = f"Следующая пара: {lesson_num}️⃣ {time_range}"
            elif lesson_end < current_hour_minute < (lesson_end + timedelta(minutes=10)):
                next_lesson_start = schedule.get(lesson_num + 1, None)
                if next_lesson_start:
                    next_start_time = next_lesson_start.split(" - ")[0]
                    next_start = datetime.strptime(next_start_time, "%H:%M")
                    time_until_next = (next_start - current_hour_minute).seconds // 60
                    formatted_schedule.append(
                        f"{lesson_num}️⃣ {time_range} ⏳ Перемена ({time_until_next} мин. до следующей пары)")
            else:
                formatted_schedule.append(f"{lesson_num}️⃣ {time_range}")

        if next_lesson:
            formatted_schedule.append(f"\n➡️ {next_lesson}")

        calls_schedule = "\n".join(formatted_schedule)

        # Создаем клавиатуру с кнопкой "Отмена"
        keyboard = [
            [InlineKeyboardButton("Отмена", callback_data="cancel")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        await query.edit_message_text(text=f"⏰ Расписание звонков:\n{calls_schedule}", reply_markup=reply_markup)
        return ConversationHandler.END

    elif query.data == "group_timetable":
        await query.edit_message_text(text="Введите название вашей группы (например, ИСпП-22-1):")
        return ENTER_GROUP


async def choose_action(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    # Проверяем, не идет ли процесс обновления файлов
    if is_update_in_progress():
        status_message = get_update_status_message()
        await update.message.reply_text(
            status_message,
            parse_mode='Markdown'
        )
        # Возвращаем тот же статус, чтобы бот продолжал слушать команды
        return CHOOSE_ACTION

    choice = update.message.text

    if choice == "Ввести другого преподавателя":
        keyboard = ReplyKeyboardMarkup([
            ['Отмена']
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Введите Фамилия И.О. преподавателя:",
            reply_markup=keyboard
        )
        return ENTER_TEACHER

    # Обработка кнопок с эмодзи
    elif choice == "⏰ Расписание звонков":
        now = datetime.now()
        weekday = now.weekday()
        if weekday == 5:  # Суббота
            schedule = {
                1: "8:30 - 10:00",
                2: "10:10 - 11:40",
                3: "11:50 - 13:20",
                4: "13:30 - 15:00",
                5: "15:10 - 16:40"
            }
        else:  # ПН-ПТ
            schedule = {
                1: "8:30 - 10:00",
                2: "10:10 - 11:40",
                3: "12:20 - 13:50",
                4: "14:20 - 15:50",
                5: "16:00 - 17:30",
                6: "17:40 - 19:10"
            }

        current_time = now.strftime("%H:%M")
        formatted_schedule = []
        next_lesson = None

        for lesson_num, time_range in schedule.items():
            start_time, end_time = time_range.split(" - ")
            lesson_start = datetime.strptime(start_time, "%H:%M")
            lesson_end = datetime.strptime(end_time, "%H:%M")
            current_hour_minute = datetime.strptime(current_time, "%H:%M")

            if lesson_start <= current_hour_minute < lesson_end:
                formatted_schedule.append(f"{lesson_num}️⃣ {time_range} ➡️ Сейчас идет пара")
                next_lesson = None
            elif current_hour_minute < lesson_start:
                time_until_start = (lesson_start - current_hour_minute).seconds // 60
                formatted_schedule.append(f"{lesson_num}️⃣ {time_range} ({time_until_start} мин. до начала)")
                if not next_lesson:
                    next_lesson = f"Следующая пара: {lesson_num}️⃣ {time_range}"
            elif lesson_end < current_hour_minute < (lesson_end + timedelta(minutes=10)):
                next_lesson_start = schedule.get(lesson_num + 1, None)
                if next_lesson_start:
                    next_start_time = next_lesson_start.split(" - ")[0]
                    next_start = datetime.strptime(next_start_time, "%H:%M")
                    time_until_next = (next_start - current_hour_minute).seconds // 60
                    formatted_schedule.append(
                        f"{lesson_num}️⃣ {time_range} ⏳ Перемена ({time_until_next} мин. до следующей пары)")
            else:
                formatted_schedule.append(f"{lesson_num}️⃣ {time_range}")

        if next_lesson:
            formatted_schedule.append(f"\n➡️ {next_lesson}")

        calls_schedule = "\n".join(formatted_schedule)
        keyboard = ReplyKeyboardMarkup([
            ['⏰ Расписание звонков'],
            ['Отмена']
        ], resize_keyboard=True)
        await update.message.reply_text(text=f"⏰ Расписание звонков:\n{calls_schedule}", reply_markup=keyboard)
        return CHOOSE_ACTION

    elif choice == "👥 Расписание группы" or choice == "Расписание группы":
        keyboard = ReplyKeyboardMarkup([
            ['Отмена']
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Введите название вашей группы (например, ИСпП-22-1):",
            reply_markup=keyboard
        )
        return ENTER_GROUP
    
    elif choice == "🚪 Расписание кабинета" or choice == "Расписание кабинета":
        keyboard = ReplyKeyboardMarkup([
            ['Отмена']
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Введите номер кабинета (например: А403, В12, К25):",
            reply_markup=keyboard
        )
        return ENTER_CLASSROOM

    elif choice == "Отмена":
        return await cancel(update, context)

    elif choice == "🎓 Расписание преподавателя" or choice == "Расписание преподавателя":
        keyboard = ReplyKeyboardMarkup([
            ['Отмена']
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Введите имя преподавателя (например, Иванов И.И.):",
            reply_markup=keyboard
        )
        return ENTER_TEACHER

    elif choice == "✏️ Подписаться на замены" or choice == "Подписаться на замены":
        await subscribe_command(update, context)
        return CHOOSE_ACTION
        
    elif choice == "❌ Отписаться от замен" or choice == "Отписаться от замен":
        await unsubscribe_command(update, context)
        return CHOOSE_ACTION

    else:
        keyboard = ReplyKeyboardMarkup([
            [KeyboardButton("⏰ Расписание звонков"), KeyboardButton("👥 Расписание группы")],
            [KeyboardButton("🎓 Расписание преподавателя"), KeyboardButton("🚪 Расписание кабинета")],
            [KeyboardButton("Подписаться на замены"), KeyboardButton("Отписаться от замен")]
        ], resize_keyboard=True)
        await update.message.reply_text("Пожалуйста, выберите одну из доступных опций.", reply_markup=keyboard)
        return CHOOSE_ACTION


async def get_schedule_for_days(group: str, subgroup: int = None, update: Update = None) -> str:
    """Gets schedule for group with replacements from local files (async version)"""
    try:
        if update:
            wait_message = await update.message.reply_text("Подождите...")

        # Check cache at the start
        logger.info(f"Проверка кэша для группы {group}, подгруппа {subgroup}")
        cached_schedule = await run_blocking(get_cached_student_schedule, group, subgroup)
        if cached_schedule:
            logger.info("Найдено кэшированное расписание")
            if update and wait_message:
                await wait_message.delete()
            return cached_schedule

        # Get schedule files list
        schedule_files = await run_blocking(lambda: [f for f in os.listdir("downloaded_files") if f.endswith('.xlsx')])

        # Find schedule file case-insensitive
        schedule_file = None
        group_upper = group.upper()
        for file in schedule_files:
            if file.upper().startswith(group_upper) and file.endswith('.xlsx'):
                schedule_file = file
                break

        if not schedule_file:
            if update and wait_message:
                await wait_message.delete()
            return f"Расписание для группы {group} не найдено."

        schedule_file_path = os.path.join("downloaded_files", schedule_file)
        replacement_files = [f for f in schedule_files if '-' in f]

        # Get dates to check
        dates_to_check = set()
        today = datetime.now().date()

        # Add dates from replacement files
        for replacement_file in replacement_files:
            try:
                dates = replacement_file.replace('.xlsx', '').split('-')
                if len(dates) == 2:
                    # Пробуем разные форматы даты
                    start_date = None
                    end_date = None
                    
                    # Сначала пробуем формат с двузначным годом
                    try:
                        start_date = datetime.strptime(dates[0], '%d.%m.%y').date()
                        end_date = datetime.strptime(dates[1], '%d.%m.%y').date()
                    except ValueError:
                        # Пробуем формат с четырехзначным годом
                        try:
                            start_date = datetime.strptime(dates[0], '%d.%m.%Y').date()
                            end_date = datetime.strptime(dates[1], '%d.%m.%Y').date()
                        except ValueError:
                            logger.warning(f"Не удалось распознать формат даты в файле: {replacement_file}")
                            continue
                    
                    logger.info(f"Обработка файла замен: {replacement_file}, даты: {start_date} - {end_date}")
                    
                    current_date = max(today, start_date)
                    while current_date <= end_date:
                        if current_date.weekday() != 6:
                            dates_to_check.add(current_date.strftime('%d.%m.%Y'))
                        current_date += timedelta(days=1)
            except Exception as e:
                logger.error(f"Ошибка при обработке файла замен {replacement_file}: {str(e)}")
                logger.error(traceback.format_exc())
                continue

        # Add next 7 days if no dates from replacement files
        if not dates_to_check:
            current_date = today
            days_added = 0
            while days_added < 7:
                if current_date.weekday() != 6:
                    dates_to_check.add(current_date.strftime('%d.%m.%Y'))
                    days_added += 1
                current_date += timedelta(days=1)

        # Create tasks for parallel processing
        async def process_date(date_str):
            try:
                schedule = None
                logger.info(f"Обработка расписания на дату: {date_str}")
                
                # Найдем подходящие файлы замен для этой даты
                applicable_replacement_files = []
                date_obj = datetime.strptime(date_str, '%d.%m.%Y').date()
                
                for replacement_file in replacement_files:
                    try:
                        dates = replacement_file.replace('.xlsx', '').split('-')
                        if len(dates) == 2:
                            # Пробуем разные форматы даты
                            try:
                                # Сначала пробуем формат с двузначным годом
                                start_date = datetime.strptime(dates[0], '%d.%m.%y').date()
                                end_date = datetime.strptime(dates[1], '%d.%m.%y').date()
                            except ValueError:
                                try:
                                    # Затем пробуем формат с четырехзначным годом
                                    start_date = datetime.strptime(dates[0], '%d.%m.%Y').date()
                                    end_date = datetime.strptime(dates[1], '%d.%m.%Y').date()
                                except ValueError:
                                    # Если не удается распознать - пропускаем файл
                                    continue
                            
                            if start_date <= date_obj <= end_date:
                                applicable_replacement_files.append(replacement_file)
                                logger.info(f"Найден подходящий файл замен {replacement_file} для даты {date_str}")
                    except Exception as e:
                        logger.error(f"Ошибка при проверке файла замен {replacement_file}: {e}")
                
                # Check replacements using the applicable files
                for replacement_file in applicable_replacement_files:
                    replacement_path = os.path.join("downloaded_files", replacement_file)
                    logger.info(f"Применение файла замен {replacement_file} для даты {date_str}")
                    temp_schedule = await run_blocking(
                        process_schedule_with_replacements,
                        schedule_file_path,
                        replacement_path,
                        date_str,
                        subgroup
                    )
                    if temp_schedule:
                        schedule = temp_schedule
                        break

                # Use regular schedule if no replacements
                if not schedule:
                    schedule = await run_blocking(parse_schedule, schedule_file_path, date_str, subgroup)

                if schedule:
                    formatted_schedule = await run_blocking(format_schedule, schedule, group, date_str, subgroup)
                    if formatted_schedule:
                        return (date_str, formatted_schedule)
            except Exception as e:
                logger.error(f"Ошибка при обработке расписания на {date_str}: {str(e)}")
                logger.error(traceback.format_exc())
            return None

        # Process all dates concurrently
        tasks = [process_date(date_str) for date_str in sorted(dates_to_check)]
        results = await asyncio.gather(*tasks)

        # Filter out None results and sort by date
        valid_schedules = [(date_str, schedule) for result in results if result for date_str, schedule in [result]]
        sorted_schedules = sorted(valid_schedules, key=lambda x: datetime.strptime(x[0], '%d.%m.%Y'))
        schedules = [schedule for _, schedule in sorted_schedules]

        if update and wait_message:
            await wait_message.delete()

        if not schedules:
            return f"Расписание для группы {group} не найдено."

        result = "\n\n".join(schedules)

        # Cache the result
        await run_blocking(cache_student_schedule, group, subgroup, result)

        return result

    except Exception as e:
        logger.error(f"Ошибка при получении расписания: {str(e)}")
        if update and wait_message:
            await wait_message.delete()
        return f"Произошла ошибка при получении расписания: {str(e)}"


async def handle_all_messages(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle all messages and maintain menu functionality"""
    if not update.message or not update.message.text:
        return ConversationHandler.END

    # Проверяем, не идет ли процесс обновления файлов
    if is_update_in_progress():
        status_message = get_update_status_message()
        await update.message.reply_text(
            status_message,
            parse_mode='Markdown'
        )
        # Возвращаем тот же статус, чтобы бот продолжал слушать команды
        return CHOOSE_ACTION

    text = update.message.text.strip()

    # Check if it's a menu command or button press
    if text in ["👥 Расписание группы", "🎓 Расписание преподавателя", "⏰ Расписание звонков", "🚪 Расписание кабинета", "Отмена", "Подписаться на замены", "Отписаться от замен"]:
        return await choose_action(update, context)
    elif text in ["Первая подгруппа", "Вторая подгруппа"]:
        return await subgroup_choice(update, context)
    elif text == "Ввести другого преподавателя":
        keyboard = ReplyKeyboardMarkup([
            ['Отмена']
        ], resize_keyboard=True)
        await update.message.reply_text(
            "Введите Фамилия И.О. преподавателя:",
            reply_markup=keyboard
        )
        return ENTER_TEACHER
    else:
        # Show main menu with all options
        keyboard = ReplyKeyboardMarkup([
            [KeyboardButton("⏰ Расписание звонков"), KeyboardButton("👥 Расписание группы")],
            [KeyboardButton("🎓 Расписание преподавателя"), KeyboardButton("🚪 Расписание кабинета")], 
            [KeyboardButton("Подписаться на замены"), KeyboardButton("Отписаться от замен")]
        ], resize_keyboard=True)

        welcome_text = """
👋 Выберите действие:
"""
        await update.message.reply_text(welcome_text, reply_markup=keyboard)
        return CHOOSE_ACTION


def clear_cache():
    """Helper function to clear the cache"""
    try:
        # Use cache_utils functions to clear the cache
        from cache_utils import clear_cache as clear_cache_impl
        clear_cache_impl()
        logger.info("Кэш успешно очищен")
        return True
    except Exception as e:
        logger.error(f"Ошибка при очистке кэша: {e}")
        return False


async def manual_clear_cache(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handler for /clear_cache command"""
    try:
        user_id = str(update.effective_user.id)
        if user_id not in ADMIN_IDS:
            await update.message.reply_text(
                f"❌ Очистка кэша доступна только администраторам.\n"
                f"Ваш ID: {user_id}"
            )
            return

        # Очищаем кэш
        if clear_cache():
            await update.message.reply_text("✅ Кэш успешно очищен")
        else:
            await update.message.reply_text("❌ Ошибка при очистке кэша")

        # Запускаем синхронизацию в отдельном потоке для загрузки новых файлов
        sync_files_async(force_check=True)
        await update.message.reply_text("🔄 Запущена синхронизация файлов в фоновом режиме")

    except Exception as e:
        logger.error(f"Ошибка при ручной очистке кэша: {e}")
        await update.message.reply_text(f"❌ Произошла ошибка: {str(e)}")


async def get_my_id(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Команда для получения ID пользователя"""
    user_id = update.effective_user.id
    username = update.effective_user.username or "Неизвестный пользователь"
    first_name = update.effective_user.first_name or ""
    last_name = update.effective_user.last_name or ""
    full_name = f"{first_name} {last_name}".strip()
    
    logger.info(f"Пользователь {username} (ID: {user_id}) запросил свой ID")
    
    message = f"👤 Информация о пользователе:\n\n"
    message += f"ID: {user_id}\n"
    if username:
        message += f"Username: @{username}\n"
    if full_name:
        message += f"Имя: {full_name}\n"
    
    message += "\nЧтобы получить права администратора, обратитесь к разработчику бота."
    
    await update.message.reply_text(message)
async def get_classroom_schedule(classroom: str, date_str: str) -> str:
    """Gets schedule for a specific classroom on a specific date"""
    try:
        # Normalize classroom format (ensure uppercase А)
        classroom = classroom.upper()
        # Replace Latin letters with Cyrillic equivalents
        latin_to_cyrillic = {
            'A': 'А', 'B': 'В', 'C': 'С', 'E': 'Е', 'H': 'Н',
            'K': 'К', 'M': 'М', 'O': 'О', 'P': 'Р', 'T': 'Т', 'X': 'Х'
        }
        for latin, cyrillic in latin_to_cyrillic.items():
            classroom = classroom.replace(latin, cyrillic)
        
        # Check cache first
        cached_result = get_cached_classroom_schedule(classroom, date_str)
        if cached_result:
            logger.info(f"Using cached schedule for classroom {classroom} on {date_str}")
            return cached_result
            
        # Use a semaphore to limit concurrent file access
        async with file_access_semaphore:
            # Log the files in the directory for debugging
            all_files = await run_blocking(lambda: os.listdir("downloaded_files"))
            logger.info(f"Files in downloaded_files directory: {all_files}")
            
            schedule_files = await run_blocking(lambda: [
                f for f in os.listdir("downloaded_files")
                if f.endswith('.xlsx')  
            ])
        
        logger.info(f"Found {len(schedule_files)} schedule files: {schedule_files}")
        
        if not schedule_files:
            result = f"❌ Ошибка: Не найдены файлы расписания групп в папке downloaded_files."
            cache_classroom_schedule(classroom, date_str, result)
            return result
            
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        weekday = date_obj.weekday()
        weekday_ru = days_ru[weekday]
        week_type = get_week_type(date_str)
        
        # Initialize the result - use a dictionary with lesson number as key
        classroom_schedule = {}
        
        # Process each schedule file with limited concurrency
        async def process_file(file):
            try:
                file_path = os.path.join("downloaded_files", file)
                
                # Check if file exists
                if not await run_blocking(lambda: os.path.exists(file_path)):
                    logger.error(f"File not found: {file_path}")
                    return []
                    
                # Extract group name from filename
                group_name = file.replace('.xlsx', '')
                
                # Parse the schedule for this group
                results = []
                async with file_access_semaphore:
                    wb = await run_blocking(lambda: openpyxl.load_workbook(file_path))
                    sheet = wb.active
                    
                    # Find the column for the current day
                    day_col, current_start_row = await run_blocking(
                        find_day_column, sheet, weekday_ru, 3, week_type == 'четная'
                    )
                
                if not day_col:
                    return []
                
                # Process each lesson
                current_row = current_start_row + 1
                while current_row < current_start_row + 20:
                    lesson_num = sheet.cell(row=current_row, column=day_col).value
                    if not lesson_num:
                        break
                        
                    # Check both subgroups for the classroom
                    room_first = str(sheet.cell(row=current_row + 1, column=day_col + 2).value or '')
                    room_second = str(sheet.cell(row=current_row + 1, column=day_col + 4).value or '')
                    
                    # Get subject and teacher information for both subgroups
                    subject_first = sheet.cell(row=current_row, column=day_col + 1).value
                    subject_second = sheet.cell(row=current_row, column=day_col + 3).value
                    teacher_first = sheet.cell(row=current_row + 1, column=day_col + 1).value
                    teacher_second = sheet.cell(row=current_row + 1, column=day_col + 3).value
                    
                    # Check if this is a course project that might use a common classroom
                    is_kp_lesson = (subject_first and '(кп)' in str(subject_first).lower()) or (subject_second and '(кп)' in str(subject_second).lower())
                    
                    # For KP lessons, we need to check if the classroom might be in either of the room cells
                    if is_kp_lesson and (classroom == room_first.strip() or classroom == room_second.strip()):
                        # Use the first subject if it's a KP, otherwise use the second
                        kp_subject = subject_first if subject_first and '(кп)' in str(subject_first).lower() else subject_second
                        kp_teacher = teacher_first if subject_first and '(кп)' in str(subject_first).lower() else teacher_second
                        
                        # For KP, don't assign a specific subgroup as they're often for the whole group
                        lesson_key = f"{lesson_num}_{group_name}"
                        results.append({
                            'lesson_num': lesson_num,
                            'lesson_key': lesson_key,
                            'subject': kp_subject,
                            'teacher': kp_teacher,
                            'group': group_name,
                            'subgroup': None  # No subgroup for KP
                        })
                    
                    # Only add a subgroup to the classroom schedule if the classroom exactly matches
                    # This fixes the issue where both subgroups were incorrectly added to the same classroom
                    if classroom == room_first.strip():
                        if not is_kp_lesson or '(кп)' not in str(subject_first).lower():  # Skip if already added as KP
                            lesson_key = f"{lesson_num}_1_{group_name}" if not is_theory_lesson(subject_first) else f"{lesson_num}_{group_name}"
                            results.append({
                                'lesson_num': lesson_num,
                                'lesson_key': lesson_key,
                                'subject': subject_first,
                                'teacher': teacher_first,
                                'group': group_name,
                                'subgroup': 1 if not is_theory_lesson(subject_first) else None
                            })
                    
                    if classroom == room_second.strip():
                        if not is_kp_lesson or '(кп)' not in str(subject_second).lower():  # Skip if already added as KP
                            lesson_key = f"{lesson_num}_2_{group_name}" if not is_theory_lesson(subject_second) and subject_second else f"{lesson_num}_{group_name}"
                            results.append({
                                'lesson_num': lesson_num,
                                'lesson_key': lesson_key,
                                'subject': subject_second if subject_second else subject_first,
                                'teacher': teacher_second if teacher_second else teacher_first,
                                'group': group_name,
                                'subgroup': 2 if not is_theory_lesson(subject_second) and subject_second else None
                            })
                    
                    current_row += 2
                return results
            except Exception as e:
                logger.error(f"Error processing file {file}: {e}")
                return []
        
        # Process files concurrently with a limit
        tasks = [process_file(file) for file in schedule_files]
        results = await asyncio.gather(*tasks)
        
        # Combine results into a dictionary with lesson_key as the key
        for file_results in results:
            for lesson in file_results:
                classroom_schedule[lesson['lesson_key']] = lesson
        
        # Check replacements for this date
# Check replacements for this date
                # Check replacements for this date
                # Check replacements for this date
                # Check replacements for this date
                # Check replacements for this date
                # Check replacements for this date
# Check replacements for this date
                # Check replacements for this date
                # Check replacements for this date
        replacement_file = await run_blocking(get_replacements_file, date_str)
        replacements_found = {}  # Initialize replacements_found dictionary
        replacements_by_lesson = {}  # Initialize replacements_by_lesson dictionary
        moved_lessons = set()  # Track lessons moved to different classrooms
        
        if replacement_file:
            async with file_access_semaphore:
                wb_replacements = await run_blocking(lambda: openpyxl.load_workbook(replacement_file))
                sheet_replacements = wb_replacements.active
                
                # Find all group columns
                group_cols = {}
                for col in range(4, sheet_replacements.max_column + 1):
                    group_cell = sheet_replacements.cell(row=2, column=col).value
                    if group_cell:
                        group_cols[col] = str(group_cell).strip()
                
                # First, identify lessons that are moved to different classrooms
                current_date = None
                for row in range(3, sheet_replacements.max_row + 1):
                    date_cell = sheet_replacements.cell(row=row, column=2).value
                    if date_cell:
                        try:
                            current_date = datetime.strptime(str(date_cell), '%d.%m.%Y').strftime('%d.%m.%Y')
                        except ValueError:
                            continue
                    
                    if current_date != date_str:
                        continue
                    
                    lesson_num = sheet_replacements.cell(row=row, column=3).value
                    if not lesson_num:
                        continue
                    
                    # Check each group for this classroom in original schedule
                    for orig_key, lesson in list(classroom_schedule.items()):
                        lesson_group = lesson['group']
                        lesson_number = str(lesson['lesson_num'])
                        
                        if str(lesson_num) == lesson_number:
                            # Check if this group has a replacement
                            for col, group in group_cols.items():
                                if group == lesson_group:
                                    replacement = sheet_replacements.cell(row=row, column=col).value
                                    if replacement and classroom not in str(replacement):
                                        # This lesson is moved to a different classroom
                                        moved_lessons.add(orig_key)
                                    elif replacement and classroom in str(replacement):
                                        # This is a replacement in the same classroom
                                        replacement_key = (str(lesson_num), lesson_group)
                                        if replacement_key not in replacements_by_lesson:
                                            replacements_by_lesson[replacement_key] = []
                                        replacements_by_lesson[replacement_key].append((col, replacement))
                
                # Remove lessons that have been moved to different classrooms
                for key in moved_lessons:
                    if key in classroom_schedule:
                        classroom_schedule.pop(key, None)
                
                # Now collect all replacements that are in this classroom
                for row in range(3, sheet_replacements.max_row + 1):
                    date_cell = sheet_replacements.cell(row=row, column=2).value
                    if date_cell:
                        try:
                            current_date = datetime.strptime(str(date_cell), '%d.%m.%Y').strftime('%d.%m.%Y')
                        except ValueError:
                            continue
                    
                    if current_date != date_str:
                        continue
                    
                    lesson_num = sheet_replacements.cell(row=row, column=3).value
                    if not lesson_num:
                        continue
                    
                    # Check each group for replacements in this classroom
                    for col, group in group_cols.items():
                        replacement = sheet_replacements.cell(row=row, column=col).value
                        if replacement and classroom in str(replacement):
                            # This is a replacement in this classroom
                            key = (str(lesson_num), group)
                            if key not in replacements_by_lesson:
                                replacements_by_lesson[key] = []
                            replacements_by_lesson[key].append((col, replacement))
                
                # Process all the replacements
                for (lesson_num, group), replacements in replacements_by_lesson.items():
                    for col, replacement in replacements:
                        replacement_text = str(replacement).strip()
                        
                        # Handle multi-line replacement text
                        lines = replacement_text.split('\n')
                        if len(lines) > 1:
                            # If we have multiple lines, the teacher is probably on the last line
                            subject_line = '\n'.join(lines[:-1]).strip()
                            teacher_line = lines[-1].strip()
                            
                            # Check if the last line looks like a teacher name
                            if re.match(r'^[А-Яа-я]+\s+[А-Я]\.[А-Я]\.$', teacher_line):
                                teacher = teacher_line
                                replacement_text = subject_line  # Use only the subject lines for further processing
                            else:
                                # If the last line doesn't look like a teacher, keep everything
                                replacement_text = '\n'.join(lines)
                        
                        # Extract teacher name with better pattern matching
                        # Try different teacher name patterns in order of specificity
                        teacher_pattern1 = r'([А-Яа-я]+)\s+([А-Я])\.\s*([А-Я])\.'  # Standard format: Фамилия И.О.
                        teacher_pattern2 = r'([А-Яа-я]+)\s+([А-Я])[а-я]+\s+([А-Я])[а-я]+'  # Full name format: Фамилия Имя Отчество
                        teacher_pattern3 = r'([А-Яа-я]+)\s+([А-Я])\.([А-Я])\.'  # No space: Фамилия И.О.
                        
                        teacher_match = None
                        if 'teacher' not in locals() or teacher == "Неизвестно":
                            teacher = "Неизвестно"
                            
                            # Try to find teacher name with first pattern
                            teacher_match = re.search(teacher_pattern1, replacement_text)
                            if teacher_match:
                                teacher = teacher_match.group(0)
                            else:
                                # Try second pattern
                                teacher_match = re.search(teacher_pattern2, replacement_text)
                                if teacher_match:
                                    # Extract full name and format as "Фамилия И.О."
                                    last_name = teacher_match.group(1)
                                    first_initial = teacher_match.group(2)
                                    middle_initial = teacher_match.group(3)
                                    teacher = f"{last_name} {first_initial}.{middle_initial}."
                                else:
                                    # Try third pattern
                                    teacher_match = re.search(teacher_pattern3, replacement_text)
                                    if teacher_match:
                                        teacher = teacher_match.group(0)
                        else:
                            # We already found a teacher from multi-line processing
                            teacher_match = re.match(r'^([А-Яа-я]+\s+[А-Я]\.[А-Я]\.)$', teacher)
                        
                        # Determine subgroup
                        subgroup_info = None
                        # Get the positions of "1." and "2." in the text
                        pos_1 = replacement_text.find('1. ')
                        pos_2 = replacement_text.find('2. ')
                        
                        # If both "1." and "2." are in the text, determine which part contains the classroom
                        if pos_1 != -1 and pos_2 != -1 and pos_1 < pos_2:
                            # Split the text into two parts for analysis
                            subgroup1_text = replacement_text[pos_1+3:pos_2].strip()
                            subgroup2_text = replacement_text[pos_2+3:].strip()
                            
                            # Check if the classroom is mentioned in either part
                            # For classroom schedule, we only care about the part that refers to this classroom
                            if classroom in subgroup1_text:
                                subgroup_info = 1
                                # Use only the first subgroup's text for further processing
                                replacement_text = subgroup1_text
                                logger.info(f"Found classroom {classroom} in first subgroup part of replacement: {subgroup1_text}")
                            elif classroom in subgroup2_text:
                                subgroup_info = 2
                                # Use only the second subgroup's text for further processing
                                replacement_text = subgroup2_text
                                logger.info(f"Found classroom {classroom} in second subgroup part of replacement: {subgroup2_text}")
                            else:
                                # If classroom isn't explicitly mentioned, use standard pattern detection
                                if '1. ' in replacement_text or '1 п/г' in replacement_text or '1 подгр' in replacement_text:
                                    subgroup_info = 1
                                elif '2. ' in replacement_text or '2 п/г' in replacement_text or '2 подгр' in replacement_text:
                                    subgroup_info = 2
                        else:
                            # Use standard pattern detection for simple cases
                            if '1. ' in replacement_text or '1 п/г' in replacement_text or '1 подгр' in replacement_text:
                                subgroup_info = 1
                            elif '2. ' in replacement_text or '2 п/г' in replacement_text or '2 подгр' in replacement_text:
                                subgroup_info = 2
                        
                        # First, try to find the classroom to remove it from subject text
                        classroom_pattern = re.compile(r'\b[А-Я]\d{3,}\b')
                        classroom_matches = classroom_pattern.findall(replacement_text)
                        clean_text = replacement_text
                        for match in classroom_matches:
                            clean_text = clean_text.replace(match, '')
                        
                        # Extract subject with improved pattern matching
                        subject_text = ""
                        # Try to find common lesson type patterns first
                        lesson_type_match = re.search(r'\([А-Яа-я]+\)', clean_text)
                        lesson_type = ""
                        if lesson_type_match:
                            lesson_type = lesson_type_match.group(0)
                        
                        # Look for specific patterns: МДК, УП, ПМ
                        if 'УП' in clean_text:
                            # Match practice lessons with their full names
                            up_match = re.search(r'УП\.\d+\.\d+(?:\.\d+)?(?:\s+[^{}]+)?', clean_text)
                            if up_match:
                                subject_start = up_match.start()
                                subject_end = len(clean_text)
                                if teacher_match:
                                    teacher_start = clean_text.find(teacher_match.group(0))
                                    if teacher_start > 0:
                                        subject_end = teacher_start
                                
                                subject_text = clean_text[subject_start:subject_end].strip()
                                if lesson_type and lesson_type not in subject_text:
                                    subject_text = f"{lesson_type} {subject_text}"
                        elif 'МДК' in clean_text:
                            # Match MDK lessons with their full names, including section numbers and titles
                            mdk_match = re.search(r'МДК\.\d+\.\d+(?:\.\d+)?(?:[\s\S]+?)?', clean_text)
                            if mdk_match:
                                # Get the full text after the MDK pattern until we hit the teacher name or end
                                mdk_start = mdk_match.start()
                                mdk_end = len(clean_text)
                                if teacher_match:
                                    teacher_start = clean_text.find(teacher_match.group(0))
                                    if teacher_start > 0:
                                        mdk_end = teacher_start
                                
                                subject_text = clean_text[mdk_start:mdk_end].strip()
                                if lesson_type and lesson_type not in subject_text:
                                    subject_text = f"{lesson_type} {subject_text}"
                        elif 'ПМ' in clean_text:
                            # Match professional modules
                            pm_match = re.search(r'ПМ\.\d+(?:\s+[^{}]+)?', clean_text)
                            if pm_match:
                                subject_start = pm_match.start()
                                subject_end = len(clean_text)
                                if teacher_match:
                                    teacher_start = clean_text.find(teacher_match.group(0))
                                    if teacher_start > 0:
                                        subject_end = teacher_start
                                
                                subject_text = clean_text[subject_start:subject_end].strip()
                                if lesson_type and lesson_type not in subject_text:
                                    subject_text = f"{lesson_type} {subject_text}"
                        else:
                            # For other types of subjects, take everything before the teacher name
                            subject_text = clean_text
                            if teacher_match:
                                teacher_text = teacher_match.group(0)
                                if teacher_text in subject_text:
                                    subject_text = subject_text[:subject_text.find(teacher_text)].strip()
                            
                            # Remove classroom numbers and other non-subject patterns
                            subject_text = re.sub(r'\b[А-Я]\d{2,}\b', '', subject_text)
                            # Remove any non-allowed characters but keep periods, spaces, etc.
                            subject_text = re.sub(r'[^А-Яа-яA-Za-z\s\.\-\(\)0-9]+', ' ', subject_text)
                            
                            # If we have a lesson type but it's not in the subject text, add it
                            if lesson_type and lesson_type not in subject_text:
                                subject_text = f"{lesson_type} {subject_text}"
                        
                        # Final cleanup of the subject text
                        subject_text = re.sub(r'\s+', ' ', subject_text)  # Replace multiple spaces with a single space
                        subject_text = subject_text.strip()
                        
                        # Create unique key for replacement
                        lesson_key = f"{lesson_num}_{subgroup_info}_{group}" if subgroup_info else f"{lesson_num}_{group}"
                        
                        # Add to replacements_found dictionary
                        replacements_found[lesson_key] = {
                            'lesson_num': lesson_num,
                            'lesson_key': lesson_key,
                            'subject': subject_text,
                            'teacher': teacher,
                            'group': group,
                            'subgroup': subgroup_info,
                            'is_replacement': True
                        }
                
                # Add all replacements to the schedule
                classroom_schedule.update(replacements_found)
                
                # For classroom schedules, prioritize replacements over originals
                # First, categorize lessons by lesson number
                replacements_by_lesson_num = {}
                original_by_lesson_num = {}
                
                # Find all replacements and originals by lesson number
                for key, lesson in list(classroom_schedule.items()):
                    lesson_num = str(lesson['lesson_num'])
                    
                    if lesson.get('is_replacement'):
                        if lesson_num not in replacements_by_lesson_num:
                            replacements_by_lesson_num[lesson_num] = []
                        replacements_by_lesson_num[lesson_num].append(key)
                    else:
                        if lesson_num not in original_by_lesson_num:
                            original_by_lesson_num[lesson_num] = []
                        original_by_lesson_num[lesson_num].append(key)
                
                # For each lesson number that has replacements, remove all original lessons
                for lesson_num, replacement_keys in replacements_by_lesson_num.items():
                    if lesson_num in original_by_lesson_num:
                        for orig_key in original_by_lesson_num[lesson_num]:
                            logger.info(f"Removing original lesson for classroom {classroom} at period {lesson_num} in favor of replacement")
                            if orig_key in classroom_schedule:
                                classroom_schedule.pop(orig_key)
        
        # Format the result
        if not classroom_schedule:
            result = f"📅 Расписание для кабинета {classroom} на {weekday_ru} {date_str} ({week_type} неделя):\n\nВ этот день занятий в кабинете нет."
            logger.info(f"Caching empty schedule for classroom {classroom} on {date_str}")
            cache_classroom_schedule(classroom, date_str, result)
            return result
            
        
                
                # Format the schedule
        formatted = [f"📅 Расписание для кабинета {classroom} на {weekday_ru} {date_str} ({week_type} неделя):"]
        
        # Convert to list and sort by lesson number
        schedule_list = list(classroom_schedule.values())
        
        # Group lessons by lesson number for better display
        lessons_by_number = {}
        for lesson in schedule_list:
            lesson_num = str(lesson['lesson_num'])
            if lesson_num not in lessons_by_number:
                lessons_by_number[lesson_num] = []
            lessons_by_number[lesson_num].append(lesson)
        
        # Sort by lesson number
        for lesson_num in sorted(lessons_by_number.keys(), key=lambda x: int(x.split('.')[0]) if isinstance(x, str) and x.split('.')[0].isdigit() else 0):
            lessons = lessons_by_number[lesson_num]
            
            # Sort lessons - replacements first
            lessons.sort(key=lambda x: 0 if x.get('is_replacement') else 1)
            
            for lesson in lessons:
                lesson_str = f"\n{str(lesson['lesson_num'])}️⃣ "
                
                # Add replacement marker
                if lesson.get('is_replacement'):
                    lesson_str += "✏️ "
                
                lesson_str += f"{lesson.get('subject', 'Неизвестно')} "
                lesson_str += f"🎓{lesson.get('teacher', 'Неизвестно')} "
                
                # Format group and subgroup information
                subgroup_text = ""
                if lesson.get('subgroup'):
                    subgroup_text = f", {lesson['subgroup']}-я подгруппа"
                
                lesson_str += f"👥 [{lesson.get('group', 'Неизвестно')}{subgroup_text}]"
                
                formatted.append(lesson_str)
        
        result = "\n".join(formatted)
        
        # Cache the result before returning
        logger.info(f"Caching schedule for classroom {classroom} on {date_str}")
        cache_classroom_schedule(classroom, date_str, result)
        
        return result
        
    except Exception as e:
        logger.error(f"Error getting classroom schedule: {e}")
        return f"Ошибка при получении расписания кабинета: {str(e)}"


async def classroom_schedule_date_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle classroom schedule date selection"""
    try:
        query = update.callback_query
        await query.answer()
        
        # Get the selected date
        date_str = query.data.replace('classroom_date_', '')
        
        # Get classroom from user data
        classroom = context.user_data.get('classroom', '')
        
        if not classroom:
            await query.message.reply_text("Ошибка: не указан кабинет")
            return ConversationHandler.END
        
        # Send searching message
        await query.message.reply_text(f"🔍 Ищу расписание для кабинета {classroom} на {date_str}...")
        
        # Get classroom schedule directly (not through run_blocking)
        schedule = await get_classroom_schedule(classroom, date_str)
        
        # Send the schedule
        await query.message.reply_text(schedule)
        
        return ConversationHandler.END
        
    except Exception as e:
        logger.error(f"Error in classroom schedule date handler: {e}")
        await update.callback_query.message.reply_text("Произошла ошибка при получении расписания кабинета")
        return ConversationHandler.END

def format_classroom_schedule(schedule: dict, classroom: str, date_str: str) -> str:
    """Formats classroom schedule for display"""
    try:
        if not schedule:
            return f"Расписание для кабинета {classroom} на {date_str} не найдено"
        
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        weekday = days_ru[date_obj.weekday()]
        week_type = get_week_type(date_str)
        
        formatted = [f"📅 Расписание для кабинета {classroom} на {weekday} {date_str} ({week_type} неделя):"]
        
        # Group lessons by lesson number
        lessons_by_number = {}
        for key, lesson in schedule.items():
            lesson_num = str(lesson['lesson_num'])
            if lesson_num not in lessons_by_number:
                lessons_by_number[lesson_num] = []
            lessons_by_number[lesson_num].append(lesson)
        
        # Sort by lesson number
        for lesson_num in sorted(lessons_by_number.keys(), key=lambda x: int(x.split('.')[0]) if isinstance(x, str) and x.split('.')[0].isdigit() else 0):
            lessons = lessons_by_number[lesson_num]
            
            # Sort lessons - replacements first
            lessons.sort(key=lambda x: 0 if x.get('is_replacement') else 1)
            
            for lesson in lessons:
                lesson_str = f"\n{str(lesson['lesson_num'])}️⃣ "
                
                # Add replacement marker
                if lesson.get('is_replacement'):
                    lesson_str += "✏️ "
                
                lesson_str += f"{lesson.get('subject', 'Неизвестно')} "
                lesson_str += f"🎓{lesson.get('teacher', 'Неизвестно')} "
                
                # Format group and subgroup information
                subgroup_text = ""
                if lesson.get('subgroup'):
                    subgroup_text = f", {lesson['subgroup']}-я подгруппа"
                
                lesson_str += f"👥 [{lesson.get('group', 'Неизвестно')}{subgroup_text}]"
                
                formatted.append(lesson_str)
        
        return "\n".join(formatted)
    except Exception as e:
        logger.error(f"Error formatting classroom schedule: {e}")
        return "Ошибка при форматировании расписания"
@non_blocking_handler
async def classroom_schedule_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handler for /classroom command"""
    try:
        # Проверяем, не идет ли процесс обновления файлов
        if is_update_in_progress():
            status_message = get_update_status_message()
            await update.message.reply_text(
                status_message,
                parse_mode='Markdown'
            )
            return
            
        # Check if command has correct format
        args = context.args
        if len(args) < 2:
            await update.message.reply_text(
                "Пожалуйста, укажите кабинет и дату в формате:\n"
                "/classroom А403 02.05.2023\n"
                "Примеры кабинетов: А403, В12, К25"
            )
            return
        
        classroom = args[0].upper()
        date_str = args[1]
        
        # Validate date format
        try:
            datetime.strptime(date_str, '%d.%m.%Y')
        except ValueError:
            await update.message.reply_text("Пожалуйста, укажите дату в формате ДД.ММ.ГГГГ")
            return
        
        # Use run_blocking to get schedule in a separate thread
        schedule = await run_blocking(get_classroom_schedule, classroom, date_str)
        formatted_schedule = await run_blocking(format_classroom_schedule, schedule, classroom, date_str)
        
        await update.message.reply_text(formatted_schedule)
        
    except Exception as e:
        logger.error(f"Error in classroom schedule command: {e}")
        await update.message.reply_text("Произошла ошибка при получении расписания кабинета")

def main() -> None:
    """Запуск бота"""
    try:
        logger.info("Инициализация бота...")
        # Инициализируем кэш
        init_cache()
        
        # Обрабатываем сигналы для корректного завершения
        signal.signal(signal.SIGINT, signal_handler)
        signal.signal(signal.SIGTERM, signal_handler)
        
        # Create the Application and pass it your bot's token.
        application = Application.builder().token(TELEGRAM_TOKEN).build()
        
        # Запускаем синхронизацию замен в отдельном потоке
        sync_thread = threading.Thread(target=schedule_sync)
        sync_thread.daemon = True  # Потом завершится вместе с основным
        sync_thread.start()
        logger.info("Запущен поток синхронизации файлов")
        
        # Установить команды бота
        loop = asyncio.get_event_loop()
        loop.run_until_complete(set_commands(application))
        
        # Add handlers for commands
        application.add_handler(CommandHandler("start", start))
        application.add_handler(CommandHandler("subscribe", subscribe_command))
        application.add_handler(CommandHandler("classroom", classroom_schedule_command))
        application.add_handler(CommandHandler("unsubscribe", unsubscribe_command))
        
        # Add conversation handler for main dialogs
        conv_handler = ConversationHandler(
            entry_points=[
                CommandHandler('start', start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_all_messages),
            ],
            states={
                CHOOSE_ACTION: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, choose_action),
                    MessageHandler(filters.Regex('^Отмена$'), cancel)
                ],
                ENTER_CLASSROOM: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, enter_classroom),
                    MessageHandler(filters.Regex('^Отмена$'), cancel)
                ],
                CHOOSE_DATE_FOR_CLASSROOM: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, choose_date_for_classroom),
                    MessageHandler(filters.Regex('^Отмена$'), cancel)
                ],
                ENTER_TEACHER: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, enter_teacher),
                    MessageHandler(filters.Regex('^Отмена$'), cancel)
                ],
                CHOOSE_DATE_FOR_TEACHER: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, choose_date_for_teacher),
                    MessageHandler(filters.Regex('^Отмена$'), cancel)
                ],
                ENTER_GROUP: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, group_input),
                    MessageHandler(filters.Regex('^Отмена$'), cancel)
                ],             
                CHOOSE_SUBGROUP: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, subgroup_choice),
                    MessageHandler(filters.Regex('^Отмена$'), cancel)
                ],
            },
            fallbacks=[CommandHandler('cancel', cancel)],
        )
        
        application.add_handler(conv_handler)
        application.add_handler(CommandHandler("clear_cache", manual_clear_cache))
        application.add_handler(CommandHandler("classroom", classroom_schedule_command))
        application.add_handler(CommandHandler("myid", get_my_id))
        
        # Добавляем обработчик ошибок
        application.add_error_handler(error_handler)
        
        # Запускаем первичную синхронизацию в фоновом режиме
        sync_files_async(force_check=False)
        
        # Run the bot until the user presses Ctrl-C
        logger.info("Бот запущен и готов к работе")
        application.run_polling()
        
    except Exception as e:
        logger.critical(f"Критическая ошибка при запуске бота: {e}")
        import traceback
        logger.critical(traceback.format_exc())


if __name__ == '__main__':
    main()


